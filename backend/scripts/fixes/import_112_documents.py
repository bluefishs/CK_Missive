# -*- coding: utf-8 -*-
"""
一次性腳本：匯入 112 收發文整理匯入範本.xlsx
- 新公文字號：INSERT
- 已存在公文字號：UPDATE（刷新欄位 + 承攬案件連結）
- Excel 內重複字號：以最後一筆為準

Usage:
    cd backend
    python scripts/fixes/import_112_documents.py [--dry-run]
"""
import asyncio
import logging
import sys
import os
from pathlib import Path
from datetime import datetime
from collections import OrderedDict
from typing import Dict, Any, Optional

# 加入 backend 到 sys.path
backend_dir = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(backend_dir))

# 環境設定
os.environ.setdefault("DOTENV_PATH", str(backend_dir.parent / ".env"))

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
logger = logging.getLogger(__name__)

# Excel 路徑
EXCEL_PATH = backend_dir.parent / "#BUG" / "112收發文整理匯入範本.xlsx"
SHEET_NAME = "公文系統匯入樣版"

# 欄位對應
FIELD_MAPPING = {
    '發文形式': 'delivery_method',
    '類別': 'category',
    '公文類型': 'doc_type',
    '公文字號': 'doc_number',
    '主旨': 'subject',
    '說明': 'content',
    '公文日期': 'doc_date',
    '收文日期': 'receive_date',
    '發文日期': 'send_date',
    '發文單位': 'sender',
    '受文單位': 'receiver',
    '備註': 'notes',
    '簡要說明(乾坤備註)': 'ck_note',
    '狀態': 'status',
    '承攬案件': 'contract_project_name',
}


def clean_string(value: Any) -> Optional[str]:
    """清理字串值"""
    if value is None:
        return None
    s = str(value).strip()
    if not s or s.lower() in ('none', 'nan', 'null'):
        return None
    return s


def parse_date(value: Any) -> Optional[Any]:
    """解析日期"""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    from datetime import date
    if isinstance(value, date):
        return value
    s = str(value).strip()
    if not s:
        return None
    for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def read_excel() -> list[Dict[str, Any]]:
    """讀取 Excel，回傳 row_data 列表（以 doc_number 去重，保留最後一筆）"""
    import openpyxl

    logger.info(f"讀取 Excel: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(str(EXCEL_PATH), read_only=True, data_only=True)

    # 嘗試指定工作表名稱
    if SHEET_NAME in wb.sheetnames:
        ws = wb[SHEET_NAME]
    else:
        ws = wb.active
        logger.warning(f"找不到工作表 '{SHEET_NAME}'，使用 active sheet")

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    logger.info(f"欄位: {[h for h in headers if h]}")

    # 使用 OrderedDict 去重（同一 doc_number 保留最後一筆）
    unique_rows: OrderedDict[str, Dict[str, Any]] = OrderedDict()
    total_rows = 0
    skipped_empty = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        total_rows += 1
        row_data = {}
        for i, header in enumerate(headers):
            if header and i < len(row):
                row_data[header] = row[i]

        doc_number = clean_string(row_data.get('公文字號'))
        if not doc_number:
            skipped_empty += 1
            continue

        unique_rows[doc_number] = row_data

    wb.close()

    dedup_count = total_rows - skipped_empty - len(unique_rows)
    logger.info(
        f"Excel 讀取完成: {total_rows} 列, "
        f"空白跳過={skipped_empty}, 去重={dedup_count}, "
        f"有效={len(unique_rows)}"
    )

    return list(unique_rows.values())


async def run_import(dry_run: bool = False):
    """執行匯入"""
    from dotenv import load_dotenv
    dotenv_path = os.environ.get("DOTENV_PATH")
    if dotenv_path and Path(dotenv_path).exists():
        load_dotenv(dotenv_path, override=True)

    # 讀取 Excel
    rows = read_excel()
    if not rows:
        logger.error("沒有有效資料")
        return

    # 建立 DB 連線
    from app.db.database import AsyncSessionLocal, engine
    from sqlalchemy import select, func
    import re

    async with AsyncSessionLocal() as db:
        # 匯入所需的服務
        from app.extended.models import OfficialDocument, ContractProject
        from app.services.strategies.agency_matcher import AgencyMatcher, ProjectMatcher
        from app.services.receiver_normalizer import (
            normalize_unit, cc_list_to_json, infer_agency_from_doc_number,
        )
        from app.services.base.validators import DocumentValidators

        agency_matcher = AgencyMatcher(db)
        project_matcher = ProjectMatcher(db)

        # 統計
        stats = {
            'inserted': 0,
            'updated': 0,
            'skipped': 0,
            'errors': 0,
            'project_linked': 0,
            'project_created': 0,
        }
        error_details = []
        update_details = []
        insert_details = []

        # 流水號計數器
        serial_counters = {'R': 0, 'S': 0}

        async def get_next_serial(category: str) -> str:
            prefix = 'S' if category == '發文' else 'R'
            if serial_counters[prefix] == 0:
                query = select(func.max(OfficialDocument.auto_serial)).where(
                    OfficialDocument.auto_serial.like(f'{prefix}%')
                )
                result = await db.execute(query)
                max_serial = result.scalar()
                if max_serial:
                    num_match = re.search(r'\d+', max_serial)
                    if num_match:
                        serial_counters[prefix] = int(num_match.group())
            serial_counters[prefix] += 1
            return f"{prefix}{serial_counters[prefix]:04d}"

        # 預先載入所有 doc_number → ID 映射
        query = select(OfficialDocument.id, OfficialDocument.doc_number).where(
            OfficialDocument.doc_number.isnot(None)
        )
        result = await db.execute(query)
        existing_map = {row[1]: row[0] for row in result.fetchall()}
        logger.info(f"資料庫現有公文: {len(existing_map)} 筆")

        # 預先載入承攬案件名稱 → ID
        query = select(ContractProject.id, ContractProject.project_name)
        result = await db.execute(query)
        project_map = {row[1]: row[0] for row in result.fetchall()}
        logger.info(f"資料庫現有承攬案件: {len(project_map)} 筆")

        for i, row_data in enumerate(rows):
            row_num = i + 2  # Excel 列號 (1-indexed header)
            doc_number = clean_string(row_data.get('公文字號'))

            try:
                # 驗證必填
                category = clean_string(row_data.get('類別'))
                subject = clean_string(row_data.get('主旨'))
                if not doc_number or not subject or not category:
                    stats['skipped'] += 1
                    continue

                if category not in ('收文', '發文'):
                    stats['skipped'] += 1
                    error_details.append(f"列{row_num}: 無效類別 '{category}'")
                    continue

                # 驗證公文類型
                doc_type = clean_string(row_data.get('公文類型')) or '函'
                if doc_type not in DocumentValidators.VALID_DOC_TYPES:
                    doc_type = '函'

                # 準備欄位資料
                sender_name = clean_string(row_data.get('發文單位'))
                receiver_name = clean_string(row_data.get('受文單位'))
                contract_name = clean_string(row_data.get('承攬案件'))

                # 智慧匹配機關
                sender_agency_id = await agency_matcher.match_or_create(sender_name) if sender_name else None
                receiver_agency_id = await agency_matcher.match_or_create(receiver_name) if receiver_name else None

                # 智慧匹配承攬案件
                contract_project_id = None
                if contract_name:
                    contract_project_id = await project_matcher.match_or_create(contract_name)
                    if contract_project_id:
                        stats['project_linked'] += 1
                        # 檢查是否為新建
                        if contract_name not in project_map:
                            stats['project_created'] += 1
                            project_map[contract_name] = contract_project_id

                # 正規化收發文單位
                s_norm = normalize_unit(sender_name) if sender_name else None
                r_norm = normalize_unit(receiver_name) if receiver_name else None

                # 根據公文字號前綴修正發文機關
                inferred_agency = infer_agency_from_doc_number(doc_number)
                if inferred_agency and s_norm and s_norm.primary != inferred_agency:
                    s_norm = normalize_unit(inferred_agency)
                    corrected_id = await agency_matcher.match_or_create(inferred_agency)
                    if corrected_id:
                        sender_agency_id = corrected_id

                doc_data = {
                    'category': category,
                    'doc_type': doc_type,
                    'doc_number': doc_number,
                    'subject': subject,
                    'content': clean_string(row_data.get('說明')),
                    'sender': sender_name,
                    'receiver': receiver_name,
                    'normalized_sender': s_norm.primary if s_norm else None,
                    'normalized_receiver': r_norm.primary if r_norm else None,
                    'cc_receivers': cc_list_to_json(r_norm.cc_list) if r_norm else None,
                    'sender_agency_id': sender_agency_id,
                    'receiver_agency_id': receiver_agency_id,
                    'contract_project_id': contract_project_id,
                    'delivery_method': clean_string(row_data.get('發文形式')) or '紙本郵寄',
                    'notes': clean_string(row_data.get('備註')),
                    'ck_note': clean_string(row_data.get('簡要說明(乾坤備註)')),
                    'status': clean_string(row_data.get('狀態')) or 'active',
                    'doc_date': parse_date(row_data.get('公文日期')),
                    'receive_date': parse_date(row_data.get('收文日期')),
                    'send_date': parse_date(row_data.get('發文日期')),
                }

                if doc_number in existing_map:
                    # === UPDATE ===
                    existing_id = existing_map[doc_number]
                    query = select(OfficialDocument).where(OfficialDocument.id == existing_id)
                    result = await db.execute(query)
                    existing_doc = result.scalar_one_or_none()

                    if existing_doc:
                        changes = []
                        for key, value in doc_data.items():
                            if key == 'doc_number':
                                continue  # 不更新 doc_number 本身
                            old_value = getattr(existing_doc, key, None)
                            if value is not None and old_value != value:
                                setattr(existing_doc, key, value)
                                changes.append(key)
                        if changes:
                            existing_doc.updated_at = datetime.now()
                            stats['updated'] += 1
                            update_details.append(
                                f"ID={existing_id} [{doc_number}]: {', '.join(changes[:5])}"
                            )
                        else:
                            stats['skipped'] += 1
                    else:
                        stats['skipped'] += 1

                else:
                    # === INSERT ===
                    auto_serial = await get_next_serial(category)
                    doc_data['auto_serial'] = auto_serial
                    doc_data['created_at'] = datetime.now()
                    doc_data['updated_at'] = datetime.now()

                    new_doc = OfficialDocument(**doc_data)
                    db.add(new_doc)
                    await db.flush()

                    existing_map[doc_number] = new_doc.id
                    stats['inserted'] += 1
                    insert_details.append(
                        f"ID={new_doc.id} [{doc_number}] {auto_serial}"
                    )

            except Exception as e:
                stats['errors'] += 1
                error_details.append(f"列{row_num} [{doc_number}]: {e}")
                logger.error(f"列{row_num} 處理失敗: {e}")

            # 進度報告
            if (i + 1) % 100 == 0:
                logger.info(f"進度: {i + 1}/{len(rows)}")

        # 提交或回滾
        if dry_run:
            await db.rollback()
            logger.info("=== DRY RUN 模式，未實際寫入 ===")
        else:
            await db.commit()
            logger.info("=== 已提交到資料庫 ===")

    # 報告結果
    logger.info("=" * 60)
    logger.info("匯入結果摘要")
    logger.info("=" * 60)
    logger.info(f"  總計處理: {len(rows)} 筆")
    logger.info(f"  新增: {stats['inserted']} 筆")
    logger.info(f"  更新: {stats['updated']} 筆")
    logger.info(f"  跳過(無變更): {stats['skipped']} 筆")
    logger.info(f"  錯誤: {stats['errors']} 筆")
    logger.info(f"  承攬案件連結: {stats['project_linked']} 筆")
    logger.info(f"  新建承攬案件: {stats['project_created']} 筆")

    if update_details:
        logger.info(f"\n--- 更新明細 (前 30 筆) ---")
        for d in update_details[:30]:
            logger.info(f"  {d}")
        if len(update_details) > 30:
            logger.info(f"  ... 還有 {len(update_details) - 30} 筆")

    if insert_details:
        logger.info(f"\n--- 新增明細 (前 30 筆) ---")
        for d in insert_details[:30]:
            logger.info(f"  {d}")
        if len(insert_details) > 30:
            logger.info(f"  ... 還有 {len(insert_details) - 30} 筆")

    if error_details:
        logger.info(f"\n--- 錯誤明細 ---")
        for d in error_details:
            logger.info(f"  {d}")

    return stats


if __name__ == "__main__":
    dry_run = "--dry-run" in sys.argv
    if dry_run:
        logger.info("*** DRY RUN 模式 — 不會實際寫入資料庫 ***")
    asyncio.run(run_import(dry_run=dry_run))
