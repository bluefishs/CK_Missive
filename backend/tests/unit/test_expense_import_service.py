"""
ExpenseImportService 單元測試

測試 QR 解析、Excel 匯入、電子發票關聯。
"""
import pytest
from unittest.mock import AsyncMock, MagicMock, patch
from decimal import Decimal
from datetime import date

from app.services.expense_import_service import ExpenseImportService


@pytest.fixture
def mock_db():
    return AsyncMock()


@pytest.fixture
def service(mock_db):
    svc = ExpenseImportService.__new__(ExpenseImportService)
    svc.db = mock_db
    svc.repo = AsyncMock()
    return svc


class TestParseQRData:
    """QR Code 解析 (財政部規範)"""

    # 銷售額 0x00001388=5000, 總額 0x00001482=5250
    SAMPLE_QR = "AB123456781150615000100001388000014820000000012345678ABCDEFGHIJKLMNOPQRSTUVWX"

    def test_parse_basic(self, service):
        result = service.parse_qr_data(self.SAMPLE_QR)
        assert result["inv_num"] == "AB12345678"
        assert result["date"] == date(2026, 6, 15)
        assert result["amount"] == Decimal("5250")
        assert result["sales_amount"] == Decimal("5000")
        assert result["tax_amount"] == Decimal("250")
        assert result["seller_ban"] == "12345678"
        assert result["buyer_ban"] is None  # 00000000
        assert result["source"] == "qr_scan"

    def test_too_short_raises(self, service):
        with pytest.raises(ValueError, match="長度不足"):
            service.parse_qr_data("AB1234567811")

    def test_with_buyer_ban(self, service):
        qr = "CD987654321150615000100001388000014828765432112345678ABCDEFGHIJKLMNOPQRSTUVWX"
        result = service.parse_qr_data(qr)
        assert result["buyer_ban"] == "87654321"

    def test_zero_tax_invoice(self, service):
        # 銷售額 = 總額 = 0x00002710 = 10000
        qr = "EF123456781150101999900002710000027100000000099887766ABCDEFGHIJKLMNOPQRSTUVWX"
        result = service.parse_qr_data(qr)
        assert result["tax_amount"] == Decimal("0")


class TestGenerateImportTemplate:
    """匯入範本 Excel 生成"""

    def test_generates_valid_xlsx(self, service):
        data = service.generate_import_template()
        assert isinstance(data, bytes)
        assert len(data) > 0
        # XLSX magic bytes: PK (ZIP format)
        assert data[:2] == b'PK'

    def test_has_headers(self, service):
        from openpyxl import load_workbook
        import io
        data = service.generate_import_template()
        wb = load_workbook(io.BytesIO(data))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        assert "發票號碼" in headers
        assert "金額" in headers
        assert "案件代碼" in headers
        wb.close()


class TestAutoLinkEInvoice:
    """電子發票關聯"""

    @pytest.mark.asyncio
    async def test_already_synced(self, service):
        inv = MagicMock()
        inv.inv_num = "AB12345678"
        inv.synced_at = "2026-01-01"
        service.repo.get_by_id = AsyncMock(return_value=inv)

        result = await service.auto_link_einvoice(1)
        assert result["status"] == "already_synced"

    @pytest.mark.asyncio
    async def test_not_found(self, service):
        service.repo.get_by_id = AsyncMock(return_value=None)
        result = await service.auto_link_einvoice(999)
        assert result is None

    @pytest.mark.asyncio
    async def test_no_inv_num(self, service):
        inv = MagicMock()
        inv.inv_num = None
        service.repo.get_by_id = AsyncMock(return_value=inv)
        result = await service.auto_link_einvoice(1)
        assert result is None
