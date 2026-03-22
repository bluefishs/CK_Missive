"""財務報表匯出服務 — 單元測試

測試範圍:
- Excel 費用報銷匯出 (表頭、資料列、合計)
- Excel 帳本匯出 (收支分列、合計)
- 空資料匯出
"""
import pytest
from datetime import date
from decimal import Decimal
from unittest.mock import AsyncMock, MagicMock, patch
from io import BytesIO

from openpyxl import load_workbook

from app.services.finance_export_service import FinanceExportService


def _make_mock_invoice(**kwargs):
    """建立模擬 ExpenseInvoice"""
    defaults = {
        "id": 1, "inv_num": "AB12345678", "date": date(2026, 3, 21),
        "amount": Decimal("1050"), "tax_amount": Decimal("50"),
        "category": "交通費", "case_code": "P113-001",
        "source": "manual", "status": "pending", "notes": "測試",
        "currency": "TWD", "original_amount": None, "exchange_rate": None,
    }
    defaults.update(kwargs)
    obj = MagicMock()
    for k, v in defaults.items():
        setattr(obj, k, v)
    return obj


def _make_mock_ledger(**kwargs):
    """建立模擬 FinanceLedger"""
    defaults = {
        "id": 1, "entry_date": date(2026, 3, 21),
        "entry_type": "expense", "amount": Decimal("500"),
        "category": "交通費", "case_code": "P113-001",
        "source_type": "manual", "description": "計程車",
        "notes": "",
    }
    defaults.update(kwargs)
    obj = MagicMock()
    for k, v in defaults.items():
        setattr(obj, k, v)
    return obj


class TestExportExpenses:
    """費用報銷 Excel 匯出"""

    @pytest.fixture
    def service(self):
        db = AsyncMock()
        svc = FinanceExportService(db)
        svc.expense_repo = AsyncMock()
        svc.ledger_repo = AsyncMock()
        return svc

    @pytest.mark.asyncio
    async def test_export_with_data(self, service):
        """匯出含資料的 Excel"""
        invoices = [
            _make_mock_invoice(inv_num="AB11111111", amount=Decimal("1000"), tax_amount=Decimal("50")),
            _make_mock_invoice(inv_num="CD22222222", amount=Decimal("2000"), tax_amount=Decimal("100")),
        ]
        service.expense_repo.query = AsyncMock(return_value=(invoices, 2))

        result = await service.export_expenses()
        assert isinstance(result, bytes)
        assert len(result) > 0

        wb = load_workbook(BytesIO(result))
        ws = wb.active
        assert ws.title == "費用報銷明細"
        # 表頭在第 3 行
        assert ws.cell(row=3, column=1).value == "發票號碼"
        # 第一筆資料在第 4 行
        assert ws.cell(row=4, column=1).value == "AB11111111"
        assert ws.cell(row=5, column=1).value == "CD22222222"
        # 合計行
        assert ws.cell(row=6, column=1).value == "合計"
        assert ws.cell(row=6, column=3).value == 3000.0

    @pytest.mark.asyncio
    async def test_export_empty(self, service):
        """空資料也能匯出"""
        service.expense_repo.query = AsyncMock(return_value=([], 0))
        result = await service.export_expenses()
        assert isinstance(result, bytes)
        wb = load_workbook(BytesIO(result))
        ws = wb.active
        assert ws.cell(row=3, column=1).value == "發票號碼"

    @pytest.mark.asyncio
    async def test_export_with_filters(self, service):
        """帶篩選條件的匯出"""
        service.expense_repo.query = AsyncMock(return_value=([], 0))
        result = await service.export_expenses(
            date_from=date(2026, 1, 1),
            date_to=date(2026, 3, 31),
            case_code="P113-001",
        )
        wb = load_workbook(BytesIO(result))
        ws = wb.active
        title = ws.cell(row=1, column=1).value
        assert "P113-001" in title
        assert "2026-01-01" in title


class TestExportLedger:
    """帳本 Excel 匯出"""

    @pytest.fixture
    def service(self):
        db = AsyncMock()
        svc = FinanceExportService(db)
        svc.expense_repo = AsyncMock()
        svc.ledger_repo = AsyncMock()
        return svc

    @pytest.mark.asyncio
    async def test_export_with_data(self, service):
        """匯出含收支資料的 Excel"""
        ledgers = [
            _make_mock_ledger(entry_type="income", amount=Decimal("5000")),
            _make_mock_ledger(entry_type="expense", amount=Decimal("1200")),
        ]
        service.ledger_repo.query = AsyncMock(return_value=(ledgers, 2))

        result = await service.export_ledger()
        assert isinstance(result, bytes)

        wb = load_workbook(BytesIO(result))
        ws = wb.active
        assert ws.title == "帳本收支明細"
        assert ws.cell(row=3, column=1).value == "日期"
        # 收支合計
        row_idx = 2 + 1 + 2 + 1  # title + empty + header + 2 data + 1
        assert ws.cell(row=row_idx + 1, column=1).value == "總收入"
        assert ws.cell(row=row_idx + 1, column=3).value == 5000.0
        assert ws.cell(row=row_idx + 2, column=1).value == "總支出"
        assert ws.cell(row=row_idx + 2, column=3).value == 1200.0

    @pytest.mark.asyncio
    async def test_export_ledger_empty(self, service):
        """空帳本匯出"""
        service.ledger_repo.query = AsyncMock(return_value=([], 0))
        result = await service.export_ledger()
        assert isinstance(result, bytes)
