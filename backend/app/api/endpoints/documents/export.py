"""
公文匯出 API 端點

包含：CSV 匯出、Excel 匯出

@version 3.0.0
@date 2026-01-18
"""
import io
import csv
from datetime import datetime
from urllib.parse import quote
from fastapi import APIRouter, Body, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy import select, or_, and_, func
from sqlalchemy.orm import selectinload

from .common import (
    logger, Depends, AsyncSession, get_async_db,
    OfficialDocument, ContractProject,
    DocumentExportQuery, ExcelExportRequest,
)

router = APIRouter()


# ============================================================================
# 公文匯出 API
# ============================================================================

@router.post("/export", summary="匯出公文資料")
async def export_documents(
    query: DocumentExportQuery = Body(...),
    db: AsyncSession = Depends(get_async_db)
):
    """
    匯出公文資料為 CSV 格式

    支援功能:
    - 依指定 ID 列表匯出
    - 依類別/年度篩選後匯出
    - 若未指定條件則匯出全部
    """
    try:
        # 構建查詢
        doc_query = select(OfficialDocument).options(
            selectinload(OfficialDocument.contract_project)
        )

        # 篩選條件
        conditions = []
        if query.document_ids:
            conditions.append(OfficialDocument.id.in_(query.document_ids))
        if query.category:
            conditions.append(OfficialDocument.category == query.category)
        if query.year:
            conditions.append(func.extract('year', OfficialDocument.doc_date) == query.year)

        if conditions:
            doc_query = doc_query.where(and_(*conditions))

        doc_query = doc_query.order_by(OfficialDocument.doc_date.desc())

        result = await db.execute(doc_query)
        documents = result.scalars().all()

        # 產生 CSV
        output = io.StringIO()
        writer = csv.writer(output, quoting=csv.QUOTE_ALL)

        # 寫入標題列
        headers = [
            '序號', '公文文號', '主旨', '類別', '發文/收文日期',
            '發文單位', '受文單位', '承攬案件', '狀態', '備註'
        ]
        writer.writerow(headers)

        # 寫入資料列
        for idx, doc in enumerate(documents, start=1):
            contract_case_name = ""
            if doc.contract_project:
                contract_case_name = doc.contract_project.project_name or ""

            row = [
                doc.auto_serial or idx,
                doc.doc_number or "",
                doc.subject or "",
                doc.category or "",
                str(doc.doc_date) if doc.doc_date else "",
                doc.sender or "",
                doc.receiver or "",
                contract_case_name,
                doc.status or "",
                doc.notes or ""
            ]
            writer.writerow(row)

        # 重置游標位置
        output.seek(0)

        # 回傳 CSV 檔案
        filename = f"documents_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

        return StreamingResponse(
            iter(['\ufeff' + output.getvalue()]),  # BOM for Excel UTF-8
            media_type="text/csv; charset=utf-8",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )

    except Exception as e:
        logger.error(f"匯出公文失敗: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"匯出公文失敗: {str(e)}")


# ============================================================================
# Excel 匯出端點
# ============================================================================

@router.post("/export/excel", summary="匯出公文為 Excel")
async def export_documents_excel(
    request: ExcelExportRequest = Body(default=ExcelExportRequest()),
    db: AsyncSession = Depends(get_async_db)
):
    """
    匯出公文資料為 Excel 格式 (.xlsx)

    檔名格式: CK公文YYYYMMDD.xlsx

    支援功能:
    - 依指定 ID 列表匯出
    - 依類別/年度/關鍵字/狀態篩選後匯出
    - 若未指定條件則匯出全部（無筆數限制）

    流水號說明:
    - S 開頭: 發文 (Send)
    - R 開頭: 收文 (Receive)
    """
    try:
        import pandas as pd
        from io import BytesIO

        # 構建查詢 - 無筆數限制，匯出全部符合條件的資料
        doc_query = select(OfficialDocument).options(
            selectinload(OfficialDocument.contract_project),
            selectinload(OfficialDocument.sender_agency),
            selectinload(OfficialDocument.receiver_agency),
            selectinload(OfficialDocument.attachments)  # 載入附件以統計數量
        )

        # 篩選條件
        conditions = []
        if request.document_ids:
            conditions.append(OfficialDocument.id.in_(request.document_ids))
        if request.category:
            conditions.append(OfficialDocument.category == request.category)
        if request.year:
            conditions.append(func.extract('year', OfficialDocument.doc_date) == request.year)
        if request.status:
            conditions.append(OfficialDocument.status == request.status)
        if request.keyword:
            keyword = f"%{request.keyword}%"
            conditions.append(
                or_(
                    OfficialDocument.subject.ilike(keyword),
                    OfficialDocument.doc_number.ilike(keyword),
                    OfficialDocument.sender.ilike(keyword),
                    OfficialDocument.receiver.ilike(keyword),
                    OfficialDocument.content.ilike(keyword),
                    OfficialDocument.notes.ilike(keyword)
                )
            )
        if request.contract_case:
            # contract_case 需要透過關聯查詢 ContractProject.project_name
            contract_case_keyword = f"%{request.contract_case}%"
            doc_query = doc_query.outerjoin(ContractProject, OfficialDocument.contract_project_id == ContractProject.id)
            conditions.append(ContractProject.project_name.ilike(contract_case_keyword))
        if request.sender:
            sender_keyword = f"%{request.sender}%"
            conditions.append(OfficialDocument.sender.ilike(sender_keyword))
        if request.receiver:
            receiver_keyword = f"%{request.receiver}%"
            conditions.append(OfficialDocument.receiver.ilike(receiver_keyword))

        if conditions:
            doc_query = doc_query.where(and_(*conditions))

        # 排序：依公文日期降序
        doc_query = doc_query.order_by(OfficialDocument.doc_date.desc())

        result = await db.execute(doc_query)
        documents = result.scalars().all()

        if not documents:
            raise HTTPException(status_code=404, detail="沒有符合條件的公文可供匯出")

        # 轉換為 DataFrame - 精簡欄位
        data = []

        def clean_agency_name(raw_text: str, agency_name: str = "") -> str:
            """清理機關名稱，移除代碼，只保留中文名稱"""
            # 優先使用關聯表的機關名稱
            if agency_name:
                return agency_name
            if not raw_text:
                return ""
            # 移除常見代碼格式
            import re
            text = raw_text.strip()
            # 移除括號內的內容提取為主名稱
            paren_match = re.search(r'[（(]([^)）]+)[)）]', text)
            if paren_match:
                return paren_match.group(1).strip()
            # 移除開頭的英數代碼（如 EB50819619、376470600A）
            text = re.sub(r'^[A-Za-z0-9]+\s*', '', text)
            return text.strip()

        def get_valid_doc_type(doc_type: str) -> str:
            """取得有效的公文類型"""
            # 保留防護：若仍有錯誤值則過濾
            if doc_type in ['收文', '發文']:
                return ""
            return doc_type or ""

        for doc in documents:
            # 取得關聯資料
            contract_case_name = ""
            if doc.contract_project:
                contract_case_name = doc.contract_project.project_name or ""

            sender_agency_name = ""
            if doc.sender_agency:
                sender_agency_name = doc.sender_agency.agency_name or ""

            receiver_agency_name = ""
            if doc.receiver_agency:
                receiver_agency_name = doc.receiver_agency.agency_name or ""

            # 統計附件數量
            attachment_count = len(doc.attachments) if doc.attachments else 0
            attachment_text = f"{attachment_count} 個附件" if attachment_count > 0 else "無"

            # 欄位順序依需求調整（公文ID對應附件資料夾 doc_{id}）
            data.append({
                "公文ID": doc.id,
                "流水號": doc.auto_serial or "",
                "發文形式": doc.delivery_method or "",
                "類別": doc.category or "",
                "公文類型": get_valid_doc_type(doc.doc_type),
                "公文字號": doc.doc_number or "",
                "主旨": doc.subject or "",
                "說明": getattr(doc, 'content', '') or "",
                "公文日期": str(doc.doc_date) if doc.doc_date else "",
                "收文日期": str(doc.receive_date) if doc.receive_date else "",
                "發文日期": str(doc.send_date) if doc.send_date else "",
                "發文單位": clean_agency_name(doc.sender or "", sender_agency_name),
                "受文單位": clean_agency_name(doc.receiver or "", receiver_agency_name),
                "附件紀錄": attachment_text,
                "備註": getattr(doc, 'notes', '') or "",
                "狀態": doc.status or "",
                "承攬案件": contract_case_name,
                "建立時間": str(doc.created_at) if doc.created_at else "",
                "更新時間": str(doc.updated_at) if doc.updated_at else "",
            })

        df = pd.DataFrame(data)

        # 產生 Excel
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='公文清單')

            # 取得工作表
            worksheet = writer.sheets['公文清單']

            # 表頭樣式：粗體 + 淺藍色背景
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            header_font = Font(bold=True, color="000000")
            header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # 套用表頭樣式
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            # 資料列樣式
            data_alignment = Alignment(vertical="center", wrap_text=True)
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                for cell in row:
                    cell.alignment = data_alignment
                    cell.border = thin_border

            # 調整欄位寬度
            for idx, col in enumerate(df.columns):
                max_length = max(
                    df[col].astype(str).map(len).max(),
                    len(col)
                ) + 2
                # 限制最大寬度
                max_length = min(max_length, 60)
                # Excel 欄位名稱 A-Z, AA-AZ...
                col_letter = chr(65 + idx) if idx < 26 else f"A{chr(65 + idx - 26)}"
                worksheet.column_dimensions[col_letter].width = max_length

            # 凍結表頭列
            worksheet.freeze_panes = 'A2'

            # 新增統計摘要工作表
            summary_data = {
                "項目": [
                    "匯出時間",
                    "公文總數",
                    "收文數量",
                    "發文數量",
                    "有附件公文",
                    "已指派案件",
                    "最早公文日期",
                    "最新公文日期"
                ],
                "數值": [
                    datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    str(len(documents)),
                    str(sum(1 for d in documents if d.category == '收文')),
                    str(sum(1 for d in documents if d.category == '發文')),
                    str(sum(1 for d in documents if d.attachments and len(d.attachments) > 0)),
                    str(sum(1 for d in documents if d.contract_project_id)),
                    str(min((d.doc_date for d in documents if d.doc_date), default="")) or "N/A",
                    str(max((d.doc_date for d in documents if d.doc_date), default="")) or "N/A"
                ]
            }
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, index=False, sheet_name='統計摘要')

            # 統計摘要工作表樣式
            summary_ws = writer.sheets['統計摘要']
            for cell in summary_ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            for row in summary_ws.iter_rows(min_row=2, max_row=summary_ws.max_row):
                for cell in row:
                    cell.alignment = data_alignment
                    cell.border = thin_border
            summary_ws.column_dimensions['A'].width = 20
            summary_ws.column_dimensions['B'].width = 30

        output.seek(0)

        # 產生檔名: 乾坤測繪公文總表YYYYMMDD.xlsx
        date_str = datetime.now().strftime('%Y%m%d')
        filename_cn = f"乾坤測繪公文總表{date_str}.xlsx"
        filename_encoded = quote(filename_cn)

        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{filename_encoded}",
                "X-Content-Type-Options": "nosniff",
                "Cache-Control": "no-cache, no-store, must-revalidate",
            }
        )

    except Exception as e:
        logger.error(f"匯出 Excel 失敗: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"匯出 Excel 失敗: {str(e)}")
