"""PM 里程碑 API 端點 (POST-only)"""
import io
import logging

from fastapi import APIRouter, Depends, HTTPException, UploadFile
from fastapi.responses import StreamingResponse

from app.core.dependencies import get_service, get_async_db
from app.services.pm import PMMilestoneService

logger = logging.getLogger(__name__)
from app.schemas.pm import (
    PMMilestoneCreate, PMMilestoneUpdate,
    PMIdRequest, PMCaseIdByFieldRequest, PMMilestoneUpdateRequest,
)
from app.schemas.common import SuccessResponse, DeleteResponse

router = APIRouter()


@router.post("/list")
async def list_milestones(
    req: PMCaseIdByFieldRequest,
    service: PMMilestoneService = Depends(get_service(PMMilestoneService)),
):
    """取得案件里程碑"""
    items = await service.get_by_case(req.pm_case_id)
    return SuccessResponse(data=items)


@router.post("/create")
async def create_milestone(
    data: PMMilestoneCreate,
    service: PMMilestoneService = Depends(get_service(PMMilestoneService)),
):
    """建立里程碑"""
    result = await service.create(data)
    return SuccessResponse(data=result, message="里程碑建立成功")


@router.post("/update")
async def update_milestone(
    req: PMMilestoneUpdateRequest,
    service: PMMilestoneService = Depends(get_service(PMMilestoneService)),
):
    """更新里程碑"""
    result = await service.update(req.id, req.data)
    if not result:
        raise HTTPException(status_code=404, detail="里程碑不存在")
    return SuccessResponse(data=result, message="里程碑更新成功")


@router.post("/delete")
async def delete_milestone(
    req: PMIdRequest,
    service: PMMilestoneService = Depends(get_service(PMMilestoneService)),
):
    """刪除里程碑"""
    success = await service.delete(req.id)
    if not success:
        raise HTTPException(status_code=404, detail="里程碑不存在")
    return DeleteResponse(deleted_id=req.id)


@router.post("/export-xlsx")
async def export_milestones_xlsx(
    body: dict = {},
    service: PMMilestoneService = Depends(get_service(PMMilestoneService)),
):
    """匯出里程碑為 XLSX

    Body: { "pm_case_id": 123 } → 匯出單一案件
    Body: {} → 匯出全部案件 (含案號對照)
    """
    import openpyxl
    from sqlalchemy import select, text
    from app.extended.models.pm import PMMilestone, PMCase

    pm_case_id = body.get("pm_case_id")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "里程碑"

    headers = [
        "id", "案件ID", "案號", "案名",
        "里程碑名稱", "類型", "預計日期", "實際完成日期",
        "狀態", "排序", "備註",
    ]
    ws.append(headers)

    # 標題列樣式
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    db = service.db
    if pm_case_id:
        query = (
            select(PMMilestone, PMCase.case_code, PMCase.case_name)
            .join(PMCase, PMCase.id == PMMilestone.pm_case_id)
            .where(PMMilestone.pm_case_id == pm_case_id)
            .order_by(PMMilestone.sort_order, PMMilestone.id)
        )
    else:
        query = (
            select(PMMilestone, PMCase.case_code, PMCase.case_name)
            .join(PMCase, PMCase.id == PMMilestone.pm_case_id)
            .order_by(PMCase.id, PMMilestone.sort_order, PMMilestone.id)
        )

    result = await db.execute(query)
    rows = result.all()

    for milestone, case_code, case_name in rows:
        ws.append([
            milestone.id,
            milestone.pm_case_id,
            case_code,
            case_name,
            milestone.milestone_name,
            milestone.milestone_type,
            str(milestone.planned_date) if milestone.planned_date else None,
            str(milestone.actual_date) if milestone.actual_date else None,
            milestone.status,
            milestone.sort_order,
            milestone.notes,
        ])

    # 如果無資料，加入範例行
    if not rows:
        ws.append([None, None, "(案號)", "(案名)", "範例：開工", "kickoff", "2026-04-01", None, "pending", 1, ""])
        ws.append([None, None, "(案號)", "(案名)", "範例：設計審查", "review", "2026-05-01", None, "pending", 2, ""])
        ws.append([None, None, "(案號)", "(案名)", "範例：驗收", "acceptance", "2026-08-01", None, "pending", 3, ""])

    # 加入說明 sheet
    ws2 = wb.create_sheet("欄位說明")
    ws2.append(["欄位", "說明", "允許值"])
    ws2.append(["id", "里程碑 ID (新增時留空)", ""])
    ws2.append(["案件ID", "PM 案件 ID (必填)", "數字"])
    ws2.append(["案號", "案件編號 (僅供參照)", ""])
    ws2.append(["案名", "案件名稱 (僅供參照)", ""])
    ws2.append(["里程碑名稱", "必填", "文字"])
    ws2.append(["類型", "里程碑類型", "kickoff/design/review/submission/acceptance/warranty/other"])
    ws2.append(["預計日期", "預定完成日", "YYYY-MM-DD"])
    ws2.append(["實際完成日期", "實際完成日", "YYYY-MM-DD"])
    ws2.append(["狀態", "執行狀態", "pending/in_progress/completed/overdue/skipped"])
    ws2.append(["排序", "顯示順序", "數字 (1, 2, 3...)"])
    ws2.append(["備註", "選填", "文字"])

    # 設定欄寬
    for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
        ws.column_dimensions[col_letter].width = 15
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 25

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"milestones_{pm_case_id or 'all'}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.post("/import-xlsx")
async def import_milestones_xlsx(
    file: UploadFile,
    service: PMMilestoneService = Depends(get_service(PMMilestoneService)),
):
    """匯入里程碑 XLSX — 新增或更新

    規則:
    - id 有值 → 更新既有里程碑
    - id 空白 + 案件ID 有值 → 新增里程碑
    - 案號/案名 僅供參照，不作為匯入依據
    """
    import openpyxl
    from datetime import date as date_type
    from app.extended.models.pm import PMMilestone
    from sqlalchemy import select

    try:
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active

        headers = [cell.value for cell in ws[1]]

        created = 0
        updated = 0
        errors = []
        db = service.db

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2):
            if not row or all(v is None for v in row):
                continue

            row_id = row[0]
            pm_case_id = row[1]
            milestone_name = row[4]
            milestone_type = row[5]
            planned_date = row[6]
            actual_date = row[7]
            status = row[8] or "pending"
            sort_order = row[9] or 0
            notes = row[10]

            if not milestone_name:
                continue

            # 日期處理
            def parse_date(val):
                if val is None:
                    return None
                if isinstance(val, date_type):
                    return val
                try:
                    from datetime import datetime
                    return datetime.strptime(str(val).strip()[:10], "%Y-%m-%d").date()
                except Exception:
                    return None

            pd = parse_date(planned_date)
            ad = parse_date(actual_date)

            if row_id:
                # 更新
                result = await db.execute(select(PMMilestone).where(PMMilestone.id == int(row_id)))
                milestone = result.scalar_one_or_none()
                if not milestone:
                    errors.append(f"Row {row_idx}: id={row_id} 不存在")
                    continue
                milestone.milestone_name = str(milestone_name)
                milestone.milestone_type = str(milestone_type) if milestone_type else None
                milestone.planned_date = pd
                milestone.actual_date = ad
                milestone.status = str(status)
                milestone.sort_order = int(sort_order) if sort_order else 0
                milestone.notes = str(notes) if notes else None
                updated += 1
            else:
                # 新增
                if not pm_case_id:
                    errors.append(f"Row {row_idx}: 缺少案件ID")
                    continue
                new_milestone = PMMilestone(
                    pm_case_id=int(pm_case_id),
                    milestone_name=str(milestone_name),
                    milestone_type=str(milestone_type) if milestone_type else None,
                    planned_date=pd,
                    actual_date=ad,
                    status=str(status),
                    sort_order=int(sort_order) if sort_order else 0,
                    notes=str(notes) if notes else None,
                )
                db.add(new_milestone)
                created += 1

        await db.commit()
        logger.info("[MILESTONE_IMPORT] created=%d updated=%d errors=%d", created, updated, len(errors))

        return {
            "success": True,
            "created": created,
            "updated": updated,
            "errors": errors,
            "total": created + updated,
        }

    except Exception as e:
        logger.error("Milestone XLSX import failed: %s", e, exc_info=True)
        return {"success": False, "error": f"匯入失敗: {str(e)[:200]}"}
