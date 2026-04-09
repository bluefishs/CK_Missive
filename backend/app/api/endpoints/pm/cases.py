"""PM 案件 API 端點 (POST-only)"""
import io
import logging

logger = logging.getLogger(__name__)

from fastapi import APIRouter, Depends, HTTPException, UploadFile
from fastapi.responses import StreamingResponse

from app.core.dependencies import get_service
from app.services.pm import PMCaseService
from app.schemas.pm import (
    PMCaseCreate, PMCaseUpdate, PMCaseResponse,
    PMCaseListRequest, PMCaseSummary,
    PMCaseIdRequest, PMCaseUpdateRequest,
    PMSummaryRequest, PMGenerateCodeRequest,
    PMCrossLookupRequest, PMLinkedDocsRequest, PMPromoteRequest,
)
from app.schemas.common import PaginatedResponse, SuccessResponse, DeleteResponse

router = APIRouter()


@router.post("/list")
async def list_cases(
    params: PMCaseListRequest,
    service: PMCaseService = Depends(get_service(PMCaseService)),
):
    """案件列表"""
    items, total = await service.list_cases(params)
    return PaginatedResponse.create(items=items, total=total, page=params.page, limit=params.limit)


@router.post("/create")
async def create_case(
    data: PMCaseCreate,
    service: PMCaseService = Depends(get_service(PMCaseService)),
):
    """建立案件 — 含重複檢查、案號格式驗證"""
    try:
        result = await service.create(data)
    except ValueError as e:
        raise HTTPException(status_code=409, detail=str(e))
    return SuccessResponse(data=result, message="案件建立成功")


@router.post("/yearly-trend")
async def get_yearly_trend(
    service: PMCaseService = Depends(get_service(PMCaseService)),
):
    """多年度案件趨勢"""
    result = await service.get_yearly_trend()
    return SuccessResponse(data=result)


@router.post("/detail")
async def get_case_detail(
    req: PMCaseIdRequest,
    service: PMCaseService = Depends(get_service(PMCaseService)),
):
    """案件詳情"""
    result = await service.get_detail(req.id)
    if not result:
        raise HTTPException(status_code=404, detail="案件不存在")
    return SuccessResponse(data=result)


@router.post("/update")
async def update_case(
    data: PMCaseUpdate,
    case_id: int = 0,
    service: PMCaseService = Depends(get_service(PMCaseService)),
):
    """更新案件 (case_id 由 request body 的 id 欄位或查詢參數傳入)

    當 status 變更為 'contracted' (已承攬) 時，自動觸發成案流程：
    - 產生 project_code
    - 建立 ContractProject
    - 同步 ERP Quotation
    """
    result = await service.update(case_id, data)
    if not result:
        raise HTTPException(status_code=404, detail="案件不存在")

    # 三方同步：核心欄位變更時同步到 ContractProject + ERPQuotation
    try:
        changed = {k: v for k, v in data.model_dump(exclude_unset=True).items()
                   if k in ("category", "case_nature", "client_name", "contract_amount")}
        if changed:
            from app.services.case_field_sync_service import CaseFieldSyncService
            sync_svc = CaseFieldSyncService(service.db)
            sync_result = await sync_svc.sync_from_pm(case_id, changed)
            if any(sync_result.values()):
                await service.db.commit()
                logger.info("[PM_UPDATE] 三方同步: case=%d fields=%s synced=%s", case_id, list(changed.keys()), sync_result)
    except Exception as e:
        logger.warning("三方同步失敗 (不影響更新): %s", e)

    # 自動成案：status 變更為已承攬時觸發
    new_status = getattr(data, 'status', None)
    if new_status == 'contracted':
        try:
            case_code = result.case_code if hasattr(result, 'case_code') else getattr(result, 'case_code', None)
            project_code = result.project_code if hasattr(result, 'project_code') else getattr(result, 'project_code', None)
            if case_code and not project_code:
                promote_result = await service.code_service.promote_to_project(case_code)
                return SuccessResponse(
                    data=result,
                    message=f"案件已承攬，自動成案: {promote_result['project_code']}",
                )
        except ValueError as e:
            logger.warning("自動成案失敗 (不影響更新): %s", e)
        except Exception as e:
            logger.warning("自動成案異常 (不影響更新): %s", e)

    return SuccessResponse(data=result, message="案件更新成功")


@router.post("/update-by-id")
async def update_case_by_id(
    req: PMCaseUpdateRequest,
    service: PMCaseService = Depends(get_service(PMCaseService)),
):
    """更新案件 (POST body 包含 id + data) — 含三方同步"""
    logger.info("[PM_UPDATE_BY_ID] id=%d fields=%s", req.id, list(req.data.model_dump(exclude_unset=True).keys()))
    result = await service.update(req.id, req.data)
    if not result:
        raise HTTPException(status_code=404, detail="案件不存在")

    # 三方同步
    try:
        changed = {k: v for k, v in req.data.model_dump(exclude_unset=True).items()
                   if k in ("category", "case_nature", "client_name", "contract_amount")}
        if changed:
            from app.services.case_field_sync_service import CaseFieldSyncService
            sync_svc = CaseFieldSyncService(service.db)
            sync_result = await sync_svc.sync_from_pm(req.id, changed)
            if any(sync_result.values()):
                await service.db.commit()
                logger.info("[PM_UPDATE_BY_ID] 三方同步: case=%d fields=%s synced=%s", req.id, list(changed.keys()), sync_result)
    except Exception as e:
        logger.warning("三方同步失敗 (不影響更新): %s", e)

    # 自動成案
    new_status = getattr(req.data, 'status', None)
    if new_status == 'contracted':
        try:
            case_code = result.case_code if hasattr(result, 'case_code') else None
            project_code = result.project_code if hasattr(result, 'project_code') else None
            if case_code and not project_code:
                promote_result = await service.code_service.promote_to_project(case_code)
                return SuccessResponse(data=result, message=f"案件已承攬，自動成案: {promote_result['project_code']}")
        except Exception as e:
            logger.warning("自動成案失敗: %s", e)

    return SuccessResponse(data=result, message="案件更新成功")


@router.post("/delete")
async def delete_case(
    req: PMCaseIdRequest,
    service: PMCaseService = Depends(get_service(PMCaseService)),
):
    """刪除案件"""
    success = await service.delete(req.id)
    if not success:
        raise HTTPException(status_code=404, detail="案件不存在")
    return DeleteResponse(deleted_id=req.id)


@router.post("/summary")
async def get_summary(
    req: PMSummaryRequest,
    service: PMCaseService = Depends(get_service(PMCaseService)),
):
    """案件統計摘要"""
    result = await service.get_summary(year=req.year)
    return SuccessResponse(data=result)


@router.post("/generate-code")
async def generate_case_code(
    req: PMGenerateCodeRequest,
    service: PMCaseService = Depends(get_service(PMCaseService)),
):
    """產生 PM 案號"""
    code = await service.generate_case_code(year=req.year, category=req.category)
    return SuccessResponse(data={"case_code": code})


@router.post("/recalculate-progress")
async def recalculate_progress(
    req: PMCaseIdRequest,
    service: PMCaseService = Depends(get_service(PMCaseService)),
):
    """根據里程碑完成率重新計算進度"""
    progress = await service.recalculate_progress(req.id)
    if progress is None:
        raise HTTPException(status_code=404, detail="案件不存在或無里程碑")
    return SuccessResponse(data={"progress": progress})


@router.post("/gantt")
async def generate_gantt(
    req: PMCaseIdRequest,
    service: PMCaseService = Depends(get_service(PMCaseService)),
):
    """產生案件里程碑甘特圖 (Mermaid Gantt 語法)"""
    gantt = await service.generate_gantt(req.id)
    if gantt is None:
        raise HTTPException(status_code=404, detail="案件不存在")
    return SuccessResponse(data={"gantt_mermaid": gantt})


@router.post("/export")
async def export_cases(
    req: PMSummaryRequest,
    service: PMCaseService = Depends(get_service(PMCaseService)),
):
    """匯出案件 CSV"""
    csv_content = await service.export_csv(year=req.year)
    return StreamingResponse(
        io.StringIO(csv_content),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=pm_cases.csv"},
    )


@router.post("/export-xlsx")
async def export_cases_xlsx(
    service: PMCaseService = Depends(get_service(PMCaseService)),
):
    """匯出全部 PM 案件為 XLSX (含 case_nature, category 等完整欄位)"""
    from app.extended.models.pm import PMCase
    from sqlalchemy import select

    result = await service.db.execute(
        select(PMCase).order_by(PMCase.id.desc())
    )
    cases = result.scalars().all()

    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PM Cases"

    headers = ["id", "案號", "案名", "年度", "計畫類別", "作業性質",
               "委託單位", "委託單位ID", "報價金額", "承攬狀態", "進度(%)",
               "開始日期", "結束日期", "作業地點", "成案編號", "說明", "備註"]
    ws.append(headers)

    for c in cases:
        ws.append([
            c.id, c.case_code, c.case_name, c.year, c.category,
            getattr(c, 'case_nature', None),
            c.client_name, c.client_vendor_id,
            float(c.contract_amount) if c.contract_amount else None,
            c.status, c.progress,
            str(c.start_date) if c.start_date else None,
            str(c.end_date) if c.end_date else None,
            c.location, c.project_code, c.description, c.notes,
        ])

    # 寫入 BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=pm_cases_full.xlsx"},
    )


@router.post("/batch-update")
async def batch_update_cases(
    body: dict,
    service: PMCaseService = Depends(get_service(PMCaseService)),
):
    """批次更新 PM 案件 (從 XLS 匯入修正資料)

    Body: { "updates": [{"id": 1, "category": "01", "case_nature": "05安全檢測", ...}, ...] }
    僅更新提供的欄位，未提供的保持不變。
    同步更新 ContractProject + ERPQuotation。
    """
    from app.extended.models.pm import PMCase
    from app.services.case_field_sync_service import CaseFieldSyncService
    from sqlalchemy import select

    updates = body.get("updates", [])
    if not updates:
        return {"success": False, "error": "未提供更新資料"}

    allowed_fields = {"category", "case_nature", "client_name", "contract_amount",
                      "status", "location", "description", "notes", "year"}

    updated = 0
    synced = 0
    errors = []
    sync_svc = CaseFieldSyncService(service.db)

    for item in updates:
        pm_id = item.get("id")
        if not pm_id:
            continue

        result = await service.db.execute(select(PMCase).where(PMCase.id == pm_id))
        pm = result.scalar_one_or_none()
        if not pm:
            errors.append(f"id={pm_id} 不存在")
            continue

        changed = {}
        for field in allowed_fields:
            if field in item and item[field] is not None:
                new_val = item[field]
                if field == "year" and isinstance(new_val, (int, float)):
                    new_val = int(new_val)
                    if new_val < 1911:
                        new_val = new_val + 1911
                if field == "category":
                    cat_str = str(new_val).strip()[:2]
                    new_val = "01" if cat_str == "01" else "02"
                if field == "contract_amount" and new_val is not None:
                    try:
                        new_val = float(new_val)
                    except (ValueError, TypeError):
                        continue
                setattr(pm, field, new_val)
                changed[field] = new_val

        if changed:
            updated += 1
            # 三方同步
            try:
                sync_result = await sync_svc.sync_from_pm(pm_id, changed)
                if any(sync_result.values()):
                    synced += 1
            except Exception as e:
                logger.warning("batch sync failed for id=%d: %s", pm_id, e)

    await service.db.commit()

    return {
        "success": True,
        "updated": updated,
        "synced": synced,
        "errors": errors,
        "total": len(updates),
    }


@router.post("/import-xlsx")
async def import_cases_xlsx(
    file: UploadFile,
    service: PMCaseService = Depends(get_service(PMCaseService)),
):
    """匯入 XLSX 修正 PM 案件 — 後端解析 + 三方同步"""
    import openpyxl

    try:
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        header_map = {
            "計畫類別": "category", "作業性質": "case_nature",
            "委託單位": "client_name", "報價金額": "contract_amount",
            "承攬狀態": "status", "作業地點": "location",
            "說明": "description", "備註": "notes", "年度": "year",
        }

        updates = []
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            if not row[0]:
                continue
            item = {"id": row[0]}
            for col_idx, header in enumerate(headers):
                if header in header_map and row[col_idx] is not None:
                    item[header_map[header]] = row[col_idx]
            updates.append(item)

        if not updates:
            return {"success": False, "error": "未找到有效資料"}

        # 複用 batch-update 邏輯
        from app.extended.models.pm import PMCase
        from app.services.case_field_sync_service import CaseFieldSyncService
        from sqlalchemy import select

        allowed_fields = {"category", "case_nature", "client_name", "contract_amount",
                          "status", "location", "description", "notes", "year"}

        updated = 0
        synced = 0
        errors = []
        sync_svc = CaseFieldSyncService(service.db)

        for item in updates:
            pm_id = item.get("id")
            result = await service.db.execute(select(PMCase).where(PMCase.id == pm_id))
            pm = result.scalar_one_or_none()
            if not pm:
                errors.append(f"id={pm_id} 不存在")
                continue

            changed = {}
            for field in allowed_fields:
                if field in item and item[field] is not None:
                    new_val = item[field]
                    # 年度：民國→西元轉換
                    if field == "year" and isinstance(new_val, (int, float)):
                        new_val = int(new_val)
                        if new_val < 1911:
                            new_val = new_val + 1911
                    # 計畫類別歸併 (03/04/05/06/07/99 → 02承攬報價, 01 → 01委辦招標)
                    if field == "category":
                        cat_str = str(new_val).strip()[:2]
                        new_val = "01" if cat_str == "01" else "02"
                    # 金額正規化
                    if field == "contract_amount" and new_val is not None:
                        try:
                            new_val = float(new_val)
                        except (ValueError, TypeError):
                            continue
                    setattr(pm, field, new_val)
                    changed[field] = new_val

            if changed:
                updated += 1
                try:
                    sync_result = await sync_svc.sync_from_pm(pm_id, changed)
                    if any(sync_result.values()):
                        synced += 1
                except Exception as e:
                    logger.warning("import sync failed for id=%d: %s", pm_id, e)

        await service.db.commit()
        logger.info("[PM_IMPORT] updated=%d synced=%d errors=%d total=%d", updated, synced, len(errors), len(updates))

        return {"success": True, "updated": updated, "synced": synced, "errors": errors, "total": len(updates)}

    except Exception as e:
        logger.error("XLSX import failed: %s", e, exc_info=True)
        return {"success": False, "error": f"匯入失敗: {str(e)[:200]}"}


@router.post("/cross-lookup")
async def cross_module_lookup(
    req: PMCrossLookupRequest,
    service: PMCaseService = Depends(get_service(PMCaseService)),
):
    """跨模組案號查詢 — 回傳 PM/ERP 兩端資料"""
    result = await service.code_service.cross_module_lookup(req.case_code)
    return SuccessResponse(data=result)


@router.post("/linked-documents")
async def get_linked_documents(
    req: PMLinkedDocsRequest,
    service: PMCaseService = Depends(get_service(PMCaseService)),
):
    """透過案號查詢相關公文 (case_code → ContractProject → OfficialDocument)"""
    docs = await service.code_service.find_linked_documents(req.case_code, req.limit)
    return SuccessResponse(data=docs)


@router.post("/promote")
async def promote_to_project(
    req: PMPromoteRequest,
    service: PMCaseService = Depends(get_service(PMCaseService)),
):
    """成案：從邀標/報價轉為正式承攬案件

    自動產生 project_code，建立 ContractProject，連結 ERP Quotation。
    """
    try:
        result = await service.code_service.promote_to_project(req.case_code)
        return SuccessResponse(data=result, message=f"成案成功，專案編號: {result['project_code']}")
    except ValueError as e:
        from fastapi import HTTPException
        raise HTTPException(status_code=400, detail=str(e))
