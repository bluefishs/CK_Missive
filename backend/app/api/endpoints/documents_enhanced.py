"""
增強版公文管理 API 端點 - POST-only 資安機制，統一回應格式
v2.3 - 權限管控升級 (2026-01-10)

變更紀錄:
- v2.3: 新增行級別權限過濾（非管理員只能查看關聯專案的公文）
- v2.2: 使用 AuditService 統一管理審計日誌（獨立 session，不會污染主交易）
- v2.2: 使用 NotificationService.safe_* 方法（獨立 session）
- v2.2: 移除對舊 log_document_change 函數的依賴
"""
import io
import os
import csv
import shutil
import logging
from typing import Optional, List
from datetime import date as date_type
from fastapi import APIRouter, Query, Depends, Body, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import text, select, or_, and_, func
from sqlalchemy.orm import selectinload
from pydantic import BaseModel, Field, field_validator

logger = logging.getLogger(__name__)


def parse_date_string(date_str: Optional[str]) -> Optional[date_type]:
    """將日期字串轉換為 Python date 物件"""
    if not date_str:
        return None
    try:
        # 支援 YYYY-MM-DD 格式
        parts = date_str.split('-')
        if len(parts) == 3:
            return date_type(int(parts[0]), int(parts[1]), int(parts[2]))
        return None
    except (ValueError, IndexError):
        logger.warning(f"無法解析日期字串: {date_str}")
        return None

from app.db.database import get_async_db
from app.extended.models import OfficialDocument, ContractProject, GovernmentAgency, DocumentAttachment
from app.services.document_service import DocumentService
from app.schemas.document import (
    DocumentFilter, DocumentListQuery, DocumentListResponse, DocumentResponse, StaffInfo,
    DocumentCreateRequest, DocumentUpdateRequest, VALID_DOC_TYPES
)
from app.extended.models import User, project_user_assignment
from app.schemas.common import (
    PaginationMeta,
    DeleteResponse,
    SuccessResponse,
    SortOrder,
)
# 統一從 schemas 匯入查詢相關型別
from app.schemas.document_query import (
    DropdownQuery,
    AgencyDropdownQuery,
    OptimizedSearchRequest,
    SearchSuggestionRequest,
    AuditLogQuery,
    AuditLogItem,
    AuditLogResponse,
    ProjectDocumentsQuery,
    DocumentExportQuery,
    ExcelExportRequest,
)
from app.core.exceptions import NotFoundException, ForbiddenException
from app.core.audit_logger import DocumentUpdateGuard
from app.services.notification_service import NotificationService, CRITICAL_FIELDS
from app.core.dependencies import require_auth, require_permission

# 使用者認證（v2.3 升級為必要）
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
security = HTTPBearer(auto_error=False)

# 統一使用 require_auth 進行認證
from app.api.endpoints.auth import get_current_user

async def get_optional_user(
    credentials: Optional[HTTPAuthorizationCredentials] = Depends(security),
    db: AsyncSession = Depends(get_async_db)
) -> Optional[User]:
    """取得當前使用者（可選，不強制認證）- 僅用於向後相容"""
    try:
        if not credentials:
            return None
        return await get_current_user(credentials, db)
    except Exception:
        return None

router = APIRouter()


# ============================================================================
# 公文列表 API（POST-only 資安機制）
# ============================================================================

@router.post(
    "/list",
    response_model=DocumentListResponse,
    summary="查詢公文列表",
    description="使用統一分頁格式查詢公文列表（POST-only 資安機制，含行級別權限過濾）"
)
async def list_documents(
    query: DocumentListQuery = Body(default=DocumentListQuery()),
    db: AsyncSession = Depends(get_async_db),
    current_user: User = Depends(require_auth())
):
    """
    查詢公文列表（POST-only 資安機制）

    🔒 權限規則：
    - 需要登入認證
    - superuser/admin: 可查看所有公文
    - 一般使用者: 只能查看關聯專案的公文，或無專案關聯的公文

    回應格式：
    ```json
    {
        "success": true,
        "items": [...],
        "pagination": {
            "total": 100,
            "page": 1,
            "limit": 20,
            "total_pages": 5,
            "has_next": true,
            "has_prev": false
        }
    }
    ```
    """
    try:
        # 詳細記錄所有查詢參數
        logger.info(f"[API] 公文查詢請求: keyword={query.keyword}, doc_number={query.doc_number}, "
                   f"doc_type={query.doc_type}, year={query.year}, "
                   f"sender={query.sender}, receiver={query.receiver}, "
                   f"delivery_method={query.delivery_method}, "
                   f"doc_date_from={query.doc_date_from}, doc_date_to={query.doc_date_to}, "
                   f"contract_case={query.contract_case}, category={query.category}")

        service = DocumentService(db)

        # 構建篩選條件
        filters = DocumentFilter(
            keyword=query.keyword,
            doc_number=query.doc_number,  # 公文字號專用篩選
            doc_type=query.doc_type,
            year=query.year,
            status=query.status,
            sender=query.sender,
            receiver=query.receiver,
            date_from=query.doc_date_from,
            date_to=query.doc_date_to,
            delivery_method=query.delivery_method,
            contract_case=query.contract_case,  # 直接設定，不用 setattr
            sort_by=query.sort_by,
            sort_order=query.sort_order.value if query.sort_order else "desc"
        )

        # 加入收發文分類篩選 (前端用 send/receive，資料庫用 發文/收文)
        if query.category:
            category_mapping = {'send': '發文', 'receive': '收文'}
            db_category = category_mapping.get(query.category, query.category)
            setattr(filters, 'category', db_category)

        # 計算 skip
        skip = (query.page - 1) * query.limit

        # 傳遞 current_user 進行行級別權限過濾
        result = await service.get_documents(
            skip=skip,
            limit=query.limit,
            filters=filters,
            current_user=current_user
        )

        # 轉換為統一回應格式
        items = result.get("items", [])
        total = result.get("total", 0)

        # 收集所有 project_id 以批次查詢
        project_ids = list(set(doc.contract_project_id for doc in items if doc.contract_project_id))

        # 批次查詢承攬案件資訊
        project_map = {}
        staff_map = {}
        if project_ids:
            # 查詢案件名稱
            project_query = select(ContractProject.id, ContractProject.project_name).where(
                ContractProject.id.in_(project_ids)
            )
            project_result = await db.execute(project_query)
            for row in project_result:
                project_map[row.id] = row.project_name

            # 查詢業務同仁
            staff_query = select(
                project_user_assignment.c.project_id,
                project_user_assignment.c.role,
                User.id.label('user_id'),
                User.full_name
            ).select_from(
                project_user_assignment.join(User, project_user_assignment.c.user_id == User.id)
            ).where(project_user_assignment.c.project_id.in_(project_ids))

            staff_result = await db.execute(staff_query)
            for row in staff_result:
                if row.project_id not in staff_map:
                    staff_map[row.project_id] = []
                staff_map[row.project_id].append(StaffInfo(
                    user_id=row.user_id,
                    name=row.full_name or '未知',
                    role=row.role or 'member'
                ))

        # 批次查詢附件數量（N+1 優化）
        attachment_count_map = {}
        doc_ids = [doc.id for doc in items]
        if doc_ids:
            attachment_query = select(
                DocumentAttachment.document_id,
                func.count(DocumentAttachment.id).label('count')
            ).where(
                DocumentAttachment.document_id.in_(doc_ids)
            ).group_by(DocumentAttachment.document_id)

            attachment_result = await db.execute(attachment_query)
            for row in attachment_result:
                attachment_count_map[row.document_id] = row.count

        # 批次查詢機關名稱（2026-01-08 新增，支援前端顯示）
        agency_map = {}
        agency_ids = set()
        for doc in items:
            if doc.sender_agency_id:
                agency_ids.add(doc.sender_agency_id)
            if doc.receiver_agency_id:
                agency_ids.add(doc.receiver_agency_id)

        if agency_ids:
            agency_query = select(
                GovernmentAgency.id,
                GovernmentAgency.agency_name
            ).where(GovernmentAgency.id.in_(agency_ids))
            agency_result = await db.execute(agency_query)
            for row in agency_result:
                agency_map[row.id] = row.agency_name

        # 轉換為 DocumentResponse
        response_items = []
        for doc in items:
            try:
                doc_dict = {
                    **doc.__dict__,
                    'contract_project_name': project_map.get(doc.contract_project_id) if doc.contract_project_id else None,
                    'assigned_staff': staff_map.get(doc.contract_project_id, []) if doc.contract_project_id else [],
                    'attachment_count': attachment_count_map.get(doc.id, 0),
                    # 機關名稱虛擬欄位
                    'sender_agency_name': agency_map.get(doc.sender_agency_id) if doc.sender_agency_id else None,
                    'receiver_agency_name': agency_map.get(doc.receiver_agency_id) if doc.receiver_agency_id else None,
                }
                # 移除 SQLAlchemy 內部屬性
                doc_dict.pop('_sa_instance_state', None)
                response_items.append(DocumentResponse.model_validate(doc_dict))
            except Exception as e:
                logger.warning(f"轉換公文資料失敗: {e}")
                continue

        return DocumentListResponse(
            items=response_items,
            pagination=PaginationMeta.create(
                total=total,
                page=query.page,
                limit=query.limit
            )
        )

    except Exception as e:
        logger.error(f"公文查詢失敗: {e}", exc_info=True)
        return DocumentListResponse(
            items=[],
            pagination=PaginationMeta.create(total=0, page=1, limit=query.limit)
        )


# ============================================================================
# 下拉選項 API（POST-only 資安機制）
# ============================================================================

@router.post(
    "/contract-projects-dropdown",
    summary="取得承攬案件下拉選項"
)
async def get_contract_projects_dropdown(
    query: DropdownQuery = Body(default=DropdownQuery()),
    db: AsyncSession = Depends(get_async_db)
):
    """取得承攬案件下拉選項 - 從 contract_projects 表查詢"""
    try:
        db_query = select(
            ContractProject.id,
            ContractProject.project_name,
            ContractProject.year,
            ContractProject.category
        )

        if query.search:
            db_query = db_query.where(
                or_(
                    ContractProject.project_name.ilike(f"%{query.search}%"),
                    ContractProject.project_code.ilike(f"%{query.search}%"),
                    ContractProject.client_agency.ilike(f"%{query.search}%")
                )
            )

        db_query = db_query.order_by(
            ContractProject.year.desc(),
            ContractProject.project_name.asc()
        ).limit(query.limit)

        result = await db.execute(db_query)
        projects = result.all()

        options = []
        for project in projects:
            options.append({
                "value": project.project_name,
                "label": f"{project.project_name} ({project.year}年)",
                "id": project.id,
                "year": project.year,
                "category": project.category
            })

        return {
            "success": True,
            "options": options,
            "total": len(options)
        }

    except Exception as e:
        logger.error(f"取得承攬案件選項失敗: {e}", exc_info=True)
        return {"success": False, "options": [], "total": 0, "error": str(e)}


def _extract_agency_names_from_raw(raw_value: str) -> list:
    """
    從資料庫原始值中提取純機關名稱

    支援格式：
    - 純名稱: "桃園市政府"
    - 代碼+名稱: "380110000G (桃園市政府工務局)"
    - 多機關: "376480000A (南投縣政府) | A01020100G (內政部國土管理署城鄉發展分署)"
    - 換行格式: "380110000G\\n(桃園市政府工務局)"

    Returns:
        提取出的純機關名稱列表
    """
    import re

    if not raw_value:
        return []

    names = []

    # 先按 | 分割多個機關
    parts = raw_value.split('|')

    for part in parts:
        part = part.strip()
        if not part:
            continue

        # 處理換行格式: "380110000G\n(桃園市政府工務局)"
        part = part.replace('\n', ' ').replace('\r', ' ')

        # 嘗試提取括號內的名稱: "380110000G (桃園市政府工務局)" -> "桃園市政府工務局"
        match = re.search(r'\(([^)]+)\)', part)
        if match:
            names.append(match.group(1).strip())
        else:
            # 嘗試移除代碼前綴: "380110000G 桃園市政府工務局" -> "桃園市政府工務局"
            # 代碼格式通常是 字母+數字 組合
            cleaned = re.sub(r'^[A-Z0-9]+\s*', '', part, flags=re.IGNORECASE)
            if cleaned:
                names.append(cleaned.strip())
            else:
                # 如果全都被移除了，就用原值（可能本身就是純名稱）
                names.append(part)

    return names


@router.post(
    "/agencies-dropdown",
    summary="取得政府機關下拉選項"
)
async def get_agencies_dropdown(
    query: AgencyDropdownQuery = Body(default=AgencyDropdownQuery()),
    db: AsyncSession = Depends(get_async_db)
):
    """
    取得政府機關下拉選項

    優化版：從 government_agencies 表取得標準化機關名稱，
    與 http://localhost:3000/agencies 頁面顯示一致。
    """
    try:
        # 從 government_agencies 表取得標準化機關名稱
        sql_query = """
        SELECT id, agency_name, agency_code, agency_short_name
        FROM government_agencies
        WHERE agency_name IS NOT NULL AND agency_name != ''
        """

        params = {}
        if query.search:
            sql_query += " AND (agency_name ILIKE :search OR agency_short_name ILIKE :search)"
            params["search"] = f"%{query.search}%"

        sql_query += " ORDER BY agency_name LIMIT :limit"
        params["limit"] = query.limit

        result = await db.execute(text(sql_query), params)
        agencies = result.fetchall()

        options = []
        for row in agencies:
            options.append({
                "value": row.agency_name,  # 使用標準化名稱作為值
                "label": row.agency_name,  # 顯示標準化名稱
                "id": row.id,
                "agency_code": row.agency_code or "",
                "agency_short_name": row.agency_short_name or ""
            })

        return {
            "success": True,
            "options": options,
            "total": len(options)
        }

    except Exception as e:
        logger.error(f"取得政府機關選項失敗: {e}", exc_info=True)
        return {"success": False, "options": [], "total": 0, "error": str(e)}


@router.post(
    "/years",
    summary="取得文檔年度選項"
)
async def get_document_years(
    db: AsyncSession = Depends(get_async_db)
):
    """取得文檔年度選項"""
    try:
        sql_query = """
        SELECT DISTINCT EXTRACT(YEAR FROM doc_date) as year
        FROM documents
        WHERE doc_date IS NOT NULL
        ORDER BY year DESC
        """

        result = await db.execute(text(sql_query))
        years = result.fetchall()

        year_list = [int(row.year) for row in years if row.year]

        return {
            "success": True,
            "years": year_list,
            "total": len(year_list)
        }

    except Exception as e:
        logger.error(f"取得年度選項失敗: {e}", exc_info=True)
        return {"success": False, "years": [], "total": 0}


@router.post(
    "/statistics",
    summary="取得公文統計資料"
)
async def get_documents_statistics(
    db: AsyncSession = Depends(get_async_db)
):
    """取得公文統計資料 (收發文分類基於 category 欄位)"""
    try:
        total_query = "SELECT COUNT(*) as count FROM documents"
        send_query = "SELECT COUNT(*) as count FROM documents WHERE category = '發文'"
        receive_query = "SELECT COUNT(*) as count FROM documents WHERE category = '收文'"
        current_year_query = "SELECT COUNT(*) as count FROM documents WHERE EXTRACT(YEAR FROM doc_date) = EXTRACT(YEAR FROM CURRENT_DATE)"

        # 發文形式統計 (僅統計發文類別)
        electronic_query = "SELECT COUNT(*) as count FROM documents WHERE category = '發文' AND delivery_method = '電子交換'"
        paper_query = "SELECT COUNT(*) as count FROM documents WHERE category = '發文' AND delivery_method = '紙本郵寄'"
        both_query = "SELECT COUNT(*) as count FROM documents WHERE category = '發文' AND delivery_method = '電子+紙本'"

        # 本年度發文數
        current_year_send_query = "SELECT COUNT(*) as count FROM documents WHERE category = '發文' AND EXTRACT(YEAR FROM doc_date) = EXTRACT(YEAR FROM CURRENT_DATE)"

        total_result = await db.execute(text(total_query))
        send_result = await db.execute(text(send_query))
        receive_result = await db.execute(text(receive_query))
        current_year_result = await db.execute(text(current_year_query))

        # 發文形式統計
        electronic_result = await db.execute(text(electronic_query))
        paper_result = await db.execute(text(paper_query))
        both_result = await db.execute(text(both_query))
        current_year_send_result = await db.execute(text(current_year_send_query))

        total = total_result.scalar() or 0
        send = send_result.scalar() or 0
        receive = receive_result.scalar() or 0

        return {
            "success": True,
            "total": total,
            "total_documents": total,
            "send": send,
            "send_count": send,
            "receive": receive,
            "receive_count": receive,
            "current_year_count": current_year_result.scalar() or 0,
            "current_year_send_count": current_year_send_result.scalar() or 0,
            "delivery_method_stats": {
                "electronic": electronic_result.scalar() or 0,
                "paper": paper_result.scalar() or 0,
                "both": both_result.scalar() or 0
            }
        }
    except Exception as e:
        logger.error(f"取得統計資料失敗: {e}", exc_info=True)
        return {
            "success": False,
            "total": 0,
            "total_documents": 0,
            "send": 0,
            "send_count": 0,
            "receive": 0,
            "receive_count": 0,
            "current_year_count": 0,
            "current_year_send_count": 0,
            "delivery_method_stats": {
                "electronic": 0,
                "paper": 0,
                "both": 0
            }
        }


@router.post(
    "/filtered-statistics",
    summary="取得篩選後的公文統計資料"
)
async def get_filtered_statistics(
    query: DocumentListQuery = Body(default=DocumentListQuery()),
    db: AsyncSession = Depends(get_async_db)
):
    """
    根據篩選條件取得動態統計資料

    回傳基於當前篩選條件的：
    - total: 符合條件的總數
    - send_count: 符合條件的發文數
    - receive_count: 符合條件的收文數

    用於前端 Tab 標籤的動態數字顯示
    """
    try:
        # 構建基本 WHERE 條件（不含 category）
        conditions = []
        params = {}

        # 公文字號專用篩選（僅搜尋 doc_number 欄位）
        if query.doc_number:
            conditions.append("doc_number ILIKE :doc_number")
            params["doc_number"] = f"%{query.doc_number}%"

        # 關鍵字搜尋（主旨、說明、備註 - 不含 doc_number）
        if query.keyword:
            conditions.append("""
                (subject ILIKE :keyword
                 OR content ILIKE :keyword OR notes ILIKE :keyword)
            """)
            params["keyword"] = f"%{query.keyword}%"

        if query.doc_type:
            conditions.append("doc_type = :doc_type")
            params["doc_type"] = query.doc_type

        if query.year:
            conditions.append("EXTRACT(YEAR FROM doc_date) = :year")
            params["year"] = query.year

        if query.sender:
            conditions.append("sender ILIKE :sender")
            params["sender"] = f"%{query.sender}%"

        if query.receiver:
            conditions.append("receiver ILIKE :receiver")
            params["receiver"] = f"%{query.receiver}%"

        if query.delivery_method:
            conditions.append("delivery_method = :delivery_method")
            params["delivery_method"] = query.delivery_method

        if query.doc_date_from:
            conditions.append("doc_date >= :doc_date_from")
            params["doc_date_from"] = query.doc_date_from

        if query.doc_date_to:
            conditions.append("doc_date <= :doc_date_to")
            params["doc_date_to"] = query.doc_date_to

        if query.contract_case:
            # 需要 JOIN contract_projects 表
            conditions.append("""
                contract_project_id IN (
                    SELECT id FROM contract_projects
                    WHERE project_name ILIKE :contract_case OR project_code ILIKE :contract_case
                )
            """)
            params["contract_case"] = f"%{query.contract_case}%"

        # 組合 WHERE 子句
        where_clause = " AND ".join(conditions) if conditions else "1=1"

        # 總數查詢
        total_query = f"SELECT COUNT(*) FROM documents WHERE {where_clause}"
        total_result = await db.execute(text(total_query), params)
        total = total_result.scalar() or 0

        # 發文數查詢
        send_query = f"SELECT COUNT(*) FROM documents WHERE {where_clause} AND category = '發文'"
        send_result = await db.execute(text(send_query), params)
        send_count = send_result.scalar() or 0

        # 收文數查詢
        receive_query = f"SELECT COUNT(*) FROM documents WHERE {where_clause} AND category = '收文'"
        receive_result = await db.execute(text(receive_query), params)
        receive_count = receive_result.scalar() or 0

        logger.info(f"篩選統計: total={total}, send={send_count}, receive={receive_count}, filters={params}")

        return {
            "success": True,
            "total": total,
            "send_count": send_count,
            "receive_count": receive_count,
            "filters_applied": bool(conditions)
        }

    except Exception as e:
        logger.error(f"取得篩選統計失敗: {e}", exc_info=True)
        return {
            "success": False,
            "total": 0,
            "send_count": 0,
            "receive_count": 0,
            "filters_applied": False,
            "error": str(e)
        }


# ============================================================================
# 優化搜尋 API
# ============================================================================

@router.post(
    "/search/optimized",
    summary="優化全文搜尋",
    description="使用智能關鍵字處理和結果排名的優化搜尋"
)
async def optimized_search(
    request: OptimizedSearchRequest,
    db: AsyncSession = Depends(get_async_db)
):
    """
    優化全文搜尋

    特點：
    - 智能關鍵字分詞處理
    - 支援公文字號格式識別
    - 多欄位權重搜尋
    - 搜尋結果快取
    """
    from app.services.search_optimizer import SearchOptimizer

    try:
        optimizer = SearchOptimizer(db)

        # 構建篩選條件
        filters = {}
        if request.category:
            category_mapping = {'send': '發文', 'receive': '收文'}
            filters['category'] = category_mapping.get(request.category, request.category)
        if request.delivery_method:
            filters['delivery_method'] = request.delivery_method
        if request.year:
            filters['year'] = request.year

        # 執行優化搜尋
        skip = (request.page - 1) * request.limit
        result = await optimizer.search_with_ranking(
            keyword=request.keyword,
            filters=filters,
            skip=skip,
            limit=request.limit
        )

        # 轉換結果格式
        items = []
        for doc in result.get("items", []):
            items.append({
                "id": doc.id,
                "doc_number": doc.doc_number,
                "doc_type": doc.doc_type,
                "subject": doc.subject,
                "sender": doc.sender,
                "receiver": doc.receiver,
                "doc_date": str(doc.doc_date) if doc.doc_date else None,
                "category": doc.category,
                "delivery_method": doc.delivery_method,
                "status": doc.status,
            })

        total = result.get("total", 0)

        return {
            "success": True,
            "items": items,
            "pagination": {
                "total": total,
                "page": request.page,
                "limit": request.limit,
                "total_pages": (total + request.limit - 1) // request.limit if request.limit > 0 else 0,
                "has_next": request.page * request.limit < total,
                "has_prev": request.page > 1
            },
            "search_info": {
                "tokens": result.get("tokens", []),
                "normalized_keyword": result.get("normalized_keyword", request.keyword)
            }
        }

    except Exception as e:
        logger.error(f"優化搜尋失敗: {e}", exc_info=True)
        return {
            "success": False,
            "items": [],
            "pagination": {"total": 0, "page": 1, "limit": request.limit},
            "error": str(e)
        }


@router.post(
    "/search/suggestions",
    summary="取得搜尋建議",
    description="根據輸入前綴提供自動完成建議"
)
async def get_search_suggestions(
    request: SearchSuggestionRequest,
    db: AsyncSession = Depends(get_async_db)
):
    """
    取得搜尋建議（自動完成）

    根據用戶輸入提供：
    - 主旨匹配建議
    - 文號匹配建議
    """
    from app.services.search_optimizer import SearchOptimizer

    try:
        optimizer = SearchOptimizer(db)
        suggestions = await optimizer.get_search_suggestions(
            prefix=request.prefix,
            limit=request.limit
        )

        return {
            "success": True,
            "suggestions": suggestions,
            "count": len(suggestions)
        }

    except Exception as e:
        logger.error(f"取得搜尋建議失敗: {e}", exc_info=True)
        return {
            "success": False,
            "suggestions": [],
            "error": str(e)
        }


@router.post(
    "/search/popular",
    summary="取得熱門搜尋詞",
    description="取得最近的熱門搜尋關鍵詞"
)
async def get_popular_searches(
    limit: int = Query(default=10, ge=1, le=20, description="數量上限"),
    db: AsyncSession = Depends(get_async_db)
):
    """取得熱門搜尋詞"""
    from app.services.search_optimizer import SearchOptimizer

    try:
        optimizer = SearchOptimizer(db)
        popular = await optimizer.get_popular_searches(limit=limit)

        return {
            "success": True,
            "popular_searches": popular,
            "count": len(popular)
        }

    except Exception as e:
        logger.error(f"取得熱門搜尋失敗: {e}", exc_info=True)
        return {
            "success": False,
            "popular_searches": [],
            "error": str(e)
        }


# ============================================================================
# 公文 CRUD API（POST-only 資安機制）
# ============================================================================

# 注意：DocumentCreateRequest, DocumentUpdateRequest, VALID_DOC_TYPES
# 已統一定義於 app/schemas/document.py，此處透過 import 使用


@router.post(
    "/{document_id}/detail",
    response_model=DocumentResponse,
    summary="取得公文詳情"
)
async def get_document_detail(
    document_id: int,
    db: AsyncSession = Depends(get_async_db),
    current_user: User = Depends(require_auth())
):
    """取得單一公文詳情（POST-only 資安機制，含擴充欄位與權限檢查）"""
    try:
        query = select(OfficialDocument).where(OfficialDocument.id == document_id)
        result = await db.execute(query)
        document = result.scalar_one_or_none()

        if not document:
            from fastapi.responses import JSONResponse
            return JSONResponse(
                status_code=404,
                content={
                    "success": False,
                    "error": {
                        "code": "ERR_NOT_FOUND",
                        "message": f"公文 (ID: {document_id}) 不存在"
                    }
                }
            )

        # 🔒 行級別權限檢查 (RLS)
        if not current_user.is_admin and not current_user.is_superuser:
            # 檢查使用者是否有權限查看此公文
            if document.contract_project_id:
                # 公文有關聯專案，檢查使用者是否為專案成員
                access_check = await db.execute(
                    select(project_user_assignment.c.id).where(
                        and_(
                            project_user_assignment.c.project_id == document.contract_project_id,
                            project_user_assignment.c.user_id == current_user.id,
                            project_user_assignment.c.status.in_(['active', 'Active', None])
                        )
                    ).limit(1)
                )
                if not access_check.scalar_one_or_none():
                    raise ForbiddenException("您沒有權限查看此公文")
            # 無專案關聯的公文視為公開，不需額外檢查

        # 準備擴充欄位
        doc_dict = {k: v for k, v in document.__dict__.items() if not k.startswith('_')}

        # 查詢承攬案件名稱
        if document.contract_project_id:
            project_query = select(ContractProject.project_name).where(
                ContractProject.id == document.contract_project_id
            )
            project_result = await db.execute(project_query)
            doc_dict['contract_project_name'] = project_result.scalar()

        # 查詢機關名稱（2026-01-08 新增）
        if document.sender_agency_id:
            agency_query = select(GovernmentAgency.agency_name).where(
                GovernmentAgency.id == document.sender_agency_id
            )
            agency_result = await db.execute(agency_query)
            doc_dict['sender_agency_name'] = agency_result.scalar()

        if document.receiver_agency_id:
            agency_query = select(GovernmentAgency.agency_name).where(
                GovernmentAgency.id == document.receiver_agency_id
            )
            agency_result = await db.execute(agency_query)
            doc_dict['receiver_agency_name'] = agency_result.scalar()

        # 查詢附件數量
        attachment_count_query = select(func.count(DocumentAttachment.id)).where(
            DocumentAttachment.document_id == document_id
        )
        attachment_result = await db.execute(attachment_count_query)
        doc_dict['attachment_count'] = attachment_result.scalar() or 0

        return DocumentResponse.model_validate(doc_dict)
    except Exception as e:
        logger.error(f"取得公文詳情失敗: {e}", exc_info=True)
        from fastapi.responses import JSONResponse
        return JSONResponse(
            status_code=500,
            content={
                "success": False,
                "error": {
                    "code": "ERR_INTERNAL",
                    "message": f"取得公文詳情失敗: {str(e)}"
                }
            }
        )


@router.post(
    "",
    response_model=DocumentResponse,
    summary="建立公文"
)
async def create_document(
    data: DocumentCreateRequest = Body(...),
    db: AsyncSession = Depends(get_async_db),
    current_user: User = Depends(require_permission("documents:create"))
):
    """
    建立新公文（POST-only 資安機制，含使用者追蹤）

    🔒 權限要求：documents:create
    """
    try:
        create_data = data.model_dump(exclude_unset=True)

        # OfficialDocument 模型的有效欄位（與資料庫 schema 對齊）
        valid_model_fields = {
            'auto_serial', 'doc_number', 'doc_type', 'subject', 'sender', 'receiver',
            'doc_date', 'receive_date', 'send_date', 'status', 'category',
            'delivery_method', 'has_attachment', 'contract_project_id',
            'sender_agency_id', 'receiver_agency_id', 'title', 'cloud_file_link',
            'dispatch_format', 'assignee', 'notes', 'ck_note', 'content'
        }

        # 過濾掉不存在於模型的欄位（避免 TypeError）
        filtered_data = {k: v for k, v in create_data.items() if k in valid_model_fields}

        # 自動產生 auto_serial（若未提供）
        if not filtered_data.get('auto_serial'):
            doc_type = filtered_data.get('doc_type', '收文')
            prefix = 'R' if doc_type == '收文' else 'S'
            # 查詢當前最大流水號
            result = await db.execute(
                select(func.max(OfficialDocument.auto_serial)).where(
                    OfficialDocument.auto_serial.like(f'{prefix}%')
                )
            )
            max_serial = result.scalar_one_or_none()
            if max_serial:
                try:
                    num = int(max_serial[1:]) + 1
                except (ValueError, IndexError):
                    num = 1
            else:
                num = 1
            filtered_data['auto_serial'] = f'{prefix}{num:04d}'

        # 日期欄位需要特別處理：字串轉換為 date 物件
        date_fields = ['doc_date', 'receive_date', 'send_date']
        for field in date_fields:
            if field in filtered_data and isinstance(filtered_data[field], str):
                filtered_data[field] = parse_date_string(filtered_data[field])

        document = OfficialDocument(**filtered_data)
        db.add(document)
        await db.commit()
        await db.refresh(document)

        # 審計日誌（使用 AuditService，自動使用獨立 session，不會污染主交易）
        user_id = current_user.id if current_user else None
        user_name = current_user.username if current_user else "Anonymous"
        logger.info(f"公文 {document.id} 建立 by {user_name}")

        from app.services.audit_service import AuditService
        await AuditService.log_document_change(
            document_id=document.id,
            action="CREATE",
            changes={"created": filtered_data},
            user_id=user_id,
            user_name=user_name,
            source="API"
        )

        return DocumentResponse.model_validate(document)
    except Exception as e:
        await db.rollback()
        logger.error(f"建立公文失敗: {e}", exc_info=True)
        raise


@router.post(
    "/{document_id}/update",
    response_model=DocumentResponse,
    summary="更新公文"
)
async def update_document(
    document_id: int,
    data: DocumentUpdateRequest = Body(...),
    db: AsyncSession = Depends(get_async_db),
    current_user: User = Depends(require_permission("documents:edit"))
):
    """
    更新公文（POST-only 資安機制，含審計日誌與使用者追蹤）

    🔒 權限要求：documents:edit
    🔒 行級別權限：一般使用者只能編輯關聯專案的公文
    """
    try:
        logger.info(f"[更新公文] 開始更新公文 ID: {document_id}")
        logger.debug(f"[更新公文] 收到資料: {data.model_dump()}")

        query = select(OfficialDocument).where(OfficialDocument.id == document_id)
        result = await db.execute(query)
        document = result.scalar_one_or_none()

        if not document:
            raise NotFoundException(resource="公文", resource_id=document_id)

        # 🔒 行級別權限檢查 (RLS) - 非管理員只能編輯關聯專案的公文
        if not current_user.is_admin and not current_user.is_superuser:
            if document.contract_project_id:
                access_check = await db.execute(
                    select(project_user_assignment.c.id).where(
                        and_(
                            project_user_assignment.c.project_id == document.contract_project_id,
                            project_user_assignment.c.user_id == current_user.id,
                            project_user_assignment.c.status.in_(['active', 'Active', None])
                        )
                    ).limit(1)
                )
                if not access_check.scalar_one_or_none():
                    raise ForbiddenException("您沒有權限編輯此公文")

        # 初始化審計保護器，記錄原始資料
        guard = DocumentUpdateGuard(db, document_id)
        original_data = {
            col.name: getattr(document, col.name)
            for col in document.__table__.columns
        }

        update_data = data.model_dump(exclude_unset=True)
        logger.debug(f"[更新公文] 過濾前 update_data: {update_data}")

        # OfficialDocument 模型的有效欄位（與資料庫 schema 對齊）
        valid_model_fields = {
            'auto_serial', 'doc_number', 'doc_type', 'subject', 'sender', 'receiver',
            'doc_date', 'receive_date', 'send_date', 'status', 'category',
            'delivery_method', 'has_attachment', 'contract_project_id',
            'sender_agency_id', 'receiver_agency_id', 'title', 'cloud_file_link',
            'dispatch_format', 'assignee', 'notes', 'ck_note', 'content'
        }

        # 過濾掉不存在於模型的欄位
        update_data = {k: v for k, v in update_data.items() if k in valid_model_fields}
        logger.debug(f"[更新公文] 過濾後 update_data: {update_data}")

        # 日期欄位需要特別處理：字串轉換為 date 物件
        date_fields = ['doc_date', 'receive_date', 'send_date']
        processed_data = {}

        for key, value in update_data.items():
            if value is not None:
                # 處理日期欄位
                if key in date_fields:
                    parsed_date = parse_date_string(value) if isinstance(value, str) else value
                    setattr(document, key, parsed_date)
                    processed_data[key] = parsed_date
                else:
                    setattr(document, key, value)
                    processed_data[key] = value

        # 記錄審計日誌（變更前後比對）
        changes = {}
        for key, new_value in processed_data.items():
            old_value = original_data.get(key)
            if old_value != new_value:
                changes[key] = {"old": str(old_value), "new": str(new_value)}

        # 先提交主要更新操作
        await db.commit()
        await db.refresh(document)

        # 審計日誌和通知（使用統一服務，自動管理獨立 session）
        if changes:
            user_id = current_user.id if current_user else None
            user_name = current_user.username if current_user else "Anonymous"
            logger.info(f"公文 {document_id} 更新 by {user_name}: {list(changes.keys())}")

            # 使用 AuditService（自動使用獨立 session，不會污染主交易）
            from app.services.audit_service import AuditService
            await AuditService.log_document_change(
                document_id=document_id,
                action="UPDATE",
                changes=changes,
                user_id=user_id,
                user_name=user_name,
                source="API"
            )

            # 關鍵欄位變更通知（使用 safe_* 方法，自動使用獨立 session）
            critical_field_names = CRITICAL_FIELDS.get("documents", {})
            for field_key, change_info in changes.items():
                if field_key in critical_field_names:
                    await NotificationService.safe_notify_critical_change(
                        document_id=document_id,
                        field=field_key,
                        old_value=change_info.get("old", ""),
                        new_value=change_info.get("new", ""),
                        user_id=user_id,
                        user_name=user_name,
                        table_name="documents"
                    )

        return DocumentResponse.model_validate(document)
    except NotFoundException:
        raise
    except Exception as e:
        await db.rollback()
        logger.error(f"更新公文失敗: {e}", exc_info=True)
        raise


@router.post(
    "/{document_id}/delete",
    response_model=DeleteResponse,
    summary="刪除公文"
)
async def delete_document(
    document_id: int,
    db: AsyncSession = Depends(get_async_db),
    current_user: User = Depends(require_permission("documents:delete"))
):
    """
    刪除公文（POST-only 資安機制）

    🔒 權限要求：documents:delete
    🔒 行級別權限：一般使用者只能刪除關聯專案的公文

    同步刪除：
    - 公文資料庫記錄
    - 附件資料庫記錄（CASCADE）
    - 實體附件檔案
    - 公文附件資料夾（若為空）
    """
    try:
        # 1. 查詢公文是否存在
        query = select(OfficialDocument).where(OfficialDocument.id == document_id)
        result = await db.execute(query)
        document = result.scalar_one_or_none()

        if not document:
            raise NotFoundException(resource="公文", resource_id=document_id)

        # 🔒 行級別權限檢查 (RLS) - 非管理員只能刪除關聯專案的公文
        if not current_user.is_admin and not current_user.is_superuser:
            if document.contract_project_id:
                access_check = await db.execute(
                    select(project_user_assignment.c.id).where(
                        and_(
                            project_user_assignment.c.project_id == document.contract_project_id,
                            project_user_assignment.c.user_id == current_user.id,
                            project_user_assignment.c.status.in_(['active', 'Active', None])
                        )
                    ).limit(1)
                )
                if not access_check.scalar_one_or_none():
                    raise ForbiddenException("您沒有權限刪除此公文")

        # 2. 查詢關聯的附件記錄（在刪除前取得檔案路徑）
        attachment_query = select(DocumentAttachment).where(
            DocumentAttachment.document_id == document_id
        )
        attachment_result = await db.execute(attachment_query)
        attachments = attachment_result.scalars().all()

        # 3. 收集需要刪除的檔案路徑和資料夾
        file_paths_to_delete = []
        folders_to_check = set()

        for attachment in attachments:
            if attachment.file_path:
                file_paths_to_delete.append(attachment.file_path)
                # 記錄父資料夾路徑（doc_{id} 層級）
                parent_folder = os.path.dirname(attachment.file_path)
                if parent_folder:
                    folders_to_check.add(parent_folder)

        # 4. 記錄公文資訊（在刪除前保存，用於後續審計日誌）
        user_id = current_user.id
        user_name = current_user.username
        doc_number = document.doc_number or ""
        subject = document.subject or ""
        attachments_count = len(attachments)
        logger.info(f"公文 {document_id} 刪除 by {user_name}")

        # 5. 刪除資料庫記錄（CASCADE 會自動刪除 document_attachments）
        await db.delete(document)
        await db.commit()

        # 6. 審計日誌和通知（使用統一服務，自動管理獨立 session）
        from app.services.audit_service import AuditService
        await AuditService.log_document_change(
            document_id=document_id,
            action="DELETE",
            changes={
                "deleted": {
                    "doc_number": doc_number,
                    "subject": subject,
                    "attachments_count": attachments_count
                }
            },
            user_id=user_id,
            user_name=user_name,
            source="API"
        )

        # 公文刪除通知（使用 safe_* 方法，自動使用獨立 session）
        await NotificationService.safe_notify_document_deleted(
            document_id=document_id,
            doc_number=doc_number,
            subject=subject,
            user_id=user_id,
            user_name=user_name
        )

        # 7. 刪除實體檔案（在資料庫成功刪除後執行）
        deleted_files = 0
        file_errors = []

        for file_path in file_paths_to_delete:
            try:
                if os.path.exists(file_path):
                    os.remove(file_path)
                    deleted_files += 1
                    logger.info(f"已刪除附件檔案: {file_path}")
            except Exception as e:
                file_errors.append(f"{file_path}: {str(e)}")
                logger.warning(f"刪除附件檔案失敗: {file_path}, 錯誤: {e}")

        # 8. 嘗試刪除空的公文資料夾（doc_{id}）
        deleted_folders = 0
        for folder in folders_to_check:
            try:
                if os.path.exists(folder) and os.path.isdir(folder):
                    # 只刪除空資料夾
                    if not os.listdir(folder):
                        os.rmdir(folder)
                        deleted_folders += 1
                        logger.info(f"已刪除空資料夾: {folder}")
            except Exception as e:
                logger.warning(f"刪除資料夾失敗: {folder}, 錯誤: {e}")

        # 9. 建構回應訊息
        message = f"公文已刪除"
        if deleted_files > 0:
            message += f"，同步刪除 {deleted_files} 個附件檔案"
        if deleted_folders > 0:
            message += f"，清理 {deleted_folders} 個空資料夾"
        if file_errors:
            message += f"（{len(file_errors)} 個檔案刪除失敗）"

        return DeleteResponse(
            success=True,
            message=message,
            deleted_id=document_id
        )
    except NotFoundException:
        raise
    except Exception as e:
        await db.rollback()
        logger.error(f"刪除公文失敗: {e}", exc_info=True)
        raise


# ============================================================================
# 審計日誌查詢 API
# ============================================================================

@router.post(
    "/audit-logs",
    response_model=AuditLogResponse,
    summary="查詢審計日誌"
)
async def get_audit_logs(
    query: AuditLogQuery = Body(default=AuditLogQuery()),
    db: AsyncSession = Depends(get_async_db)
):
    """
    查詢審計日誌

    支援依公文 ID、操作類型、使用者、日期範圍等條件篩選
    """
    try:
        # 構建查詢條件
        conditions = []
        params = {}

        if query.document_id:
            conditions.append("record_id = :document_id")
            params["document_id"] = query.document_id
        if query.table_name:
            conditions.append("table_name = :table_name")
            params["table_name"] = query.table_name
        if query.action:
            conditions.append("action = :action")
            params["action"] = query.action
        if query.user_id:
            conditions.append("user_id = :user_id")
            params["user_id"] = query.user_id
        if query.is_critical is not None:
            conditions.append("is_critical = :is_critical")
            params["is_critical"] = query.is_critical
        if query.date_from:
            conditions.append("created_at >= :date_from")
            params["date_from"] = query.date_from
        if query.date_to:
            conditions.append("created_at <= :date_to")
            params["date_to"] = query.date_to + " 23:59:59"

        where_clause = " AND ".join(conditions) if conditions else "1=1"

        # 計算總筆數
        count_sql = f"SELECT COUNT(*) FROM audit_logs WHERE {where_clause}"
        count_result = await db.execute(text(count_sql), params)
        total = count_result.scalar() or 0

        # 查詢資料（分頁）
        offset = (query.page - 1) * query.limit
        data_sql = f"""
            SELECT id, table_name, record_id, action, changes,
                   user_id, user_name, source, is_critical,
                   TO_CHAR(created_at, 'YYYY-MM-DD HH24:MI:SS') as created_at
            FROM audit_logs
            WHERE {where_clause}
            ORDER BY created_at DESC
            LIMIT :limit OFFSET :offset
        """
        params["limit"] = query.limit
        params["offset"] = offset

        result = await db.execute(text(data_sql), params)
        rows = result.fetchall()

        items = [
            AuditLogItem(
                id=row.id,
                table_name=row.table_name,
                record_id=row.record_id,
                action=row.action,
                changes=row.changes,
                user_id=row.user_id,
                user_name=row.user_name,
                source=row.source,
                is_critical=row.is_critical or False,
                created_at=row.created_at
            )
            for row in rows
        ]

        total_pages = (total + query.limit - 1) // query.limit

        return AuditLogResponse(
            success=True,
            items=items,
            pagination=PaginationMeta(
                total=total,
                page=query.page,
                limit=query.limit,
                total_pages=total_pages,
                has_next=query.page < total_pages,
                has_prev=query.page > 1
            )
        )
    except Exception as e:
        logger.error(f"查詢審計日誌失敗: {e}", exc_info=True)
        return AuditLogResponse(
            success=False,
            items=[],
            pagination=PaginationMeta(
                total=0, page=1, limit=query.limit,
                total_pages=0, has_next=False, has_prev=False
            )
        )


@router.post(
    "/{document_id}/audit-history",
    response_model=AuditLogResponse,
    summary="查詢公文變更歷史"
)
async def get_document_audit_history(
    document_id: int,
    db: AsyncSession = Depends(get_async_db)
):
    """查詢特定公文的變更歷史記錄"""
    query = AuditLogQuery(document_id=document_id, table_name="documents", limit=50)
    return await get_audit_logs(query, db)


# ============================================================================
# 向後相容：保留 GET 端點（已棄用，將在未來版本移除）
# ============================================================================

@router.post(
    "/integrated-search",
    summary="整合式公文搜尋（已棄用，請改用 POST /list）",
    deprecated=True
)
async def integrated_document_search_legacy(
    skip: int = Query(0, ge=0, description="跳過筆數"),
    limit: int = Query(50, ge=1, le=1000, description="取得筆數"),
    keyword: Optional[str] = Query(None, description="關鍵字搜尋"),
    doc_type: Optional[str] = Query(None, description="公文類型"),
    year: Optional[int] = Query(None, description="年度"),
    status: Optional[str] = Query(None, description="狀態"),
    contract_case: Optional[str] = Query(None, description="承攬案件"),
    sender: Optional[str] = Query(None, description="發文單位"),
    receiver: Optional[str] = Query(None, description="受文單位"),
    doc_date_from: Optional[str] = Query(None, description="公文日期起"),
    doc_date_to: Optional[str] = Query(None, description="公文日期迄"),
    sort_by: Optional[str] = Query("updated_at", description="排序欄位"),
    sort_order: Optional[str] = Query("desc", description="排序順序"),
    db: AsyncSession = Depends(get_async_db)
):
    """
    整合式公文搜尋（已棄用）

    請改用 POST /documents-enhanced/list 端點
    """
    try:
        service = DocumentService(db)

        filters = DocumentFilter(
            keyword=keyword,
            doc_type=doc_type,
            year=year,
            status=status,
            sender=sender,
            receiver=receiver,
            date_from=doc_date_from,
            date_to=doc_date_to,
            sort_by=sort_by,
            sort_order=sort_order
        )

        if contract_case:
            setattr(filters, 'contract_case', contract_case)

        result = await service.get_documents(
            skip=skip,
            limit=limit,
            filters=filters
        )

        return result

    except Exception as e:
        logger.error(f"整合搜尋失敗: {e}", exc_info=True)
        return {"items": [], "total": 0, "page": 1, "limit": limit, "total_pages": 0}


@router.post("/document-years", summary="取得年度選項（已棄用）", deprecated=True)
async def get_document_years_legacy(db: AsyncSession = Depends(get_async_db)):
    """已棄用，請改用 POST /documents-enhanced/years"""
    return await get_document_years(db)


# ============================================================================
# 專案關聯公文 API（自動關聯機制）
# ============================================================================

@router.post(
    "/by-project",
    response_model=DocumentListResponse,
    summary="查詢專案關聯公文",
    description="根據 project_id 自動查詢該專案的所有關聯公文"
)
async def get_documents_by_project(
    query: ProjectDocumentsQuery = Body(...),
    db: AsyncSession = Depends(get_async_db)
):
    """
    根據專案 ID 查詢關聯公文（自動關聯機制）

    關聯邏輯：
    依據 documents.contract_project_id = project_id 查詢

    回傳該專案的所有公文紀錄
    """
    try:
        # 構建查詢條件：依 contract_project_id 匹配
        doc_query = select(OfficialDocument).where(
            OfficialDocument.contract_project_id == query.project_id
        ).order_by(
            OfficialDocument.doc_date.desc(),
            OfficialDocument.id.desc()
        )

        # 計算總數
        count_query = select(func.count()).select_from(doc_query.subquery())
        total_result = await db.execute(count_query)
        total = total_result.scalar() or 0

        # 分頁
        skip = (query.page - 1) * query.limit
        doc_query = doc_query.offset(skip).limit(query.limit)

        result = await db.execute(doc_query)
        documents = result.scalars().all()

        # 查詢專案名稱（所有文件共用同一個專案）
        project_name = None
        if query.project_id:
            project_query = select(ContractProject.project_name).where(
                ContractProject.id == query.project_id
            )
            project_result = await db.execute(project_query)
            project_name = project_result.scalar()

        # 查詢專案承辦同仁（使用 project_user_assignment 關聯表）
        assigned_staff = []
        if query.project_id:
            # 從關聯表查詢專案成員，並 JOIN users 表獲取姓名
            staff_query = (
                select(
                    project_user_assignment.c.user_id,
                    project_user_assignment.c.role,
                    User.full_name,
                    User.username
                )
                .join(User, User.id == project_user_assignment.c.user_id)
                .where(
                    project_user_assignment.c.project_id == query.project_id,
                    project_user_assignment.c.status == 'active'
                )
            )
            staff_result = await db.execute(staff_query)
            staff_rows = staff_result.all()
            assigned_staff = [
                StaffInfo(
                    user_id=row.user_id,
                    name=row.full_name or row.username or f"User {row.user_id}",
                    role=row.role or "member"
                )
                for row in staff_rows
            ]

        # 轉換為回應格式（包含專案關聯資訊）
        response_items = []
        for doc in documents:
            try:
                doc_dict = {
                    **{k: v for k, v in doc.__dict__.items() if not k.startswith('_')},
                    'contract_project_name': project_name,
                    'assigned_staff': assigned_staff
                }
                response_items.append(DocumentResponse.model_validate(doc_dict))
            except Exception as e:
                logger.warning(f"轉換公文資料失敗: {e}")
                continue

        return DocumentListResponse(
            items=response_items,
            pagination=PaginationMeta.create(
                total=total,
                page=query.page,
                limit=query.limit
            )
        )

    except Exception as e:
        logger.error(f"查詢專案關聯公文失敗: {e}", exc_info=True)
        return DocumentListResponse(
            items=[],
            pagination=PaginationMeta.create(total=0, page=1, limit=query.limit)
        )


# ============================================================================
# 公文匯出 API
# ============================================================================

@router.post("/export", summary="匯出公文資料")
async def export_documents(
    query: DocumentExportQuery = Body(...),
    db: AsyncSession = Depends(get_async_db)
):
    """
    匯出公文資料為 CSV 格式

    支援功能:
    - 依指定 ID 列表匯出
    - 依類別/年度篩選後匯出
    - 若未指定條件則匯出全部
    """
    try:
        # 構建查詢
        doc_query = select(OfficialDocument).options(
            selectinload(OfficialDocument.contract_project)
        )

        # 篩選條件
        conditions = []
        if query.document_ids:
            conditions.append(OfficialDocument.id.in_(query.document_ids))
        if query.category:
            conditions.append(OfficialDocument.category == query.category)
        if query.year:
            conditions.append(func.extract('year', OfficialDocument.doc_date) == query.year)

        if conditions:
            doc_query = doc_query.where(and_(*conditions))

        doc_query = doc_query.order_by(OfficialDocument.doc_date.desc())

        result = await db.execute(doc_query)
        documents = result.scalars().all()

        # 產生 CSV
        output = io.StringIO()
        writer = csv.writer(output, quoting=csv.QUOTE_ALL)

        # 寫入標題列
        headers = [
            '序號', '公文文號', '主旨', '類別', '發文/收文日期',
            '發文單位', '受文單位', '承攬案件', '狀態', '備註'
        ]
        writer.writerow(headers)

        # 寫入資料列
        for idx, doc in enumerate(documents, start=1):
            contract_case_name = ""
            if doc.contract_project:
                contract_case_name = doc.contract_project.project_name or ""

            row = [
                doc.auto_serial or idx,
                doc.doc_number or "",
                doc.subject or "",
                doc.category or "",
                str(doc.doc_date) if doc.doc_date else "",
                doc.sender or "",
                doc.receiver or "",
                contract_case_name,
                doc.status or "",
                doc.notes or ""
            ]
            writer.writerow(row)

        # 重置游標位置
        output.seek(0)

        # 回傳 CSV 檔案
        from datetime import datetime
        filename = f"documents_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

        return StreamingResponse(
            iter(['\ufeff' + output.getvalue()]),  # BOM for Excel UTF-8
            media_type="text/csv; charset=utf-8",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )

    except Exception as e:
        logger.error(f"匯出公文失敗: {e}", exc_info=True)
        from fastapi import HTTPException
        raise HTTPException(status_code=500, detail=f"匯出公文失敗: {str(e)}")


# ============================================================================
# Excel 匯出端點
# ============================================================================

@router.post("/export/excel", summary="匯出公文為 Excel")
async def export_documents_excel(
    request: ExcelExportRequest = Body(default=ExcelExportRequest()),
    db: AsyncSession = Depends(get_async_db)
):
    """
    匯出公文資料為 Excel 格式 (.xlsx)

    檔名格式: CK公文YYYYMMDD.xlsx

    支援功能:
    - 依指定 ID 列表匯出
    - 依類別/年度/關鍵字/狀態篩選後匯出
    - 若未指定條件則匯出全部（無筆數限制）

    流水號說明:
    - S 開頭: 發文 (Send)
    - R 開頭: 收文 (Receive)
    """
    try:
        import pandas as pd
        from io import BytesIO
        from datetime import datetime

        # 構建查詢 - 無筆數限制，匯出全部符合條件的資料
        doc_query = select(OfficialDocument).options(
            selectinload(OfficialDocument.contract_project),
            selectinload(OfficialDocument.sender_agency),
            selectinload(OfficialDocument.receiver_agency),
            selectinload(OfficialDocument.attachments)  # 載入附件以統計數量
        )

        # 篩選條件
        conditions = []
        if request.document_ids:
            conditions.append(OfficialDocument.id.in_(request.document_ids))
        if request.category:
            conditions.append(OfficialDocument.category == request.category)
        if request.year:
            conditions.append(func.extract('year', OfficialDocument.doc_date) == request.year)
        if request.status:
            conditions.append(OfficialDocument.status == request.status)
        if request.keyword:
            keyword = f"%{request.keyword}%"
            conditions.append(
                or_(
                    OfficialDocument.subject.ilike(keyword),
                    OfficialDocument.doc_number.ilike(keyword),
                    OfficialDocument.sender.ilike(keyword),
                    OfficialDocument.receiver.ilike(keyword),
                    OfficialDocument.content.ilike(keyword),
                    OfficialDocument.notes.ilike(keyword)
                )
            )
        if request.contract_case:
            # contract_case 需要透過關聯查詢 ContractProject.project_name
            contract_case_keyword = f"%{request.contract_case}%"
            doc_query = doc_query.outerjoin(ContractProject, OfficialDocument.contract_project_id == ContractProject.id)
            conditions.append(ContractProject.project_name.ilike(contract_case_keyword))
        if request.sender:
            sender_keyword = f"%{request.sender}%"
            conditions.append(OfficialDocument.sender.ilike(sender_keyword))
        if request.receiver:
            receiver_keyword = f"%{request.receiver}%"
            conditions.append(OfficialDocument.receiver.ilike(receiver_keyword))

        if conditions:
            doc_query = doc_query.where(and_(*conditions))

        # 排序：依公文日期降序
        doc_query = doc_query.order_by(OfficialDocument.doc_date.desc())

        result = await db.execute(doc_query)
        documents = result.scalars().all()

        if not documents:
            from fastapi import HTTPException
            raise HTTPException(status_code=404, detail="沒有符合條件的公文可供匯出")

        # 轉換為 DataFrame - 精簡欄位
        data = []

        def clean_agency_name(raw_text: str, agency_name: str = "") -> str:
            """清理機關名稱，移除代碼，只保留中文名稱"""
            # 優先使用關聯表的機關名稱
            if agency_name:
                return agency_name
            if not raw_text:
                return ""
            # 移除常見代碼格式：
            # "EB50819619 乾坤測繪科技有限公司" → "乾坤測繪科技有限公司"
            # "376470600A (彰化縣和美地政事務所)" → "彰化縣和美地政事務所"
            # "376470600A\n(彰化縣和美地政事務所)" → "彰化縣和美地政事務所"
            import re
            text = raw_text.strip()
            # 移除括號內的內容提取為主名稱
            paren_match = re.search(r'[（(]([^)）]+)[)）]', text)
            if paren_match:
                return paren_match.group(1).strip()
            # 移除開頭的英數代碼（如 EB50819619、376470600A）
            text = re.sub(r'^[A-Za-z0-9]+\s*', '', text)
            return text.strip()

        def get_valid_doc_type(doc_type: str) -> str:
            """取得有效的公文類型

            有效值: 函、開會通知單、會勘通知單、書函等
            注意: 2026-01-07 已修復 8 筆錯誤資料（doc_type 誤為收文/發文）
            """
            # 保留防護：若仍有錯誤值則過濾
            if doc_type in ['收文', '發文']:
                return ""
            return doc_type or ""

        for doc in documents:
            # 取得關聯資料
            contract_case_name = ""
            if doc.contract_project:
                contract_case_name = doc.contract_project.project_name or ""

            sender_agency_name = ""
            if doc.sender_agency:
                sender_agency_name = doc.sender_agency.agency_name or ""

            receiver_agency_name = ""
            if doc.receiver_agency:
                receiver_agency_name = doc.receiver_agency.agency_name or ""

            # 統計附件數量
            attachment_count = len(doc.attachments) if doc.attachments else 0
            attachment_text = f"{attachment_count} 個附件" if attachment_count > 0 else "無"

            # 欄位順序依需求調整（公文ID對應附件資料夾 doc_{id}）
            data.append({
                "公文ID": doc.id,
                "流水號": doc.auto_serial or "",
                "發文形式": doc.delivery_method or "",
                "類別": doc.category or "",
                "公文類型": get_valid_doc_type(doc.doc_type),
                "公文字號": doc.doc_number or "",
                "主旨": doc.subject or "",
                "說明": getattr(doc, 'content', '') or "",
                "公文日期": str(doc.doc_date) if doc.doc_date else "",
                "收文日期": str(doc.receive_date) if doc.receive_date else "",
                "發文日期": str(doc.send_date) if doc.send_date else "",
                "發文單位": clean_agency_name(doc.sender or "", sender_agency_name),
                "受文單位": clean_agency_name(doc.receiver or "", receiver_agency_name),
                "附件紀錄": attachment_text,
                "備註": getattr(doc, 'notes', '') or "",
                "狀態": doc.status or "",
                "承攬案件": contract_case_name,
                "建立時間": str(doc.created_at) if doc.created_at else "",
                "更新時間": str(doc.updated_at) if doc.updated_at else "",
            })

        df = pd.DataFrame(data)

        # 產生 Excel
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='公文清單')

            # 取得工作表
            worksheet = writer.sheets['公文清單']

            # 表頭樣式：粗體 + 淺藍色背景
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            header_font = Font(bold=True, color="000000")
            header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # 套用表頭樣式
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            # 資料列樣式
            data_alignment = Alignment(vertical="center", wrap_text=True)
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                for cell in row:
                    cell.alignment = data_alignment
                    cell.border = thin_border

            # 調整欄位寬度
            for idx, col in enumerate(df.columns):
                max_length = max(
                    df[col].astype(str).map(len).max(),
                    len(col)
                ) + 2
                # 限制最大寬度
                max_length = min(max_length, 60)
                # Excel 欄位名稱 A-Z, AA-AZ...
                col_letter = chr(65 + idx) if idx < 26 else f"A{chr(65 + idx - 26)}"
                worksheet.column_dimensions[col_letter].width = max_length

            # 凍結表頭列
            worksheet.freeze_panes = 'A2'

            # 新增統計摘要工作表
            summary_data = {
                "項目": [
                    "匯出時間",
                    "公文總數",
                    "收文數量",
                    "發文數量",
                    "有附件公文",
                    "已指派案件",
                    "最早公文日期",
                    "最新公文日期"
                ],
                "數值": [
                    datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    str(len(documents)),
                    str(sum(1 for d in documents if d.category == '收文')),
                    str(sum(1 for d in documents if d.category == '發文')),
                    str(sum(1 for d in documents if d.attachments and len(d.attachments) > 0)),
                    str(sum(1 for d in documents if d.contract_project_id)),
                    str(min((d.doc_date for d in documents if d.doc_date), default="")) or "N/A",
                    str(max((d.doc_date for d in documents if d.doc_date), default="")) or "N/A"
                ]
            }
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, index=False, sheet_name='統計摘要')

            # 統計摘要工作表樣式
            summary_ws = writer.sheets['統計摘要']
            for cell in summary_ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            for row in summary_ws.iter_rows(min_row=2, max_row=summary_ws.max_row):
                for cell in row:
                    cell.alignment = data_alignment
                    cell.border = thin_border
            summary_ws.column_dimensions['A'].width = 20
            summary_ws.column_dimensions['B'].width = 30

        output.seek(0)

        # 產生檔名: 乾坤測繪公文總表YYYYMMDD.xlsx
        from urllib.parse import quote
        date_str = datetime.now().strftime('%Y%m%d')
        filename_cn = f"乾坤測繪公文總表{date_str}.xlsx"
        filename_encoded = quote(filename_cn)

        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{filename_encoded}",
                "X-Content-Type-Options": "nosniff",
                "Cache-Control": "no-cache, no-store, must-revalidate",
            }
        )

    except Exception as e:
        logger.error(f"匯出 Excel 失敗: {e}", exc_info=True)
        from fastapi import HTTPException
        raise HTTPException(status_code=500, detail=f"匯出 Excel 失敗: {str(e)}")


# ============================================================================
# Excel 匯入（手動公文匯入）
# ============================================================================

@router.post("/import/excel/preview", summary="Excel 匯入預覽")
async def preview_excel_import(
    file: UploadFile = File(..., description="要預覽的 Excel 檔案（.xlsx）"),
    preview_rows: int = Query(default=10, ge=1, le=50, description="預覽筆數"),
    check_duplicates: bool = Query(default=True, description="是否檢查資料庫重複"),
    db: AsyncSession = Depends(get_async_db)
):
    """
    預覽 Excel 檔案內容（不執行匯入）

    功能：
    - 顯示前 N 筆資料預覽
    - 驗證欄位格式
    - 標示可能的問題（重複、缺欄位等）
    - 檢查資料庫中已存在的公文字號
    - 統計預計新增/更新筆數

    使用情境：
    - 使用者上傳檔案後，先預覽確認再正式匯入
    """
    from fastapi import HTTPException

    if not file.filename:
        raise HTTPException(status_code=400, detail="未提供檔案")

    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=400,
            detail="檔案格式不正確，僅支援 Excel 檔案（.xlsx, .xls）"
        )

    try:
        file_content = await file.read()
        filename = file.filename

        logger.info(f"Excel 匯入預覽: {filename}, 大小: {len(file_content)} bytes")

        from app.services.excel_import_service import ExcelImportService
        import_service = ExcelImportService(db)
        result = await import_service.preview_excel(
            file_content, filename, preview_rows, check_duplicates
        )

        return result

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Excel 預覽失敗: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"預覽失敗: {str(e)}")


@router.post("/import/excel", summary="手動公文匯入（Excel）")
async def import_documents_excel(
    file: UploadFile = File(..., description="要匯入的 Excel 檔案（.xlsx）"),
    db: AsyncSession = Depends(get_async_db)
):
    """
    從 Excel 檔案匯入公文資料（手動公文匯入）

    適用情境：
    - 紙本郵寄紀錄
    - 手動輸入的公文資料
    - 匯出後修改再匯入

    匯入規則：
    - 公文ID 有值：更新現有資料
    - 公文ID 空白：新增資料（自動生成流水號）
    - 必填欄位：公文字號、主旨、類別

    與「電子公文檔匯入」(CSV) 的差異：
    - CSV 匯入：電子公文系統匯出的固定格式
    - Excel 匯入：本系統匯出格式，支援新增/更新
    """
    from fastapi import HTTPException

    # 驗證檔案格式
    if not file.filename:
        raise HTTPException(status_code=400, detail="未提供檔案")

    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=400,
            detail="檔案格式不正確，僅支援 Excel 檔案（.xlsx, .xls）"
        )

    try:
        # 讀取檔案內容
        file_content = await file.read()
        filename = file.filename

        logger.info(f"開始 Excel 匯入: {filename}, 大小: {len(file_content)} bytes")

        # 使用 ExcelImportService 處理
        from app.services.excel_import_service import ExcelImportService
        import_service = ExcelImportService(db)
        result = await import_service.import_from_excel(file_content, filename)

        return result

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Excel 匯入失敗: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Excel 匯入失敗: {str(e)}")


@router.post("/import/excel/template", summary="下載 Excel 匯入範本")
async def download_excel_template():
    """
    下載 Excel 匯入範本

    範本包含：
    - 標題列（欄位名稱）
    - 範例資料（1-2 筆）
    - 欄位說明
    """
    try:
        import pandas as pd
        from io import BytesIO
        from urllib.parse import quote

        # 建立範本資料（欄位順序與匯出一致：19 欄）
        template_data = [
            {
                "公文ID": "",  # 空白=新增
                "流水號": "",  # 系統自動生成
                "發文形式": "紙本郵寄",
                "類別": "收文",
                "公文類型": "函",
                "公文字號": "XX字第1140000001號",
                "主旨": "（請輸入公文主旨）",
                "說明": "（請輸入公文內容說明）",
                "公文日期": "2026-01-07",
                "收文日期": "2026-01-07",
                "發文日期": "",
                "發文單位": "○○單位",
                "受文單位": "乾坤測繪科技有限公司",
                "附件紀錄": "",  # 僅供參考，匯入忽略
                "備註": "",
                "狀態": "active",
                "承攬案件": "",
                "建立時間": "",  # 系統自動
                "更新時間": "",  # 系統自動
            }
        ]

        df = pd.DataFrame(template_data)

        # 產生 Excel
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='公文匯入')

            # 調整欄位寬度
            worksheet = writer.sheets['公文匯入']
            for idx, col in enumerate(df.columns):
                col_letter = chr(65 + idx) if idx < 26 else f"A{chr(65 + idx - 26)}"
                worksheet.column_dimensions[col_letter].width = 15

        output.seek(0)

        filename_cn = "公文匯入範本.xlsx"
        filename_encoded = quote(filename_cn)

        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{filename_encoded}"
            }
        )

    except Exception as e:
        logger.error(f"下載範本失敗: {e}", exc_info=True)
        from fastapi import HTTPException
        raise HTTPException(status_code=500, detail=f"下載範本失敗: {str(e)}")
