"""
派工紀錄 Excel 增強匯入服務 (v2.0)

從「分派案件紀錄表」主表 Excel 讀取：
1. 價金資料 (7 種作業類別的派工日期/金額 + 彙總)
2. 公文原始文號 (機關來文 Z 欄 + 公司發文 AB 欄，不過度解析)

以「項次」為索引對應現有派工單 (項次 N → 112年_派工單號{N:03d})。

v2.0: 全面遷移至 Repository 模式 (消除 8 個直接 db.execute)
v1.1: 解析工具函數提取至 dispatch_document_parser.py
"""
import re
import logging
from datetime import date
from typing import Optional, Dict, Any, List
from io import BytesIO

import openpyxl
from sqlalchemy.ext.asyncio import AsyncSession

from app.extended.models.taoyuan import (
    TaoyuanDispatchOrder,
    TaoyuanContractPayment,
    TaoyuanDispatchDocumentLink,
)
from app.extended.models.document import OfficialDocument
from app.repositories.document_repository import DocumentRepository
from app.repositories.taoyuan.dispatch_order_repository import DispatchOrderRepository
from app.repositories.taoyuan.dispatch_doc_link_repository import DispatchDocLinkRepository
from app.repositories.taoyuan.payment_repository import PaymentRepository

# 從解析器模組匯入（向後相容：外部可繼續從本模組 import）
from app.services.taoyuan.dispatch_document_parser import (  # noqa: F401
    AGENCY_MAP,
    COMPANY_NAME,
    parse_roc_date,
    parse_sequence_no,
    parse_amount,
    safe_cell,
    parse_doc_line,
)

logger = logging.getLogger(__name__)


# ── 主服務 ────────────────────────────────────────────────────

class DispatchEnrichmentService:
    """從主表 Excel 增強匯入價金 + 公文原始值"""

    # 主表欄位索引 (0-based, values_only=True)
    COL_SEQ = 0           # A: 項次
    # 7 種作業類別 (日期/金額 pairs)
    WORK_COLS = [
        (2,  3,  '01'),  # C/D: 土地改良物查估
        (4,  5,  '02'),  # E/F: 土地協議市價查估
        (6,  7,  '03'),  # G/H: 土地徵收市價查估
        (8,  9,  '04'),  # I/J: 相關計畫書製作
        (10, 11, '05'),  # K/L: 測量作業
        (12, 13, '06'),  # M/N: 樁位測釘費
        (14, 15, '07'),  # O/P: 辦理教育訓練
    ]
    COL_ACCEPTANCE = 17   # R: 完成驗收日期
    COL_CURRENT = 18      # S: 本次派工金額
    COL_CUMUL = 19        # T: 累進派工金額
    COL_REMAIN = 20       # U: 剩餘金額
    COL_AGENCY = 25       # Z: 機關來文
    COL_COMPANY = 27      # AB: 公司發文

    def __init__(self, db: AsyncSession):
        self.db = db
        self._dispatch_repo = DispatchOrderRepository(db)
        self._doc_repo = DocumentRepository(db)
        self._doc_link_repo = DispatchDocLinkRepository(db)
        self._payment_repo = PaymentRepository(db)

    async def enrich_from_master_excel(
        self,
        file_content: bytes,
        dispatch_no_prefix: str = "112年_派工單號",
        data_start_row: int = 4,
        sheet_name: str = "派工單",
    ) -> Dict[str, Any]:
        """
        從主表 Excel 增強匯入。

        Args:
            file_content: Excel 檔案 bytes
            dispatch_no_prefix: 派工單號前綴
            data_start_row: 資料起始行 (1-based)
            sheet_name: Sheet 名稱

        Returns:
            {total_rows, matched, payment_created, payment_updated,
             doc_updated, skipped, errors, message}
        """
        wb = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]

        # 預載 dispatch_no → id 映射 (委派至 Repository)
        dispatch_map = await self._dispatch_repo.build_dispatch_no_map()

        stats: Dict[str, Any] = {
            'total_rows': 0,
            'matched': 0,
            'payment_created': 0,
            'payment_updated': 0,
            'doc_updated': 0,
            'skipped': 0,
            'errors': [],
        }

        for row in ws.iter_rows(min_row=data_start_row, max_row=ws.max_row, values_only=True):
            seq = parse_sequence_no(safe_cell(row, self.COL_SEQ))
            if seq is None:
                continue

            stats['total_rows'] += 1
            dispatch_no = f"{dispatch_no_prefix}{seq:03d}"
            dispatch_info = dispatch_map.get(dispatch_no)

            if not dispatch_info:
                stats['skipped'] += 1
                stats['errors'].append({
                    'seq': seq,
                    'dispatch_no': dispatch_no,
                    'reason': '未找到派工單',
                })
                continue

            stats['matched'] += 1
            dispatch_id = dispatch_info['id']

            # ① 更新公文原始值 (Z/AB 欄，不過度解析，直接存原始值)
            agency_raw = self._extract_raw_text(safe_cell(row, self.COL_AGENCY))
            company_raw = self._extract_raw_text(safe_cell(row, self.COL_COMPANY))

            if agency_raw or company_raw:
                update_vals = {}
                if agency_raw:
                    update_vals['agency_doc_number_raw'] = agency_raw[:500]
                if company_raw:
                    update_vals['company_doc_number_raw'] = company_raw[:500]

                await self._dispatch_repo.update(dispatch_id, update_vals, auto_commit=False)
                stats['doc_updated'] += 1

            # ② 建立/更新價金記錄 (C-P + S/T/U + R)
            payment_data = self._extract_payment(row)
            if payment_data:
                result = await self._upsert_payment(dispatch_id, payment_data)
                if result == 'created':
                    stats['payment_created'] += 1
                elif result == 'updated':
                    stats['payment_updated'] += 1

        await self.db.commit()

        stats['message'] = (
            f"增強匯入完成：掃描 {stats['total_rows']} 行，"
            f"匹配 {stats['matched']} 筆派工單，"
            f"更新公文 {stats['doc_updated']} 筆，"
            f"價金新增 {stats['payment_created']} / 更新 {stats['payment_updated']} 筆"
        )
        logger.info(stats['message'])
        return stats

    # ── 公文 Stub 反建 ─────────────────────────────────────────

    async def create_document_stubs(
        self,
        contract_project_id: int,
    ) -> Dict[str, Any]:
        """
        從派工單已存的原始文號反建 OfficialDocument Stub + 自動關聯。

        不需要原始 Excel，直接從 DB 的 agency_doc_number_raw / company_doc_number_raw 解析。

        Args:
            contract_project_id: 承攬案件 ID

        Returns:
            {total_dispatches, total_lines, docs_created, docs_existing,
             links_created, errors, message}
        """
        # ① 查詢該案件有原始文號的派工單 (委派至 Repository)
        dispatches = await self._dispatch_repo.get_by_project(contract_project_id)

        # ② 預載現有文號 → id 映射 (委派至 Repository)
        existing_docs = await self._doc_repo.build_doc_number_map()

        # ③ 計算下一個 auto_serial 起始值 (委派至 Repository)
        next_r_serial = await self._get_max_serial('R') + 1
        next_s_serial = await self._get_max_serial('S') + 1

        stats: Dict[str, Any] = {
            'total_dispatches': len(dispatches),
            'total_lines': 0,
            'docs_created': 0,
            'docs_existing': 0,
            'links_created': 0,
            'parse_skipped': 0,
            'errors': [],
        }

        for dispatch in dispatches:
            # 處理機關來文 (收文)
            if dispatch.agency_doc_number_raw:
                next_r_serial = await self._process_raw_docs(
                    dispatch=dispatch,
                    raw_text=dispatch.agency_doc_number_raw,
                    is_incoming=True,
                    existing_docs=existing_docs,
                    serial_counter=next_r_serial,
                    contract_project_id=contract_project_id,
                    stats=stats,
                )

            # 處理公司發文 (發文)
            if dispatch.company_doc_number_raw:
                next_s_serial = await self._process_raw_docs(
                    dispatch=dispatch,
                    raw_text=dispatch.company_doc_number_raw,
                    is_incoming=False,
                    existing_docs=existing_docs,
                    serial_counter=next_s_serial,
                    contract_project_id=contract_project_id,
                    stats=stats,
                )

        await self.db.commit()

        stats['message'] = (
            f"公文 Stub 建立完成：掃描 {stats['total_dispatches']} 筆派工單，"
            f"解析 {stats['total_lines']} 行文號，"
            f"新建公文 {stats['docs_created']} 筆，"
            f"已存在 {stats['docs_existing']} 筆，"
            f"建立關聯 {stats['links_created']} 筆"
        )
        logger.info(stats['message'])
        return stats

    async def _process_raw_docs(
        self,
        dispatch: TaoyuanDispatchOrder,
        raw_text: str,
        is_incoming: bool,
        existing_docs: Dict[str, int],
        serial_counter: int,
        contract_project_id: int,
        stats: Dict[str, Any],
    ) -> int:
        """處理一個派工單的原始文號，建立 Stub + Link。回傳更新後的 serial_counter。"""
        lines = raw_text.split('\n')
        link_type = 'agency_incoming' if is_incoming else 'company_outgoing'
        first_doc_id = None

        for line in lines:
            parsed = parse_doc_line(line)
            if not parsed:
                if line.strip() and not line.strip().startswith('('):
                    stats['parse_skipped'] += 1
                continue

            stats['total_lines'] += 1
            doc_number = parsed['doc_number']

            # 找或建公文
            doc_id = existing_docs.get(doc_number)
            if doc_id:
                stats['docs_existing'] += 1
            else:
                # 建立 Stub
                prefix = 'R' if is_incoming else 'S'
                auto_serial = f"{prefix}{serial_counter:04d}"
                serial_counter += 1

                stub = OfficialDocument(
                    auto_serial=auto_serial,
                    doc_number=doc_number,
                    doc_type='收文' if is_incoming else '發文',
                    category='收文' if is_incoming else '發文',
                    doc_date=parsed['doc_date'],
                    subject=f"派工關聯 — {dispatch.project_name or dispatch.dispatch_no}",
                    sender=parsed['sender'] or ('桃園市政府工務局' if is_incoming else COMPANY_NAME),
                    receiver=COMPANY_NAME if is_incoming else '桃園市政府工務局',
                    status='已辦畢',
                    contract_project_id=contract_project_id,
                )
                self.db.add(stub)
                await self.db.flush()
                doc_id = stub.id
                existing_docs[doc_number] = doc_id
                stats['docs_created'] += 1

            # 建立關聯 (冪等) — 委派至 Repository
            existing_link = await self._doc_link_repo.find_dispatch_document_link(
                dispatch.id, doc_id
            )
            if not existing_link:
                link = TaoyuanDispatchDocumentLink(
                    dispatch_order_id=dispatch.id,
                    document_id=doc_id,
                    link_type=link_type,
                )
                self.db.add(link)
                await self.db.flush()
                stats['links_created'] += 1

            # 第一筆設為向下相容 FK
            if first_doc_id is None:
                first_doc_id = doc_id

        # 更新向下相容 FK — 委派至 Repository
        if first_doc_id:
            fk_field = 'agency_doc_id' if is_incoming else 'company_doc_id'
            await self._dispatch_repo.update(
                dispatch.id, {fk_field: first_doc_id}, auto_commit=False
            )

        return serial_counter

    async def _get_max_serial(self, prefix: str) -> int:
        """取得指定前綴的最大流水序號數字 (委派至 Repository)"""
        max_serial = await self._doc_repo.get_max_serial_by_prefix(prefix)
        if max_serial:
            m = re.search(r'(\d+)$', max_serial)
            return int(m.group(1)) if m else 0
        return 0

    # ── 私有方法 ──────────────────────────────────────────────

    @staticmethod
    def _extract_raw_text(cell_value: Any) -> Optional[str]:
        """取得儲存格的原始文字，無值回傳 None"""
        if cell_value is None:
            return None
        text = str(cell_value).strip()
        return text if text else None

    def _extract_payment(self, row: tuple) -> Optional[Dict[str, Any]]:
        """從 Excel 行提取價金資料，無任何有效值回傳 None"""
        data: Dict[str, Any] = {}

        # 7 work types
        for date_col, amt_col, suffix in self.WORK_COLS:
            d = parse_roc_date(safe_cell(row, date_col))
            a = parse_amount(safe_cell(row, amt_col))
            if d is not None:
                data[f'work_{suffix}_date'] = d
            if a is not None:
                data[f'work_{suffix}_amount'] = a

        # 彙總欄位
        current = parse_amount(safe_cell(row, self.COL_CURRENT))
        cumul = parse_amount(safe_cell(row, self.COL_CUMUL))
        remain = parse_amount(safe_cell(row, self.COL_REMAIN))
        accept = parse_roc_date(safe_cell(row, self.COL_ACCEPTANCE))

        if current is not None:
            data['current_amount'] = current
        if cumul is not None:
            data['cumulative_amount'] = cumul
        if remain is not None:
            data['remaining_amount'] = remain
        if accept is not None:
            data['acceptance_date'] = accept

        # 至少有一個非 None 且非零的值才建立
        has_meaningful = any(
            v is not None and v != 0
            for v in data.values()
            if not isinstance(v, date)
        ) or any(isinstance(v, date) for v in data.values())

        return data if has_meaningful else None

    async def _upsert_payment(
        self, dispatch_id: int, data: Dict[str, Any]
    ) -> str:
        """建立或更新價金記錄 (冪等，flush 確保可見性)"""
        await self.db.flush()
        existing = await self._payment_repo.get_by_dispatch_order(dispatch_id)

        if existing:
            for key, val in data.items():
                setattr(existing, key, val)
            return 'updated'
        else:
            payment = TaoyuanContractPayment(
                dispatch_order_id=dispatch_id,
                **data,
            )
            self.db.add(payment)
            await self.db.flush()
            return 'created'
