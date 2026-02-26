"""
DispatchExportService - 派工單總表多工作表 Excel 匯出服務

產生 5 個工作表的 Excel 總表匯出：
1. 派工總表 - 每張派工單一列摘要
2. 作業紀錄明細 - 跨派工單所有作業歷程
3. 公文對照矩陣 - 來文/覆文配對
4. 契金摘要 - 各派工單 7 項作業金額與彙總
5. 統計摘要 - 匯出範圍 key-value 統計

@version 1.0.0
@date 2026-02-25
"""

import logging
from datetime import datetime, date
from io import BytesIO
from typing import Optional, List, Dict, Any

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select, and_, or_
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.extended.models import (
    TaoyuanDispatchOrder,
    TaoyuanDispatchDocumentLink,
    TaoyuanContractPayment,
    TaoyuanWorkRecord,
)
from app.utils.doc_helpers import is_outgoing_doc_number

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# 常數：作業類別 / 狀態 / 契金工種 標籤
# ---------------------------------------------------------------------------

WORK_CATEGORY_LABELS: Dict[str, str] = {
    'dispatch_notice': '派工通知',
    'work_result': '作業成果',
    'meeting_notice': '會議通知',
    'meeting_record': '會議紀錄',
    'survey_notice': '查估通知',
    'survey_record': '查估紀錄',
    'other': '其他',
}

STATUS_LABELS: Dict[str, str] = {
    'pending': '待處理',
    'in_progress': '進行中',
    'completed': '已完成',
    'overdue': '逾期',
    'on_hold': '暫緩',
}

PAYMENT_WORK_LABELS: List[str] = [
    '01.地上物查估',
    '02.土地協議市價查估',
    '03.土地徵收市價查估',
    '04.相關計畫書製作',
    '05.測量作業',
    '06.樁位測釘作業',
    '07.辦理教育訓練',
]

# ---------------------------------------------------------------------------
# 樣式常數
# ---------------------------------------------------------------------------

_HEADER_FILL = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
_HEADER_FONT = Font(bold=True)
_HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)
_THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)


MAX_EXPORT_ROWS = 2000


class DispatchExportService:
    """派工單總表 Excel 匯出服務"""

    def __init__(self, db: AsyncSession):
        self.db = db

    # =========================================================================
    # 公開 API
    # =========================================================================

    async def export_master_matrix(
        self,
        contract_project_id: Optional[int] = None,
        work_type: Optional[str] = None,
        search: Optional[str] = None,
    ) -> BytesIO:
        """Export all matching dispatch orders as a multi-sheet Excel file.

        Args:
            contract_project_id: 篩選特定承攬案件
            work_type: 篩選作業類別
            search: 關鍵字搜尋 (派工單號/工程名稱)

        Returns:
            BytesIO 物件 (application/vnd.openxmlformats-officedocument.spreadsheetml.sheet)

        Raises:
            ValueError: 篩選結果超過 MAX_EXPORT_ROWS 上限
        """
        # --- Step 1: query dispatch orders (no pagination) ---
        dispatches = await self._query_dispatches(
            contract_project_id=contract_project_id,
            work_type=work_type,
            search=search,
        )

        if len(dispatches) > MAX_EXPORT_ROWS:
            raise ValueError(
                f"匯出上限 {MAX_EXPORT_ROWS} 筆，目前篩選結果 {len(dispatches)} 筆，請縮小篩選範圍"
            )

        dispatch_ids = [d.id for d in dispatches]

        # --- Step 2: query work records for those dispatches ---
        work_records = await self._query_work_records(dispatch_ids) if dispatch_ids else []

        # Group work records by dispatch_order_id for quick lookup
        wr_by_dispatch: Dict[int, List[TaoyuanWorkRecord]] = {}
        for wr in work_records:
            wr_by_dispatch.setdefault(wr.dispatch_order_id, []).append(wr)

        # --- Step 3: build filter description for summary sheet ---
        filter_parts: List[str] = []
        if contract_project_id is not None:
            filter_parts.append(f'承攬案件ID={contract_project_id}')
        if work_type:
            filter_parts.append(f'作業類別={work_type}')
        if search:
            filter_parts.append(f'關鍵字={search}')
        filter_desc = ', '.join(filter_parts) if filter_parts else '全部'

        # --- Step 4: build DataFrames ---
        df_summary = self._build_sheet1(dispatches, wr_by_dispatch)
        df_records = self._build_sheet2(dispatches, work_records)
        df_doc_matrix = self._build_sheet3(dispatches, wr_by_dispatch)
        df_payment = self._build_sheet4(dispatches)
        df_stats = self._build_sheet5(dispatches, work_records, filter_desc)

        # --- Step 5: write to Excel with styling ---
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_summary.to_excel(writer, index=False, sheet_name='派工總表')
            df_records.to_excel(writer, index=False, sheet_name='作業紀錄明細')
            df_doc_matrix.to_excel(writer, index=False, sheet_name='公文對照矩陣')
            df_payment.to_excel(writer, index=False, sheet_name='契金摘要')
            df_stats.to_excel(writer, index=False, sheet_name='統計摘要')

            # Apply styling to every sheet
            for sheet_name in writer.sheets:
                self._apply_sheet_styling(writer.sheets[sheet_name])

        output.seek(0)
        return output

    # =========================================================================
    # 資料查詢
    # =========================================================================

    async def _query_dispatches(
        self,
        contract_project_id: Optional[int] = None,
        work_type: Optional[str] = None,
        search: Optional[str] = None,
    ) -> List[TaoyuanDispatchOrder]:
        """Query all matching dispatch orders with eager-loaded relations (no pagination)."""
        query = select(TaoyuanDispatchOrder).options(
            selectinload(TaoyuanDispatchOrder.document_links).selectinload(
                TaoyuanDispatchDocumentLink.document
            ),
            selectinload(TaoyuanDispatchOrder.attachments),
            selectinload(TaoyuanDispatchOrder.payment),
        )

        conditions = []
        if contract_project_id is not None:
            conditions.append(TaoyuanDispatchOrder.contract_project_id == contract_project_id)
        if work_type:
            conditions.append(TaoyuanDispatchOrder.work_type == work_type)
        if search:
            search_pattern = f"%{search}%"
            conditions.append(
                or_(
                    TaoyuanDispatchOrder.dispatch_no.ilike(search_pattern),
                    TaoyuanDispatchOrder.project_name.ilike(search_pattern),
                )
            )

        if conditions:
            query = query.where(and_(*conditions))

        query = query.order_by(TaoyuanDispatchOrder.id.desc())

        result = await self.db.execute(query)
        return list(result.scalars().unique().all())

    async def _query_work_records(
        self, dispatch_ids: List[int]
    ) -> List[TaoyuanWorkRecord]:
        """Query all work records for the given dispatch IDs (chunked to avoid SQL IN overflow)."""
        if not dispatch_ids:
            return []

        chunk_size = 500
        all_records: List[TaoyuanWorkRecord] = []

        for i in range(0, len(dispatch_ids), chunk_size):
            chunk = dispatch_ids[i:i + chunk_size]
            query = (
                select(TaoyuanWorkRecord)
                .options(
                    selectinload(TaoyuanWorkRecord.document),
                    selectinload(TaoyuanWorkRecord.incoming_doc),
                    selectinload(TaoyuanWorkRecord.outgoing_doc),
                )
                .where(TaoyuanWorkRecord.dispatch_order_id.in_(chunk))
                .order_by(TaoyuanWorkRecord.dispatch_order_id, TaoyuanWorkRecord.sort_order)
            )
            result = await self.db.execute(query)
            all_records.extend(result.scalars().unique().all())

        return all_records

    # =========================================================================
    # Sheet builders
    # =========================================================================

    def _build_sheet1(
        self,
        dispatches: List[TaoyuanDispatchOrder],
        wr_by_dispatch: Dict[int, List[TaoyuanWorkRecord]],
    ) -> pd.DataFrame:
        """Sheet 1: 派工總表 - one row per dispatch order."""
        rows: List[Dict[str, Any]] = []
        for d in dispatches:
            doc_links = d.document_links or []
            incoming_count = sum(1 for lk in doc_links if lk.link_type == 'agency_incoming')
            outgoing_count = sum(1 for lk in doc_links if lk.link_type == 'company_outgoing')

            records = wr_by_dispatch.get(d.id, [])
            total_records = len(records)
            completed_records = sum(1 for r in records if r.status == 'completed')
            current_stage = self._get_current_stage(records)

            payment = d.payment
            current_amount = payment.current_amount if payment else None
            cumulative_amount = payment.cumulative_amount if payment else None

            rows.append({
                '派工單號': d.dispatch_no or '',
                '工程名稱': d.project_name or '',
                '分案備註': d.sub_case_name or '',
                '作業類別': d.work_type or '',
                '承辦人': d.case_handler or '',
                '測量組': d.survey_unit or '',
                '期限': d.deadline or '',
                '來文數': incoming_count,
                '覆文數': outgoing_count,
                '作業紀錄數': total_records,
                '已完成數': completed_records,
                '當前階段': current_stage,
                '本次金額': current_amount,
                '累進金額': cumulative_amount,
                '附件數': len(d.attachments) if d.attachments else 0,
                '建立日期': self._fmt_datetime(d.created_at),
            })

        return pd.DataFrame(rows)

    def _build_sheet2(
        self,
        dispatches: List[TaoyuanDispatchOrder],
        work_records: List[TaoyuanWorkRecord],
    ) -> pd.DataFrame:
        """Sheet 2: 作業紀錄明細 - one row per work record."""
        # Build dispatch_id -> dispatch_no/project_name lookup
        dispatch_lookup: Dict[int, Dict[str, str]] = {
            d.id: {
                'dispatch_no': d.dispatch_no or '',
                'project_name': d.project_name or '',
            }
            for d in dispatches
        }

        rows: List[Dict[str, Any]] = []
        for wr in work_records:
            info = dispatch_lookup.get(wr.dispatch_order_id, {})
            category_label = WORK_CATEGORY_LABELS.get(wr.work_category or '', wr.work_category or '')
            status_label = STATUS_LABELS.get(wr.status or '', wr.status or '')

            rows.append({
                '派工單號': info.get('dispatch_no', ''),
                '工程名稱': info.get('project_name', ''),
                '序號': wr.sort_order,
                '分類': category_label,
                '說明': wr.description or '',
                '紀錄日期': self._fmt_date(wr.record_date),
                '期限日期': self._fmt_date(wr.deadline_date),
                '完成日期': self._fmt_date(wr.completed_date),
                '狀態': status_label,
                '關聯公文字號': self._get_record_doc_number(wr),
            })

        return pd.DataFrame(rows)

    def _build_sheet3(
        self,
        dispatches: List[TaoyuanDispatchOrder],
        wr_by_dispatch: Dict[int, List[TaoyuanWorkRecord]],
    ) -> pd.DataFrame:
        """Sheet 3: 公文對照矩陣 - 3 階段配對演算法 (chain → date proximity → standalone)."""
        rows: List[Dict[str, Any]] = []

        for d in dispatches:
            records = wr_by_dispatch.get(d.id, [])
            doc_links = d.document_links or []

            paired_rows = self._pair_documents_for_dispatch(records, doc_links)
            for inc, out in paired_rows:
                rows.append({
                    '派工單號': d.dispatch_no or '',
                    '工程名稱': d.project_name or '',
                    '來文字號': inc.get('doc_number', '') if inc else '',
                    '來文日期': inc.get('doc_date', '') if inc else '',
                    '來文主旨': inc.get('subject', '') if inc else '',
                    '\u2192': '\u2192',
                    '覆文字號': out.get('doc_number', '') if out else '',
                    '覆文日期': out.get('doc_date', '') if out else '',
                    '覆文主旨': out.get('subject', '') if out else '',
                })

        return pd.DataFrame(rows)

    def _pair_documents_for_dispatch(
        self,
        records: List[TaoyuanWorkRecord],
        doc_links: list,
    ) -> List[tuple]:
        """3 階段公文配對演算法 (與前端 buildCorrespondenceMatrix 一致)

        Phase 1: parent_record_id chain pairing
        Phase 2: date proximity greedy pairing (assigned + unassigned)
        Phase 3: remaining standalone rows

        Returns:
            List of (incoming_dict | None, outgoing_dict | None) tuples
        """
        # --- 分離作業紀錄中的來文/覆文 ---
        assigned_in: List[Dict[str, Any]] = []
        assigned_out: List[Dict[str, Any]] = []
        record_doc_ids: set = set()  # 已出現在 work records 的 document IDs

        for r in records:
            # Old format: incoming_doc / outgoing_doc
            if r.incoming_doc:
                doc = r.incoming_doc
                assigned_in.append({
                    'record_id': r.id,
                    'parent_record_id': getattr(r, 'parent_record_id', None),
                    'doc_number': doc.doc_number or '',
                    'doc_date': self._fmt_date(doc.doc_date),
                    'subject': r.description or (doc.subject or ''),
                })
                if doc.id:
                    record_doc_ids.add(doc.id)
            if r.outgoing_doc:
                doc = r.outgoing_doc
                assigned_out.append({
                    'record_id': r.id,
                    'parent_record_id': getattr(r, 'parent_record_id', None),
                    'doc_number': doc.doc_number or '',
                    'doc_date': self._fmt_date(doc.doc_date),
                    'subject': r.description or (doc.subject or ''),
                })
                if doc.id:
                    record_doc_ids.add(doc.id)
            # New format: document (判斷方向)
            if r.document and not r.incoming_doc and not r.outgoing_doc:
                doc = r.document
                doc_number = doc.doc_number or ''
                item = {
                    'record_id': r.id,
                    'parent_record_id': getattr(r, 'parent_record_id', None),
                    'doc_number': doc_number,
                    'doc_date': self._fmt_date(doc.doc_date),
                    'subject': r.description or (doc.subject or ''),
                }
                if is_outgoing_doc_number(doc_number):
                    assigned_out.append(item)
                else:
                    assigned_in.append(item)
                if doc.id:
                    record_doc_ids.add(doc.id)

        # --- 收集未指派到 work record 的公文 (unassigned) ---
        unassigned_in: List[Dict[str, Any]] = []
        unassigned_out: List[Dict[str, Any]] = []
        for lk in doc_links:
            doc = lk.document
            if not doc or doc.id in record_doc_ids:
                continue
            item = {
                'doc_number': doc.doc_number or '',
                'doc_date': self._fmt_date(doc.doc_date),
                'subject': doc.subject or '',
            }
            if lk.link_type == 'company_outgoing':
                unassigned_out.append(item)
            else:
                unassigned_in.append(item)

        # --- Phase 1: parent_record_id chain pairing ---
        result: List[tuple] = []
        used_in_ids: set = set()
        used_out_ids: set = set()

        for out_item in assigned_out:
            pid = out_item.get('parent_record_id')
            if not pid:
                continue
            for in_item in assigned_in:
                if in_item['record_id'] == pid and in_item['record_id'] not in used_in_ids:
                    result.append((in_item, out_item))
                    used_in_ids.add(in_item['record_id'])
                    used_out_ids.add(out_item['record_id'])
                    break

        # --- Phase 2: date proximity greedy pairing ---
        remain_in = [
            x for x in assigned_in if x.get('record_id') not in used_in_ids
        ]
        remain_out = [
            x for x in assigned_out if x.get('record_id') not in used_out_ids
        ]

        all_in = sorted(remain_in + unassigned_in, key=lambda x: x.get('doc_date', ''))
        all_out = sorted(remain_out + unassigned_out, key=lambda x: x.get('doc_date', ''))

        used_out_idx: set = set()
        date_matched: List[tuple] = []

        for in_item in all_in:
            best_idx = -1
            best_date = ''
            in_date = in_item.get('doc_date', '')

            for j, out_item in enumerate(all_out):
                if j in used_out_idx:
                    continue
                out_date = out_item.get('doc_date', '')
                if out_date >= in_date:
                    if best_idx == -1 or out_date < best_date:
                        best_idx = j
                        best_date = out_date

            if best_idx >= 0:
                date_matched.append((in_item, all_out[best_idx]))
                used_out_idx.add(best_idx)
            else:
                date_matched.append((in_item, None))

        # --- Phase 3: remaining outgoing standalone ---
        for j, out_item in enumerate(all_out):
            if j not in used_out_idx:
                date_matched.append((None, out_item))

        result.extend(date_matched)

        # --- 按最早日期排序 ---
        result.sort(key=lambda pair: (
            pair[0].get('doc_date', '') if pair[0] else (pair[1].get('doc_date', '') if pair[1] else '')
        ))

        return result

    def _build_sheet4(
        self,
        dispatches: List[TaoyuanDispatchOrder],
    ) -> pd.DataFrame:
        """Sheet 4: 契金摘要 - one row per dispatch order with payment data."""
        rows: List[Dict[str, Any]] = []
        for d in dispatches:
            payment: Optional[TaoyuanContractPayment] = d.payment
            if not payment:
                continue

            row: Dict[str, Any] = {
                '派工單號': d.dispatch_no or '',
                '工程名稱': d.project_name or '',
            }

            # 7 work types: date + amount columns
            for idx in range(1, 8):
                label = PAYMENT_WORK_LABELS[idx - 1]
                date_attr = f'work_{idx:02d}_date'
                amount_attr = f'work_{idx:02d}_amount'
                row[f'{label}(日期)'] = self._fmt_date(getattr(payment, date_attr, None))
                row[f'{label}(金額)'] = getattr(payment, amount_attr, None)

            row['本次金額'] = payment.current_amount
            row['累進金額'] = payment.cumulative_amount
            row['剩餘金額'] = payment.remaining_amount
            row['驗收日期'] = self._fmt_date(payment.acceptance_date)

            rows.append(row)

        return pd.DataFrame(rows)

    def _build_sheet5(
        self,
        dispatches: List[TaoyuanDispatchOrder],
        work_records: List[TaoyuanWorkRecord],
        filter_desc: str,
    ) -> pd.DataFrame:
        """Sheet 5: 統計摘要 - key-value pairs."""
        total_dispatches = len(dispatches)
        total_records = len(work_records)
        completed_records = sum(1 for wr in work_records if wr.status == 'completed')

        total_incoming = 0
        total_outgoing = 0
        for d in dispatches:
            for lk in (d.document_links or []):
                if lk.link_type == 'agency_incoming':
                    total_incoming += 1
                elif lk.link_type == 'company_outgoing':
                    total_outgoing += 1

        total_cumulative = sum(
            d.payment.cumulative_amount
            for d in dispatches
            if d.payment and d.payment.cumulative_amount
        )

        rows = [
            {'項目': '匯出時間', '值': datetime.now().strftime('%Y-%m-%d %H:%M:%S')},
            {'項目': '篩選條件', '值': filter_desc},
            {'項目': '派工單總數', '值': total_dispatches},
            {'項目': '作業紀錄總數', '值': total_records},
            {'項目': '已完成紀錄', '值': completed_records},
            {'項目': '來文總數', '值': total_incoming},
            {'項目': '覆文總數', '值': total_outgoing},
            {'項目': '契金累計總額', '值': total_cumulative},
        ]

        return pd.DataFrame(rows)

    # =========================================================================
    # 輔助方法
    # =========================================================================

    @staticmethod
    def _get_record_doc_number(wr: TaoyuanWorkRecord) -> str:
        """Get the effective document number from a work record."""
        if wr.document and wr.document.doc_number:
            return wr.document.doc_number
        if wr.incoming_doc and wr.incoming_doc.doc_number:
            return wr.incoming_doc.doc_number
        if wr.outgoing_doc and wr.outgoing_doc.doc_number:
            return wr.outgoing_doc.doc_number
        return ''

    @staticmethod
    def _get_current_stage(records: List[TaoyuanWorkRecord]) -> str:
        """Find the latest non-completed work record's category as the current stage."""
        # Records are already sorted by sort_order; iterate in reverse to find latest non-completed
        for wr in reversed(records):
            if wr.status != 'completed':
                return WORK_CATEGORY_LABELS.get(wr.work_category or '', wr.work_category or '')
        # All completed or no records
        if records:
            return '全部完成'
        return ''

    @staticmethod
    def _fmt_date(value: Any) -> str:
        """Format a date/datetime to YYYY-MM-DD string."""
        if value is None:
            return ''
        if isinstance(value, datetime):
            return value.strftime('%Y-%m-%d')
        if isinstance(value, date):
            return value.strftime('%Y-%m-%d')
        return str(value)

    @staticmethod
    def _fmt_datetime(value: Any) -> str:
        """Format a datetime to YYYY-MM-DD HH:MM string."""
        if value is None:
            return ''
        if isinstance(value, datetime):
            return value.strftime('%Y-%m-%d %H:%M')
        if isinstance(value, date):
            return value.strftime('%Y-%m-%d')
        return str(value)

    @staticmethod
    def _apply_sheet_styling(ws) -> None:
        """Apply header styling, borders, column widths, and freeze panes to a worksheet."""
        # Freeze first row
        ws.freeze_panes = 'A2'

        # Style header row
        for cell in ws[1]:
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = _HEADER_ALIGNMENT
            cell.border = _THIN_BORDER

        # Apply borders to all data cells and calculate column widths
        col_widths: Dict[int, float] = {}
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.border = _THIN_BORDER
                # Track column width
                val_len = len(str(cell.value or ''))
                # CJK characters count as ~2.2 width units
                cjk_count = sum(1 for c in str(cell.value or '') if '\u4e00' <= c <= '\u9fff')
                estimated_width = val_len + cjk_count * 1.2
                current = col_widths.get(cell.column, 0)
                if estimated_width > current:
                    col_widths[cell.column] = estimated_width

        # Set column widths (capped between 10 and 40)
        for col_idx, width in col_widths.items():
            letter = get_column_letter(col_idx)
            ws.column_dimensions[letter].width = max(10, min(40, width + 2))
