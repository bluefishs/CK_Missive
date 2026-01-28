"""
DocumentExportService - 公文匯出業務邏輯層

提供公文匯出相關的業務邏輯處理。

@version 1.0.0
@date 2026-01-28
"""

import io
import csv
import re
import logging
from typing import Optional, List, Dict, Any
from datetime import date, datetime
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_, func, or_
from sqlalchemy.orm import selectinload

from app.extended.models import OfficialDocument, ContractProject

logger = logging.getLogger(__name__)


class DocumentExportService:
    """
    公文匯出業務邏輯服務

    職責:
    - CSV 匯出
    - Excel 匯出
    - 資料格式化
    """

    def __init__(self, db: AsyncSession):
        self.db = db

    # =========================================================================
    # CSV 匯出
    # =========================================================================

    async def export_to_csv(
        self,
        document_ids: Optional[List[int]] = None,
        category: Optional[str] = None,
        year: Optional[int] = None,
    ) -> bytes:
        """
        匯出公文為 CSV 格式

        Args:
            document_ids: 公文 ID 列表
            category: 類別篩選
            year: 年度篩選

        Returns:
            CSV 位元組
        """
        documents = await self._query_documents(
            document_ids=document_ids,
            category=category,
            year=year
        )

        output = io.StringIO()
        writer = csv.writer(output, quoting=csv.QUOTE_ALL)

        # 寫入標題列
        headers = [
            '序號', '公文文號', '主旨', '類別', '發文/收文日期',
            '發文單位', '受文單位', '承攬案件', '狀態', '備註'
        ]
        writer.writerow(headers)

        # 寫入資料列
        for idx, doc in enumerate(documents, start=1):
            contract_case_name = ""
            if doc.contract_project:
                contract_case_name = doc.contract_project.project_name or ""

            row = [
                doc.auto_serial or idx,
                doc.doc_number or "",
                doc.subject or "",
                doc.category or "",
                str(doc.doc_date) if doc.doc_date else "",
                doc.sender or "",
                doc.receiver or "",
                contract_case_name,
                doc.status or "",
                doc.notes or ""
            ]
            writer.writerow(row)

        output.seek(0)
        return ('\ufeff' + output.getvalue()).encode('utf-8')

    # =========================================================================
    # Excel 匯出
    # =========================================================================

    async def export_to_excel(
        self,
        document_ids: Optional[List[int]] = None,
        category: Optional[str] = None,
        year: Optional[int] = None,
        status: Optional[str] = None,
        keyword: Optional[str] = None,
        contract_case: Optional[str] = None,
        sender: Optional[str] = None,
        receiver: Optional[str] = None,
    ) -> bytes:
        """
        匯出公文為 Excel 格式

        Returns:
            Excel 位元組
        """
        import pandas as pd
        from io import BytesIO

        documents = await self._query_documents(
            document_ids=document_ids,
            category=category,
            year=year,
            status=status,
            keyword=keyword,
            contract_case=contract_case,
            sender=sender,
            receiver=receiver,
            include_attachments=True
        )

        if not documents:
            raise ValueError("沒有符合條件的公文可供匯出")

        # 轉換為 DataFrame
        data = []
        for doc in documents:
            contract_case_name = ""
            if doc.contract_project:
                contract_case_name = doc.contract_project.project_name or ""

            sender_agency_name = ""
            if doc.sender_agency:
                sender_agency_name = doc.sender_agency.agency_name or ""

            receiver_agency_name = ""
            if doc.receiver_agency:
                receiver_agency_name = doc.receiver_agency.agency_name or ""

            attachment_count = len(doc.attachments) if doc.attachments else 0
            attachment_text = f"{attachment_count} 個附件" if attachment_count > 0 else "無"

            data.append({
                "公文ID": doc.id,
                "流水號": doc.auto_serial or "",
                "發文形式": doc.delivery_method or "",
                "類別": doc.category or "",
                "公文類型": self._get_valid_doc_type(doc.doc_type),
                "公文字號": doc.doc_number or "",
                "主旨": doc.subject or "",
                "說明": getattr(doc, 'content', '') or "",
                "公文日期": str(doc.doc_date) if doc.doc_date else "",
                "收文日期": str(doc.receive_date) if doc.receive_date else "",
                "發文日期": str(doc.send_date) if doc.send_date else "",
                "發文單位": self._clean_agency_name(doc.sender or "", sender_agency_name),
                "受文單位": self._clean_agency_name(doc.receiver or "", receiver_agency_name),
                "附件紀錄": attachment_text,
                "備註": getattr(doc, 'notes', '') or "",
                "狀態": doc.status or "",
                "承攬案件": contract_case_name,
                "建立時間": str(doc.created_at) if doc.created_at else "",
                "更新時間": str(doc.updated_at) if doc.updated_at else "",
            })

        df = pd.DataFrame(data)

        # 產生 Excel
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='公文清單')

            # 套用樣式
            self._apply_excel_styles(writer, df, documents)

        output.seek(0)
        return output.getvalue()

    # =========================================================================
    # 輔助方法
    # =========================================================================

    async def _query_documents(
        self,
        document_ids: Optional[List[int]] = None,
        category: Optional[str] = None,
        year: Optional[int] = None,
        status: Optional[str] = None,
        keyword: Optional[str] = None,
        contract_case: Optional[str] = None,
        sender: Optional[str] = None,
        receiver: Optional[str] = None,
        include_attachments: bool = False,
    ) -> List[OfficialDocument]:
        """查詢公文"""
        query = select(OfficialDocument).options(
            selectinload(OfficialDocument.contract_project),
            selectinload(OfficialDocument.sender_agency),
            selectinload(OfficialDocument.receiver_agency),
        )

        if include_attachments:
            query = query.options(selectinload(OfficialDocument.attachments))

        conditions = []
        if document_ids:
            conditions.append(OfficialDocument.id.in_(document_ids))
        if category:
            conditions.append(OfficialDocument.category == category)
        if year:
            conditions.append(func.extract('year', OfficialDocument.doc_date) == year)
        if status:
            conditions.append(OfficialDocument.status == status)
        if keyword:
            kw = f"%{keyword}%"
            conditions.append(
                or_(
                    OfficialDocument.subject.ilike(kw),
                    OfficialDocument.doc_number.ilike(kw),
                    OfficialDocument.sender.ilike(kw),
                    OfficialDocument.receiver.ilike(kw),
                    OfficialDocument.content.ilike(kw),
                    OfficialDocument.notes.ilike(kw)
                )
            )
        if contract_case:
            query = query.outerjoin(
                ContractProject,
                OfficialDocument.contract_project_id == ContractProject.id
            )
            conditions.append(ContractProject.project_name.ilike(f"%{contract_case}%"))
        if sender:
            conditions.append(OfficialDocument.sender.ilike(f"%{sender}%"))
        if receiver:
            conditions.append(OfficialDocument.receiver.ilike(f"%{receiver}%"))

        if conditions:
            query = query.where(and_(*conditions))

        query = query.order_by(OfficialDocument.doc_date.desc())

        result = await self.db.execute(query)
        return list(result.scalars().all())

    def _clean_agency_name(self, raw_text: str, agency_name: str = "") -> str:
        """清理機關名稱"""
        if agency_name:
            return agency_name
        if not raw_text:
            return ""

        text = raw_text.strip()
        paren_match = re.search(r'[（(]([^)）]+)[)）]', text)
        if paren_match:
            return paren_match.group(1).strip()

        text = re.sub(r'^[A-Za-z0-9]+\s*', '', text)
        return text.strip()

    def _get_valid_doc_type(self, doc_type: str) -> str:
        """取得有效的公文類型"""
        if doc_type in ['收文', '發文']:
            return ""
        return doc_type or ""

    def _apply_excel_styles(self, writer, df, documents) -> None:
        """套用 Excel 樣式"""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        worksheet = writer.sheets['公文清單']

        header_font = Font(bold=True, color="000000")
        header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 套用表頭樣式
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 資料列樣式
        data_alignment = Alignment(vertical="center", wrap_text=True)
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                cell.alignment = data_alignment
                cell.border = thin_border

        # 調整欄位寬度
        for idx, col in enumerate(df.columns):
            max_length = max(
                df[col].astype(str).map(len).max(),
                len(col)
            ) + 2
            max_length = min(max_length, 60)
            col_letter = chr(65 + idx) if idx < 26 else f"A{chr(65 + idx - 26)}"
            worksheet.column_dimensions[col_letter].width = max_length

        # 凍結表頭列
        worksheet.freeze_panes = 'A2'

        # 新增統計摘要工作表
        summary_data = {
            "項目": [
                "匯出時間",
                "公文總數",
                "收文數量",
                "發文數量",
                "有附件公文",
                "已指派案件",
            ],
            "數值": [
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                str(len(documents)),
                str(sum(1 for d in documents if d.category == '收文')),
                str(sum(1 for d in documents if d.category == '發文')),
                str(sum(1 for d in documents if d.attachments and len(d.attachments) > 0)),
                str(sum(1 for d in documents if d.contract_project_id)),
            ]
        }

        import pandas as pd
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, index=False, sheet_name='統計摘要')

        summary_ws = writer.sheets['統計摘要']
        for cell in summary_ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        for row in summary_ws.iter_rows(min_row=2, max_row=summary_ws.max_row):
            for cell in row:
                cell.alignment = data_alignment
                cell.border = thin_border

        summary_ws.column_dimensions['A'].width = 20
        summary_ws.column_dimensions['B'].width = 30
