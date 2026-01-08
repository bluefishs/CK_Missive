# -*- coding: utf-8 -*-
"""
Excel 匯入服務 - 手動公文匯入（紙本郵寄等）

繼承 ImportBaseService，使用統一的驗證與處理邏輯。

與 CSV 匯入（電子公文檔匯入）不同：
- CSV 匯入：電子公文系統匯出的格式，欄位固定
- Excel 匯入：使用本系統匯出的 Excel 格式，支援新增/更新

匯入欄位對應：
- 公文ID: 有值=更新現有資料，空白=新增
- 流水號: 新增時自動生成（S/R+序號）
- 其他欄位: 直接匯入

智慧關聯機制：
- 發文單位/受文單位：使用 AgencyMatcher 智慧匹配/建立機關
- 承攬案件：使用 ProjectMatcher 智慧匹配/建立案件
"""
import logging
from typing import Dict, Any, List, Optional
from datetime import datetime
from io import BytesIO

from sqlalchemy.ext.asyncio import AsyncSession

from app.extended.models import OfficialDocument
from app.services.base.import_base import ImportBaseService
from app.services.base.response import ImportResult, ImportRowResult
from app.services.calendar.event_auto_builder import CalendarEventAutoBuilder

logger = logging.getLogger(__name__)


class ExcelImportService(ImportBaseService):
    """Excel 匯入服務"""

    # 匯入欄位對應（Excel 欄位名 → 資料庫欄位）
    FIELD_MAPPING = {
        '公文ID': 'id',
        '流水號': 'auto_serial',
        '發文形式': 'delivery_method',
        '類別': 'category',
        '公文類型': 'doc_type',
        '公文字號': 'doc_number',
        '主旨': 'subject',
        '說明': 'content',
        '公文日期': 'doc_date',
        '收文日期': 'receive_date',
        '發文日期': 'send_date',
        '發文單位': 'sender',
        '受文單位': 'receiver',
        '備註': 'notes',
        '狀態': 'status',
        '承攬案件': 'contract_project_name',
    }

    # 忽略的欄位（系統管理，僅匯出用）
    IGNORED_FIELDS = ['附件紀錄', '建立時間', '更新時間']

    # 必填欄位
    REQUIRED_FIELDS = ['公文字號', '主旨', '類別']

    def __init__(self, db: AsyncSession, auto_create_events: bool = True):
        super().__init__(db)
        self._auto_create_events = auto_create_events
        self._event_builder = CalendarEventAutoBuilder(db) if auto_create_events else None

    async def preview_excel(
        self,
        file_content: bytes,
        filename: str,
        preview_rows: int = 10,
        check_db_duplicates: bool = True
    ) -> Dict[str, Any]:
        """
        預覽 Excel 檔案內容（不執行匯入）

        Args:
            file_content: Excel 檔案內容
            filename: 檔案名稱
            preview_rows: 預覽筆數（預設 10 筆）
            check_db_duplicates: 是否檢查資料庫重複

        Returns:
            預覽結果，包含資料和驗證結果
        """
        try:
            import openpyxl
        except ImportError:
            raise ImportError("需要安裝 openpyxl: pip install openpyxl")

        result = {
            "success": True,
            "filename": filename,
            "total_rows": 0,
            "preview_rows": [],
            "headers": [],
            "field_mapping": self.FIELD_MAPPING,
            "validation": {
                "missing_required_fields": [],
                "invalid_categories": [],
                "invalid_doc_types": [],
                "duplicate_doc_numbers": [],
                "existing_in_db": [],
                "will_insert": 0,
                "will_update": 0,
            },
            "errors": []
        }

        # 預先查詢資料庫中的公文字號
        existing_doc_numbers = set()
        if check_db_duplicates and self.db:
            try:
                from sqlalchemy import select
                query = select(OfficialDocument.doc_number).where(
                    OfficialDocument.doc_number.isnot(None)
                )
                db_result = await self.db.execute(query)
                existing_doc_numbers = {row[0] for row in db_result.fetchall()}
            except Exception as e:
                logger.warning(f"無法查詢資料庫重複: {e}")

        try:
            wb = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
            ws = wb.active

            if ws.max_row < 2:
                result["success"] = False
                result["errors"].append("Excel 檔案沒有資料列")
                return result

            # 取得標題列
            headers = [cell.value for cell in ws[1]]
            result["headers"] = [h for h in headers if h]
            result["total_rows"] = ws.max_row - 1

            # 驗證必要欄位
            for field in self.REQUIRED_FIELDS:
                if field not in headers:
                    result["validation"]["missing_required_fields"].append(field)

            if result["validation"]["missing_required_fields"]:
                result["success"] = False
                result["errors"].append(
                    f"缺少必要欄位: {', '.join(result['validation']['missing_required_fields'])}"
                )

            # 預覽資料
            doc_numbers_seen = set()
            max_preview = min(ws.max_row, preview_rows + 1)

            for row_num in range(2, max_preview + 1):
                row_data = {}
                for col_idx, header in enumerate(headers, 1):
                    if header:
                        cell_value = ws.cell(row=row_num, column=col_idx).value
                        row_data[header] = cell_value

                validation_status = self._validate_preview_row(
                    row_num, row_data, doc_numbers_seen, existing_doc_numbers, result
                )
                result["preview_rows"].append(validation_status)

            # 統計全部資料
            for row_num in range(max_preview + 1, ws.max_row + 1):
                doc_id = ws.cell(row=row_num, column=headers.index('公文ID') + 1 if '公文ID' in headers else 1).value
                if doc_id and str(doc_id).strip():
                    result["validation"]["will_update"] += 1
                else:
                    result["validation"]["will_insert"] += 1

        except Exception as e:
            logger.error(f"Excel 預覽失敗: {e}", exc_info=True)
            result["success"] = False
            result["errors"].append(f"預覽失敗: {str(e)}")

        return result

    def _validate_preview_row(
        self,
        row_num: int,
        row_data: Dict,
        doc_numbers_seen: set,
        existing_doc_numbers: set,
        result: Dict
    ) -> Dict:
        """驗證預覽列"""
        validation_status = {
            "row": row_num,
            "data": row_data,
            "status": "valid",
            "issues": [],
            "action": "insert"
        }

        # 檢查公文ID判斷新增/更新
        doc_id = row_data.get('公文ID')
        if doc_id and str(doc_id).strip():
            validation_status["action"] = "update"
            result["validation"]["will_update"] += 1
        else:
            result["validation"]["will_insert"] += 1

        # 檢查類別
        category = str(row_data.get('類別', '')).strip()
        if category and category not in self.validators.VALID_CATEGORIES:
            validation_status["issues"].append(f"無效類別: {category}")
            result["validation"]["invalid_categories"].append(row_num)

        # 檢查公文類型
        doc_type = str(row_data.get('公文類型', '')).strip()
        if doc_type and doc_type not in self.validators.VALID_DOC_TYPES:
            validation_status["issues"].append(f"無效公文類型: {doc_type}")
            result["validation"]["invalid_doc_types"].append(row_num)

        # 檢查重複公文字號
        doc_number = str(row_data.get('公文字號', '')).strip()
        if doc_number:
            if doc_number in doc_numbers_seen:
                validation_status["issues"].append("檔案內重複公文字號")
                result["validation"]["duplicate_doc_numbers"].append(row_num)
            doc_numbers_seen.add(doc_number)

            if doc_number in existing_doc_numbers and validation_status["action"] == "insert":
                validation_status["issues"].append("資料庫已存在此公文字號")
                result["validation"]["existing_in_db"].append(row_num)

        # 缺少必填欄位
        for field in self.REQUIRED_FIELDS:
            value = row_data.get(field)
            if not value or (isinstance(value, str) and not value.strip()):
                validation_status["issues"].append(f"缺少必填欄位: {field}")
                validation_status["status"] = "warning"

        if validation_status["issues"]:
            validation_status["status"] = "warning"

        return validation_status

    async def import_from_file(
        self,
        file_content: bytes,
        filename: str
    ) -> ImportResult:
        """
        從 Excel 檔案匯入公文資料

        Args:
            file_content: Excel 檔案內容
            filename: 檔案名稱

        Returns:
            匯入結果
        """
        try:
            import openpyxl
        except ImportError:
            raise ImportError("需要安裝 openpyxl: pip install openpyxl")

        result = ImportResult(
            success=True,
            filename=filename,
            total_rows=0,
        )

        try:
            # 重置流水號計數器
            self.reset_serial_counters()

            # 讀取 Excel
            wb = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
            ws = wb.active

            if ws.max_row < 2:
                result.success = False
                result.errors.append(ImportRowResult(
                    row=0, status="error", message="Excel 檔案沒有資料列"
                ))
                return result

            # 取得標題列
            headers = [cell.value for cell in ws[1]]
            result.total_rows = ws.max_row - 1

            # 驗證必要欄位
            missing_fields = [f for f in self.REQUIRED_FIELDS if f not in headers]
            if missing_fields:
                result.success = False
                result.errors.append(ImportRowResult(
                    row=0, status="error", message=f"缺少必要欄位: {', '.join(missing_fields)}"
                ))
                return result

            # 逐行處理
            for row_num in range(2, ws.max_row + 1):
                row_data = {}
                for col_idx, header in enumerate(headers, 1):
                    if header and header not in self.IGNORED_FIELDS:
                        cell_value = ws.cell(row=row_num, column=col_idx).value
                        row_data[header] = cell_value

                row_result = await self.process_row(row_num, row_data)
                result.details.append(row_result)

                if row_result.status == "inserted":
                    result.inserted += 1
                elif row_result.status == "updated":
                    result.updated += 1
                elif row_result.status == "skipped":
                    result.skipped += 1
                    result.warnings.append(row_result)
                elif row_result.status == "error":
                    result.errors.append(row_result)

            # 提交交易
            await self.db.commit()

            logger.info(
                f"[ExcelImport] 匯入完成: {filename}, "
                f"新增={result.inserted}, 更新={result.updated}, "
                f"跳過={result.skipped}, 錯誤={result.error_count}"
            )

        except Exception as e:
            await self.db.rollback()
            logger.error(f"[ExcelImport] 匯入失敗: {e}", exc_info=True)
            result.success = False
            result.errors.append(ImportRowResult(
                row=0, status="error", message=f"匯入失敗: {str(e)}"
            ))

        return result

    async def import_from_excel(
        self,
        file_content: bytes,
        filename: str
    ) -> Dict[str, Any]:
        """
        相容舊 API 的匯入方法

        Args:
            file_content: Excel 檔案內容
            filename: 檔案名稱

        Returns:
            匯入結果（字典格式）
        """
        result = await self.import_from_file(file_content, filename)
        return result.to_dict()

    async def process_row(
        self,
        row_num: int,
        row_data: Dict[str, Any]
    ) -> ImportRowResult:
        """處理單列資料"""
        result = ImportRowResult(
            row=row_num,
            status="error",
            message="",
            doc_number=self.clean_string(row_data.get('公文字號', '')) or ''
        )

        try:
            # 1. 驗證必填欄位
            error_msg = self.validate_required_fields(row_data, self.REQUIRED_FIELDS)
            if error_msg:
                result.status = "skipped"
                result.message = error_msg
                return result

            # 2. 驗證類別
            category = self.clean_string(row_data.get('類別', ''))
            try:
                category = self.validate_category(category)
            except ValueError as e:
                result.status = "skipped"
                result.message = str(e)
                return result

            # 3. 驗證/修正公文類型
            doc_type = self.validate_doc_type(
                self.clean_string(row_data.get('公文類型', '')),
                auto_fix=True
            )

            # 4. 判斷新增或更新
            doc_id = row_data.get('公文ID')
            doc_number = result.doc_number
            existing_doc = None

            if doc_id and str(doc_id).strip():
                try:
                    existing_doc = await self.check_duplicate_by_id(int(doc_id))
                except (ValueError, TypeError):
                    pass
            elif doc_number:
                existing_by_number = await self.check_duplicate_by_doc_number(doc_number)
                if existing_by_number:
                    result.status = "skipped"
                    result.message = f"公文字號 '{doc_number}' 已存在 (ID={existing_by_number.id})"
                    return result

            # 5. 準備資料
            doc_data = await self._prepare_document_data(row_data, category, doc_type)

            if existing_doc:
                # 更新現有資料
                for key, value in doc_data.items():
                    if hasattr(existing_doc, key) and value is not None:
                        setattr(existing_doc, key, value)
                existing_doc.updated_at = datetime.now()
                result.status = "updated"
                result.message = f"已更新公文 ID={doc_id}"
                result.doc_id = existing_doc.id
            else:
                # 新增資料
                auto_serial = await self.generate_auto_serial(category)
                doc_data['auto_serial'] = auto_serial
                doc_data['created_at'] = datetime.now()
                doc_data['updated_at'] = datetime.now()

                new_doc = OfficialDocument(**doc_data)
                self.db.add(new_doc)
                await self.db.flush()  # 取得 ID

                # 自動建立行事曆事件
                if self._auto_create_events and self._event_builder:
                    await self._event_builder.auto_create_event(new_doc, skip_if_exists=False)

                result.status = "inserted"
                result.message = f"已新增公文，流水號={auto_serial}"
                result.doc_id = new_doc.id

        except Exception as e:
            logger.error(f"[ExcelImport] 處理第 {row_num} 列時發生錯誤: {e}")
            result.message = str(e)

        return result

    async def _prepare_document_data(
        self,
        row_data: Dict[str, Any],
        category: str,
        doc_type: str
    ) -> Dict[str, Any]:
        """準備公文資料"""
        sender_name = self.clean_string(row_data.get('發文單位'))
        receiver_name = self.clean_string(row_data.get('受文單位'))
        contract_name = self.clean_string(row_data.get('承攬案件'))

        # 智慧關聯匹配
        sender_agency_id = await self.match_agency(sender_name) if sender_name else None
        receiver_agency_id = await self.match_agency(receiver_name) if receiver_name else None
        contract_project_id = await self.match_project(contract_name) if contract_name else None

        data = {
            'category': category,
            'doc_type': doc_type or '函',
            'doc_number': self.clean_string(row_data.get('公文字號')) or '',
            'subject': self.clean_string(row_data.get('主旨')) or '',
            'content': self.clean_string(row_data.get('說明')),
            'sender': sender_name,
            'receiver': receiver_name,
            'sender_agency_id': sender_agency_id,
            'receiver_agency_id': receiver_agency_id,
            'contract_project_id': contract_project_id,
            'delivery_method': self.clean_string(row_data.get('發文形式')) or '紙本郵寄',
            'notes': self.clean_string(row_data.get('備註')),
            'status': self.clean_string(row_data.get('狀態')) or 'active',
        }

        # 處理日期欄位
        data['doc_date'] = self.parse_date(row_data.get('公文日期'))
        data['receive_date'] = self.parse_date(row_data.get('收文日期'))
        data['send_date'] = self.parse_date(row_data.get('發文日期'))

        return data
