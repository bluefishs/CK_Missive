from sqlalchemy.ext.asyncio import AsyncSession
from typing import Optional, List, Tuple
from decimal import Decimal
from datetime import date, datetime as dt

from app.extended.models.invoice import ExpenseInvoice, ExpenseInvoiceItem
from app.schemas.erp.expense import (
    ExpenseInvoiceCreate, ExpenseInvoiceUpdate, ExpenseInvoiceQuery,
    APPROVAL_THRESHOLD, APPROVAL_TRANSITIONS,
    BUDGET_WARNING_PCT, BUDGET_BLOCK_PCT,
)
from app.repositories.erp.expense_invoice_repository import ExpenseInvoiceRepository
from app.services.finance_ledger_service import FinanceLedgerService
from app.services.audit_mixin import AuditableServiceMixin

import logging

logger = logging.getLogger(__name__)

class ExpenseInvoiceService(AuditableServiceMixin):
    """費用報銷發票業務服務層"""

    AUDIT_TABLE = "expense_invoices"

    def __init__(self, db: AsyncSession):
        self.db = db
        self.repo = ExpenseInvoiceRepository(db)
        self.ledger_service = FinanceLedgerService(db)

    async def create(self, data: ExpenseInvoiceCreate, user_id: Optional[int] = None) -> ExpenseInvoice:
        """建立報銷發票 (含重複檢查 + case_code 驗證)，狀態為 pending 待審核

        帳本記錄在 approve() 審核通過時才寫入，避免雙重記帳。
        """
        # 0. 驗證 case_code 對應專案存在 (Phase 13-2)
        if data.case_code:
            await self._validate_case_code(data.case_code)

        # 1. 檢查是否有重複發票
        is_duplicate = await self.repo.check_duplicate(data.inv_num)
        if is_duplicate:
            raise ValueError(f"發票號碼 {data.inv_num} 已存在，請確認是否重複報銷。")

        # 1.5 自動由 seller_ban 配對 vendor_id
        vendor_id = await self._resolve_vendor_by_ban(data.seller_ban) if data.seller_ban else None

        # 2. 建立 ExpenseInvoice 主檔 (status=pending，等待審核)
        invoice = ExpenseInvoice(
            inv_num=data.inv_num,
            date=data.date,
            amount=data.amount,
            tax_amount=data.tax_amount,
            buyer_ban=data.buyer_ban,
            seller_ban=data.seller_ban,
            case_code=data.case_code,
            attribution_type=getattr(data, "attribution_type", None) or ("project" if data.case_code else "none"),
            operational_account_id=getattr(data, "operational_account_id", None),
            category=data.category,
            source=data.source,
            notes=data.notes,
            user_id=user_id,
            status="pending",
            vendor_id=vendor_id,
            receipt_image_path=getattr(data, "receipt_image_path", None),
            currency=data.currency,
            original_amount=data.original_amount,
            exchange_rate=data.exchange_rate,
        )

        # 3. 建立 Items
        items = []
        if data.items:
            for item_in in data.items:
                items.append(ExpenseInvoiceItem(
                    item_name=item_in.item_name,
                    qty=item_in.qty,
                    unit_price=item_in.unit_price,
                    amount=item_in.amount
                ))

        result = await self.repo.create_with_items(invoice, items)
        await self.audit_create(result.id, {
            "inv_num": data.inv_num,
            "amount": str(data.amount),
            "case_code": data.case_code,
            "category": data.category,
            "source": data.source,
        }, user_id=user_id)
        return result

    async def get_by_id(self, invoice_id: int) -> Optional[ExpenseInvoice]:
        """取得發票詳情"""
        return await self.repo.get_by_id(invoice_id)

    async def update(self, invoice_id: int, data: "ExpenseInvoiceUpdate") -> Optional[ExpenseInvoice]:
        """更新發票部分欄位"""
        invoice = await self.repo.get_by_id(invoice_id)
        if not invoice:
            return None
        update_data = data.model_dump(exclude_unset=True) if hasattr(data, 'model_dump') else data
        result = await self.repo.update_fields(invoice, update_data)
        if result:
            await self.audit_update(invoice_id, update_data)
        return result

    async def approve(self, invoice_id: int) -> Optional[ExpenseInvoice]:
        """多層審核推進 — 依金額門檻自動決定下一狀態

        ≤30K TWD: pending → manager_approved → verified (二級)
        >30K TWD: pending → manager_approved → finance_approved → verified (三級)
        僅 verified 狀態觸發帳本入帳。

        預算聯防：即將進入 verified 時檢查專案預算水位
        - >100%: 攔截審核 (需總經理介入)
        - >80%: 警告但放行 (附帶預警訊息)

        Returns:
            invoice: 更新後的 ExpenseInvoice (含 _budget_warning 動態屬性)
        """
        invoice = await self.repo.get_by_id(invoice_id)
        if not invoice:
            return None

        current = invoice.status
        if current in ("verified", "rejected"):
            raise ValueError(f"此發票狀態為「{current}」，不可進行審核操作")

        allowed = APPROVAL_TRANSITIONS.get(current, [])
        if "rejected" in allowed:
            allowed = [s for s in allowed if s != "rejected"]
        if not allowed:
            raise ValueError(f"狀態「{current}」無法進行審核推進")

        next_status = self._determine_next_approval(current, invoice.amount)

        if next_status not in APPROVAL_TRANSITIONS.get(current, []):
            raise ValueError(f"非法狀態流轉: {current} → {next_status}")

        # === 預算聯防控制 ===
        budget_warning: Optional[str] = None
        if next_status == "verified" and invoice.case_code:
            budget_warning = await self._check_budget(invoice.case_code, invoice.amount)

        await self.repo.update_status(invoice, next_status)

        # 僅最終 verified 才寫入帳本
        if next_status == "verified":
            await self.ledger_service.record_from_expense(invoice)

        await self.repo.commit()

        # 通知推送
        await self._notify_status_change(invoice, current, next_status, budget_warning)

        # 將預算警告附加為動態屬性，API 層可讀取
        invoice._budget_warning = budget_warning  # type: ignore[attr-defined]
        await self.audit_update(invoice_id, {"status": next_status, "action": "approve"})
        return invoice

    async def _notify_status_change(
        self, invoice, old_status: str, new_status: str, budget_warning: Optional[str] = None
    ) -> None:
        """核銷狀態變更通知"""
        try:
            from app.services.notification_helpers import _safe_create_notification

            STATUS_LABELS = {
                "pending": "待主管審核", "manager_approved": "主管已核准",
                "finance_approved": "財務已核准", "verified": "最終通過",
                "rejected": "已駁回",
            }
            title = f"核銷審核: {invoice.inv_num} → {STATUS_LABELS.get(new_status, new_status)}"
            msg = f"發票 {invoice.inv_num} (NT$ {invoice.amount:,.0f}) 狀態: {STATUS_LABELS.get(old_status, old_status)} → {STATUS_LABELS.get(new_status, new_status)}"
            if budget_warning:
                msg += f"\n⚠️ {budget_warning}"

            severity = "info"
            if new_status == "verified":
                severity = "success"
            elif new_status == "rejected":
                severity = "warning"

            await _safe_create_notification(
                notification_type="expense_approval",
                severity=severity,
                title=title,
                message=msg,
                source_table="expense_invoices",
                source_id=invoice.id,
                user_id=invoice.user_id,
            )
        except Exception as e:
            logger.debug(f"通知推送失敗 (非阻塞): {e}")

    async def _resolve_vendor_by_ban(self, seller_ban: str) -> Optional[int]:
        """由賣方統編查找 partner_vendors.id (稅籍號碼 = vendor_code 慣例)"""
        from app.repositories.vendor_repository import VendorRepository
        vendor_repo = VendorRepository(self.db)
        return await vendor_repo.get_id_by_vendor_code(seller_ban)

    async def _validate_case_code(self, case_code: str) -> None:
        """驗證 case_code 對應的專案或報價存在

        檢查順序: ContractProject.project_code → PMCase.case_code → ERPQuotation.case_code
        任一匹配即通過，全部不匹配則拋出 ValueError。
        """
        from app.repositories import ProjectRepository

        # 1. 檢查 ContractProject
        project_repo = ProjectRepository(self.db)
        if await project_repo.exists_by_project_code(case_code):
            return

        # 2. 檢查 PMCase (soft match)
        try:
            from sqlalchemy import select, exists
            from app.extended.models.pm import PMCase
            stmt2 = select(exists().where(PMCase.case_code == case_code))
            found2 = await self.db.scalar(stmt2)
            if found2:
                return
        except ImportError:
            pass

        # 3. 檢查 ERPQuotation
        try:
            from sqlalchemy import select, exists
            from app.extended.models.erp import ERPQuotation
            stmt3 = select(exists().where(ERPQuotation.case_code == case_code))
            found3 = await self.db.scalar(stmt3)
            if found3:
                return
        except ImportError:
            pass

        raise ValueError(f"案號 {case_code} 不存在，請確認後再提交。")

    async def _check_budget(self, case_code: str, invoice_amount: Decimal) -> Optional[str]:
        """預算聯防 — 檢查專案預算水位

        Returns:
            None: 預算充足
            str: 預警訊息 (>80%) 或 raises ValueError (>100%)
        """
        from sqlalchemy import select
        from app.extended.models.erp import ERPQuotation

        # 1. 取得該專案的預算上限 (可能有多張報價單，取最大 budget_limit)
        stmt = (
            select(ERPQuotation.budget_limit)
            .where(ERPQuotation.case_code == case_code)
            .where(ERPQuotation.budget_limit.is_not(None))
            .order_by(ERPQuotation.budget_limit.desc())
            .limit(1)
        )
        result = await self.db.execute(stmt)
        budget_limit = result.scalar_one_or_none()

        if not budget_limit or budget_limit <= 0:
            return None  # 無預算設定，略過檢查

        budget = Decimal(str(budget_limit))
        amount = Decimal(str(invoice_amount)) if not isinstance(invoice_amount, Decimal) else invoice_amount

        # 2. 取得累計支出 (已入帳的 expense 總額)
        balance = await self.ledger_service.get_case_balance(case_code)
        cumulative_expense = Decimal(str(balance.get("expense", 0)))

        # 3. 預測核准後的支出總額
        projected = cumulative_expense + amount
        usage_pct = (projected / budget) * Decimal("100")

        logger.info(
            f"預算檢查 [{case_code}]: 累計支出={cumulative_expense}, "
            f"本筆={amount}, 預算={budget}, 預測使用率={usage_pct:.1f}%"
        )

        # 4. 判定
        if usage_pct > BUDGET_BLOCK_PCT:
            raise ValueError(
                f"預算超限！專案 {case_code} 累計支出將達 {projected:,.0f} 元 "
                f"(預算 {budget:,.0f} 元，使用率 {usage_pct:.1f}%)。"
                f"請聯繫總經理核准後再行操作。"
            )

        if usage_pct > BUDGET_WARNING_PCT:
            return (
                f"⚠️ 預算警告：專案 {case_code} 核准後累計支出將達 {projected:,.0f} 元 "
                f"(預算 {budget:,.0f} 元，使用率 {usage_pct:.1f}%)"
            )

        return None

    def _determine_next_approval(self, current_status: str, amount: Decimal) -> str:
        """根據當前狀態與金額決定下一審核狀態"""
        amount_val = Decimal(str(amount)) if not isinstance(amount, Decimal) else amount
        is_high_value = amount_val > APPROVAL_THRESHOLD

        if current_status == "pending":
            return "manager_approved"
        elif current_status == "pending_receipt":
            return "pending"
        elif current_status == "manager_approved":
            return "finance_approved" if is_high_value else "verified"
        elif current_status == "finance_approved":
            return "verified"
        else:
            raise ValueError(f"狀態「{current_status}」無法推進審核")

    async def reject(self, invoice_id: int, reason: Optional[str] = None) -> Optional[ExpenseInvoice]:
        """駁回報銷 — 任何非終態階段皆可駁回"""
        invoice = await self.repo.get_by_id(invoice_id)
        if not invoice:
            return None
        if invoice.status in ("verified", "rejected"):
            raise ValueError(f"此發票狀態為「{invoice.status}」，不可駁回")

        allowed = APPROVAL_TRANSITIONS.get(invoice.status, [])
        if "rejected" not in allowed:
            raise ValueError(f"狀態「{invoice.status}」不允許駁回")

        notes_append = f"[駁回] {reason}" if reason else None
        result = await self.repo.update_status(invoice, "rejected", notes_append=notes_append)
        if result:
            await self.audit_update(invoice_id, {"status": "rejected", "action": "reject", "reason": reason})
        return result

    def parse_qr_data(self, raw_qr: str) -> dict:
        """解析台灣電子發票 QR Code 資料 (財政部規範)

        Head QR 格式 (77+ 字元):
        - [0:10]   發票號碼 (2英+8數)
        - [10:17]  民國日期 YYYMMDD
        - [17:21]  隨機碼 4 碼
        - [21:29]  銷售額 hex 8 碼 (未稅)
        - [29:37]  總額 hex 8 碼 (含稅)
        - [37:45]  買方統編 8 碼 (無則 00000000)
        - [45:53]  賣方統編 8 碼
        - [53:77]  驗證碼 24 碼
        """
        if len(raw_qr) < 53:
            raise ValueError("QR 資料格式不正確，長度不足 (需至少 53 字元)")

        inv_num = raw_qr[0:10]
        date_str = raw_qr[10:17]  # 民國年 YYYMMDD

        # 民國年轉西元
        from datetime import date as date_type
        roc_year = int(date_str[0:3])
        month = int(date_str[3:5])
        day = int(date_str[5:7])
        inv_date = date_type(roc_year + 1911, month, day)

        # 隨機碼
        random_code = raw_qr[17:21]

        # 銷售額 (未稅) + 總額 (含稅) — hex 8 碼
        sales_hex = raw_qr[21:29]
        total_hex = raw_qr[29:37]
        sales_amount = Decimal(str(int(sales_hex, 16)))
        total_amount = Decimal(str(int(total_hex, 16)))
        tax_amount = total_amount - sales_amount

        # 統編
        buyer_ban = raw_qr[37:45]
        seller_ban = raw_qr[45:53]
        if buyer_ban == "00000000":
            buyer_ban = None

        return {
            "inv_num": inv_num,
            "date": inv_date,
            "random_code": random_code,
            "sales_amount": sales_amount,
            "amount": total_amount,       # 含稅總額
            "tax_amount": tax_amount,
            "buyer_ban": buyer_ban,
            "seller_ban": seller_ban,
            "source": "qr_scan",
            "raw_qr_data": raw_qr,
        }

    async def create_from_qr(
        self, raw_qr: str, case_code: Optional[str] = None,
        category: Optional[str] = None, user_id: Optional[int] = None,
    ) -> ExpenseInvoice:
        """從 QR Code 建立報銷發票"""
        parsed = self.parse_qr_data(raw_qr)

        data = ExpenseInvoiceCreate(
            inv_num=parsed["inv_num"],
            date=parsed["date"],
            amount=parsed["amount"],
            buyer_ban=parsed["buyer_ban"],
            seller_ban=parsed["seller_ban"],
            case_code=case_code,
            category=category,
            source="qr_scan",
        )
        return await self.create(data, user_id=user_id)

    async def attach_receipt(self, invoice_id: int, receipt_path: str) -> Optional[ExpenseInvoice]:
        """附加收據影像至發票 (不變更狀態)"""
        invoice = await self.repo.get_by_id(invoice_id)
        if not invoice:
            return None
        return await self.repo.update_fields(invoice, {"receipt_image_path": receipt_path})

    async def list_by_case(self, case_code: str, skip=0, limit=20) -> Tuple[List[ExpenseInvoice], int]:
        return await self.repo.find_by_case_code(case_code, skip, limit)

    async def query(self, params: ExpenseInvoiceQuery) -> Tuple[List[ExpenseInvoice], int]:
        return await self.repo.query(params)

    async def auto_link_einvoice(self, expense_id: int) -> Optional[dict]:
        """自動關聯電子發票 — 用 inv_num 查詢 einvoice_sync_logs 是否已同步

        若該發票已由財政部同步過 (synced_at 已有值)，回傳 already_synced。
        若尚未同步，嘗試從 einvoice_sync_logs 找到涵蓋該發票日期的成功同步批次，
        並將 synced_at 更新為批次完成時間。
        """
        from datetime import datetime
        from sqlalchemy import select

        expense = await self.repo.get_by_id(expense_id)
        if not expense or not expense.inv_num:
            return None

        # Already synced
        if expense.synced_at:
            return {"status": "already_synced", "synced_at": str(expense.synced_at)}

        # Try to find a successful sync batch that covers this invoice's date
        from app.extended.models.einvoice_sync import EInvoiceSyncLog
        stmt = (
            select(EInvoiceSyncLog)
            .where(EInvoiceSyncLog.status.in_(["success", "partial"]))
            .where(EInvoiceSyncLog.query_start <= expense.date)
            .where(EInvoiceSyncLog.query_end >= expense.date)
            .order_by(EInvoiceSyncLog.completed_at.desc())
            .limit(1)
        )
        result = await self.db.execute(stmt)
        sync_log = result.scalars().first()

        if sync_log:
            expense.synced_at = sync_log.completed_at or datetime.now()
            if expense.source == "manual":
                expense.source = "mof_sync"
            await self.db.commit()
            return {
                "status": "linked",
                "sync_log_id": sync_log.id,
                "synced_at": str(expense.synced_at),
            }

        return {"status": "not_found", "inv_num": expense.inv_num}

    def generate_import_template(self) -> bytes:
        """產生費用報銷匯入範本 Excel"""
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "費用報銷匯入"

        headers = ["發票號碼", "日期", "金額", "稅額", "買方統編", "賣方統編",
                   "案件代碼", "類別", "備註"]

        fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        font = Font(color="FFFFFF", bold=True)
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.fill = fill
            cell.font = font

        # 範例資料
        ws.append(["AB12345678", "2025-06-15", 5000, 250, "12345678", "87654321",
                   "B114-B001", "交通費", "出差交通"])

        # 說明頁
        ws2 = wb.create_sheet("說明")
        ws2.append(["欄位", "說明", "必填"])
        ws2.append(["發票號碼", "10碼發票號碼 (唯一)", "是"])
        ws2.append(["日期", "YYYY-MM-DD 格式", "是"])
        ws2.append(["金額", "含稅金額 (數字)", "是"])
        ws2.append(["稅額", "稅額 (數字)，預設 0", "否"])
        ws2.append(["買方統編", "8碼統一編號", "否"])
        ws2.append(["賣方統編", "8碼統一編號", "否"])
        ws2.append(["案件代碼", "對應專案或報價的案號", "否"])
        ws2.append(["類別", "交通費/差旅費/文具及印刷/郵電費/水電費/保險費/租金/維修費/雜費/設備採購/外包及勞務/訓練費/材料費/報銷及費用/其他", "否"])
        ws2.append(["備註", "自由文字", "否"])

        # 自動欄寬
        for col in ws.columns:
            mx = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(mx + 4, 25)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    async def import_from_excel(self, file_bytes: bytes, user_id: Optional[int] = None) -> dict:
        """匯入費用報銷 Excel，回傳 {total, created, skipped, errors}"""
        from app.services.base.excel_reader import load_workbook_any

        wb = load_workbook_any(file_bytes)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))

        created = 0
        skipped = 0
        errors: list[dict] = []

        for idx, row in enumerate(rows, start=2):
            try:
                if not row or len(row) < 3:
                    continue
                if not row[0] or not row[2]:  # inv_num and amount required
                    continue

                import unicodedata
                inv_num = unicodedata.normalize('NFKC', str(row[0]).strip())

                # 重複檢查
                existing = await self.repo.find_by_inv_num(inv_num)
                if existing:
                    skipped += 1
                    continue

                def _parse_num(v) -> float:
                    """解析數字，支援千分位逗號和貨幣符號"""
                    if v is None:
                        return 0
                    if isinstance(v, (int, float)):
                        return float(v)
                    import re
                    s = re.sub(r'[NT$￥¥€£\s,]', '', str(v).strip())
                    return float(s) if s else 0

                def _clean(v) -> str | None:
                    if v is None:
                        return None
                    return unicodedata.normalize('NFKC', str(v).strip()) or None

                data = {
                    "inv_num": inv_num,
                    "amount": _parse_num(row[2]),
                    "tax_amount": _parse_num(row[3]) if row[3] else 0,
                    "buyer_ban": _clean(row[4]),
                    "seller_ban": _clean(row[5]),
                    "case_code": _clean(row[6]),
                    "category": _clean(row[7]) or "其他",
                    "notes": _clean(row[8]) if len(row) > 8 else None,
                    "status": "pending",
                    "source": "manual",
                    "user_id": user_id,
                }

                # 解析日期
                if row[1]:
                    if isinstance(row[1], (date, dt)):
                        data["date"] = row[1] if isinstance(row[1], date) else row[1].date()
                    else:
                        data["date"] = dt.strptime(str(row[1]).strip(), "%Y-%m-%d").date()

                # 自動配對 vendor
                vendor_id = await self._resolve_vendor_by_ban(data["seller_ban"]) if data.get("seller_ban") else None

                invoice = ExpenseInvoice(
                    inv_num=data["inv_num"],
                    date=data.get("date"),
                    amount=Decimal(str(data["amount"])),
                    tax_amount=Decimal(str(data["tax_amount"])) if data["tax_amount"] else Decimal("0"),
                    buyer_ban=data["buyer_ban"],
                    seller_ban=data["seller_ban"],
                    case_code=data["case_code"],
                    category=data["category"],
                    notes=data["notes"],
                    status="pending",
                    source="manual",
                    user_id=user_id,
                    vendor_id=vendor_id,
                )
                self.db.add(invoice)
                created += 1
            except Exception as e:
                errors.append({"row": idx, "error": str(e)})

        if created > 0:
            await self.db.commit()
        wb.close()
        return {"total": len(rows), "created": created, "skipped": skipped, "errors": errors}

    @staticmethod
    def get_approval_info(invoice: ExpenseInvoice) -> dict:
        """計算發票的審核層級資訊 (用於 Response 填充)"""
        status = invoice.status
        amount = Decimal(str(invoice.amount)) if invoice.amount else Decimal("0")
        is_high_value = amount > APPROVAL_THRESHOLD

        level_map = {
            "pending": "pending",
            "pending_receipt": "pending",
            "manager_approved": "manager",
            "finance_approved": "finance",
            "verified": "final",
            "rejected": None,
        }

        next_map: dict[str, Optional[str]] = {
            "pending": "manager",
            "pending_receipt": "manager",
            "manager_approved": "finance" if is_high_value else "final",
            "finance_approved": "final",
            "verified": None,
            "rejected": None,
        }

        return {
            "approval_level": level_map.get(status),
            "next_approval": next_map.get(status),
        }
