"""
費用報銷匯入/匯出服務 — QR 解析、Excel 匯入、電子發票關聯

拆分自 expense_invoice_service.py，處理資料匯入/匯出邏輯。

Version: 1.0.0
"""
from sqlalchemy.ext.asyncio import AsyncSession
from typing import Optional, List, Tuple
from decimal import Decimal
from datetime import date, datetime as dt

from app.extended.models.invoice import ExpenseInvoice
from app.schemas.erp.expense import ExpenseInvoiceCreate
from app.repositories.erp.expense_invoice_repository import ExpenseInvoiceRepository

import logging

logger = logging.getLogger(__name__)


class ExpenseImportService:
    """費用報銷匯入/匯出"""

    def __init__(self, db: AsyncSession):
        self.db = db
        self.repo = ExpenseInvoiceRepository(db)

    def parse_qr_data(self, raw_qr: str) -> dict:
        """解析台灣電子發票 QR Code 資料 (財政部規範)

        Head QR 格式 (77+ 字元):
        - [0:10]   發票號碼 (2英+8數)
        - [10:17]  民國日期 YYYMMDD
        - [17:21]  隨機碼 4 碼
        - [21:29]  銷售額 hex 8 碼 (未稅)
        - [29:37]  總額 hex 8 碼 (含稅)
        - [37:45]  買方統編 8 碼 (無則 00000000)
        - [45:53]  賣方統編 8 碼
        - [53:77]  驗證碼 24 碼
        """
        if len(raw_qr) < 53:
            raise ValueError("QR 資料格式不正確，長度不足 (需至少 53 字元)")

        inv_num = raw_qr[0:10]
        date_str = raw_qr[10:17]  # 民國年 YYYMMDD

        # 民國年轉西元
        from datetime import date as date_type
        roc_year = int(date_str[0:3])
        month = int(date_str[3:5])
        day = int(date_str[5:7])
        inv_date = date_type(roc_year + 1911, month, day)

        # 隨機碼
        random_code = raw_qr[17:21]

        # 銷售額 (未稅) + 總額 (含稅) — hex 8 碼
        sales_hex = raw_qr[21:29]
        total_hex = raw_qr[29:37]
        sales_amount = Decimal(str(int(sales_hex, 16)))
        total_amount = Decimal(str(int(total_hex, 16)))
        tax_amount = total_amount - sales_amount

        # 統編
        buyer_ban = raw_qr[37:45]
        seller_ban = raw_qr[45:53]
        if buyer_ban == "00000000":
            buyer_ban = None

        return {
            "inv_num": inv_num,
            "date": inv_date,
            "random_code": random_code,
            "sales_amount": sales_amount,
            "amount": total_amount,       # 含稅總額
            "tax_amount": tax_amount,
            "buyer_ban": buyer_ban,
            "seller_ban": seller_ban,
            "source": "qr_scan",
            "raw_qr_data": raw_qr,
        }

    async def auto_link_einvoice(self, expense_id: int) -> Optional[dict]:
        """自動關聯電子發票 — 用 inv_num 查詢 einvoice_sync_logs 是否已同步"""
        from datetime import datetime
        from sqlalchemy import select

        expense = await self.repo.get_by_id(expense_id)
        if not expense or not expense.inv_num:
            return None

        # Already synced
        if expense.synced_at:
            return {"status": "already_synced", "synced_at": str(expense.synced_at)}

        # Try to find a successful sync batch that covers this invoice's date
        from app.extended.models.einvoice_sync import EInvoiceSyncLog
        stmt = (
            select(EInvoiceSyncLog)
            .where(EInvoiceSyncLog.status.in_(["success", "partial"]))
            .where(EInvoiceSyncLog.query_start <= expense.date)
            .where(EInvoiceSyncLog.query_end >= expense.date)
            .order_by(EInvoiceSyncLog.completed_at.desc())
            .limit(1)
        )
        result = await self.db.execute(stmt)
        sync_log = result.scalars().first()

        if sync_log:
            expense.synced_at = sync_log.completed_at or datetime.now()
            if expense.source == "manual":
                expense.source = "mof_sync"
            await self.db.commit()
            return {
                "status": "linked",
                "sync_log_id": sync_log.id,
                "synced_at": str(expense.synced_at),
            }

        return {"status": "not_found", "inv_num": expense.inv_num}

    def generate_import_template(self) -> bytes:
        """產生費用報銷匯入範本 Excel"""
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "費用報銷匯入"

        # 必填 3 欄用藍色，可選 6 欄用灰色 — 使用端反應欄位太多
        headers = ["發票號碼", "日期", "金額", "案件代碼", "類別", "備註",
                   "稅額", "買方統編", "賣方統編"]

        fill_req = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        fill_opt = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        font_w = Font(color="FFFFFF", bold=True)
        font_b = Font(color="333333", bold=True)
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            if col_idx <= 3:  # 必填
                cell.fill = fill_req
                cell.font = font_w
            else:  # 可選
                cell.fill = fill_opt
                cell.font = font_b

        # 範例資料 — 最簡 3 欄即可匯入
        ws.append(["AB12345678", "2025-06-15", 5000, "B114-B001", "交通費", "出差交通",
                   250, "12345678", "87654321"])

        # 說明頁
        ws2 = wb.create_sheet("說明")
        ws2.append(["欄位", "說明", "必填"])
        ws2.append(["發票號碼", "10碼發票號碼 (唯一鍵)", "是 (藍色)"])
        ws2.append(["日期", "YYYY-MM-DD 格式", "是"])
        ws2.append(["金額", "含稅金額 (數字)", "是"])
        ws2.append(["案件代碼", "對應專案的案號 (如 B114-B001)", "否 (灰色)"])
        ws2.append(["類別", "交通費/差旅費/文具及印刷/郵電費/水電費/保險費/租金/維修費/雜費/設備採購/外包及勞務/材料費/其他", "否"])
        ws2.append(["備註", "自由文字", "否"])
        ws2.append(["稅額", "稅額 (數字)，預設 0", "否"])
        ws2.append(["買方統編", "8碼統一編號", "否"])
        ws2.append(["賣方統編", "8碼統一編號，自動配對廠商", "否"])
        ws2.append([])
        ws2.append(["提示", "只填前 3 欄（發票號碼、日期、金額）也能匯入"])

        # 自動欄寬
        for col in ws.columns:
            mx = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(mx + 4, 25)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    async def import_from_excel(self, file_bytes: bytes, user_id: Optional[int] = None) -> dict:
        """匯入費用報銷 Excel，回傳 {total, created, skipped, errors}"""
        from app.services.base.excel_reader import load_workbook_any

        wb = load_workbook_any(file_bytes)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))

        created = 0
        skipped = 0
        errors: list[dict] = []

        for idx, row in enumerate(rows, start=2):
            try:
                if not row or len(row) < 3:
                    continue
                if not row[0] or not row[2]:  # inv_num and amount required
                    continue

                import unicodedata
                inv_num = unicodedata.normalize('NFKC', str(row[0]).strip())

                # 重複檢查
                existing = await self.repo.find_by_inv_num(inv_num)
                if existing:
                    skipped += 1
                    continue

                def _parse_num(v) -> float:
                    """解析數字，支援千分位逗號和貨幣符號"""
                    if v is None:
                        return 0
                    if isinstance(v, (int, float)):
                        return float(v)
                    import re
                    s = re.sub(r'[NT$￥¥€£\s,]', '', str(v).strip())
                    return float(s) if s else 0

                def _clean(v) -> str | None:
                    if v is None:
                        return None
                    return unicodedata.normalize('NFKC', str(v).strip()) or None

                # 欄位順序: 發票號碼/日期/金額/案件代碼/類別/備註/稅額/買方統編/賣方統編
                # 前 3 欄必填，其餘可選 (v5.5.4 簡化)
                data = {
                    "inv_num": inv_num,
                    "amount": _parse_num(row[2]),
                    "case_code": _clean(row[3]) if len(row) > 3 else None,
                    "category": _clean(row[4]) if len(row) > 4 else "其他",
                    "notes": _clean(row[5]) if len(row) > 5 else None,
                    "tax_amount": _parse_num(row[6]) if len(row) > 6 and row[6] else 0,
                    "buyer_ban": _clean(row[7]) if len(row) > 7 else None,
                    "seller_ban": _clean(row[8]) if len(row) > 8 else None,
                    "status": "pending",
                    "source": "manual",
                    "user_id": user_id,
                }

                # 解析日期
                if row[1]:
                    if isinstance(row[1], (date, dt)):
                        data["date"] = row[1] if isinstance(row[1], date) else row[1].date()
                    else:
                        data["date"] = dt.strptime(str(row[1]).strip(), "%Y-%m-%d").date()

                # 自動配對 vendor
                vendor_id = await self._resolve_vendor_by_ban(data["seller_ban"]) if data.get("seller_ban") else None

                invoice = ExpenseInvoice(
                    inv_num=data["inv_num"],
                    date=data.get("date"),
                    amount=Decimal(str(data["amount"])),
                    tax_amount=Decimal(str(data["tax_amount"])) if data["tax_amount"] else Decimal("0"),
                    buyer_ban=data["buyer_ban"],
                    seller_ban=data["seller_ban"],
                    case_code=data["case_code"],
                    category=data["category"],
                    notes=data["notes"],
                    status="pending",
                    source="manual",
                    user_id=user_id,
                    vendor_id=vendor_id,
                )
                self.db.add(invoice)
                created += 1
            except Exception as e:
                errors.append({"row": idx, "error": str(e)})

        if created > 0:
            await self.db.commit()
        wb.close()
        return {"total": len(rows), "created": created, "skipped": skipped, "errors": errors}

    async def _resolve_vendor_by_ban(self, seller_ban: str) -> Optional[int]:
        """由賣方統編查找 partner_vendors.id"""
        from app.repositories.vendor_repository import VendorRepository
        vendor_repo = VendorRepository(self.db)
        return await vendor_repo.get_id_by_vendor_code(seller_ban)
