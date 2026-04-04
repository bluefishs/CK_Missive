"""ERP 財務報表匯出服務

提供費用報銷與帳本資料的 Excel 匯出功能：
- 費用報銷明細表 (按日期區間)
- 帳本收支明細表
- 專案費用彙總表

Version: 1.0.0
Created: 2026-03-21
"""
import io
import logging
from datetime import date, datetime
from typing import Optional, List, Tuple
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.ext.asyncio import AsyncSession

from app.repositories.erp.expense_invoice_repository import ExpenseInvoiceRepository
from app.repositories.erp.ledger_repository import LedgerRepository
from app.schemas.erp.expense import ExpenseInvoiceQuery

logger = logging.getLogger(__name__)

# 樣式常數
_HEADER_FONT = Font(bold=True, size=11)
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")
_TITLE_FONT = Font(bold=True, size=14)
_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_NUM_FORMAT = '#,##0'
_NUM_FORMAT_2D = '#,##0.00'

# 狀態/來源中文對照
_STATUS_LABELS = {
    "pending": "待主管審核", "pending_receipt": "待上傳收據",
    "manager_approved": "主管已核准", "finance_approved": "財務已核准",
    "verified": "最終通過", "rejected": "已駁回",
}
_SOURCE_LABELS = {
    "manual": "手動", "qr_scan": "QR掃描", "api": "API",
    "ocr": "OCR辨識", "mof_sync": "財政部同步",
}


class FinanceExportService:
    """ERP 財務報表 Excel 匯出"""

    def __init__(self, db: AsyncSession):
        self.db = db
        self.expense_repo = ExpenseInvoiceRepository(db)
        self.ledger_repo = LedgerRepository(db)

    async def export_expenses(
        self,
        date_from: Optional[date] = None,
        date_to: Optional[date] = None,
        case_code: Optional[str] = None,
        status: Optional[str] = None,
        attribution_type: Optional[str] = None,
    ) -> bytes:
        """匯出費用報銷明細 Excel (按歸屬分組，每組一個 Sheet)"""
        query = ExpenseInvoiceQuery(
            date_from=date_from, date_to=date_to,
            case_code=case_code, status=status,
            attribution_type=attribution_type,
            skip=0, limit=100,
        )
        items, total = await self.expense_repo.query(query)

        # 按歸屬分組
        from collections import defaultdict
        groups: dict = defaultdict(list)
        for item in items:
            attr = str(getattr(item, 'attribution_type', 'none') or 'none')
            cc = str(getattr(item, 'case_code', '') or '')
            key = cc if attr == 'project' and cc else attr
            groups[key].append(item)

        wb = Workbook()
        first_sheet = True
        for group_key, group_items in sorted(groups.items(), key=lambda x: str(x[0])):
            sheet_name = group_key[:31]  # Excel sheet name max 31 chars
            if first_sheet:
                ws = wb.active
                ws.title = sheet_name
                first_sheet = False
            else:
                ws = wb.create_sheet(title=sheet_name)
            self._write_expense_sheet(ws, group_items, group_key, date_from, date_to)

        if first_sheet:
            ws = wb.active
            ws.title = "費用報銷明細"
            self._write_expense_sheet(ws, [], "無資料", date_from, date_to)

        return self._save_to_bytes(wb)

    def _write_expense_sheet(self, ws, items, group_key: str, date_from=None, date_to=None):
        """寫入單一 Sheet 的費用資料"""
        ws.merge_cells("A1:L1")
        title_cell = ws["A1"]
        title_cell.value = self._build_title(f"費用明細 — {group_key}", date_from, date_to, None)
        title_cell.font = _TITLE_FONT
        title_cell.alignment = Alignment(horizontal="center")

        ws.append([])

        headers = ["發票號碼", "日期", "金額", "幣別", "原幣金額", "匯率", "稅額", "分類", "案號", "來源", "狀態", "備註"]
        ws.append(headers)
        for col_idx, _ in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx)
            cell.font = _HEADER_FONT_WHITE
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = _BORDER

        total_amount = Decimal("0")
        total_tax = Decimal("0")
        for item in items:
            amount = Decimal(str(item.amount)) if item.amount else Decimal("0")
            tax = Decimal(str(item.tax_amount)) if item.tax_amount else Decimal("0")
            total_amount += amount
            total_tax += tax

            currency = getattr(item, 'currency', 'TWD') or 'TWD'
            orig_amt = getattr(item, 'original_amount', None)
            ex_rate = getattr(item, 'exchange_rate', None)

            ws.append([
                item.inv_num,
                str(item.date) if item.date else "",
                float(amount),
                currency,
                float(Decimal(str(orig_amt))) if orig_amt else "",
                float(Decimal(str(ex_rate))) if ex_rate else "",
                float(tax) if item.tax_amount else "",
                item.category or "",
                item.case_code or "營運",
                _SOURCE_LABELS.get(item.source, item.source),
                _STATUS_LABELS.get(item.status, item.status),
                item.notes or "",
            ])

        row_idx = len(items) + 4
        ws.cell(row=row_idx, column=1, value="合計").font = Font(bold=True)
        ws.cell(row=row_idx, column=3, value=float(total_amount)).font = Font(bold=True)
        ws.cell(row=row_idx, column=3).number_format = _NUM_FORMAT
        ws.cell(row=row_idx, column=7, value=float(total_tax)).font = Font(bold=True)
        ws.cell(row=row_idx, column=7).number_format = _NUM_FORMAT

        ws.cell(row=row_idx + 2, column=1, value=f"匯出時間: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        ws.cell(row=row_idx + 3, column=1, value=f"共 {len(items)} 筆")

        self._auto_column_width(ws, headers)

        for row in ws.iter_rows(min_row=4, max_row=row_idx - 1, min_col=3, max_col=7):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = _NUM_FORMAT
                    cell.alignment = Alignment(horizontal="right")

    async def export_ledger(
        self,
        date_from: Optional[date] = None,
        date_to: Optional[date] = None,
        case_code: Optional[str] = None,
        entry_type: Optional[str] = None,
    ) -> bytes:
        """匯出帳本收支明細 Excel"""
        from app.schemas.erp.ledger import LedgerQuery
        query = LedgerQuery(
            date_from=date_from, date_to=date_to,
            case_code=case_code, entry_type=entry_type,
            skip=0, limit=100,
        )
        items, total = await self.ledger_repo.query(query)

        wb = Workbook()
        ws = wb.active
        ws.title = "帳本收支明細"

        # 標題
        ws.merge_cells("A1:H1")
        title_cell = ws["A1"]
        title_cell.value = self._build_title("帳本收支明細表", date_from, date_to, case_code)
        title_cell.font = _TITLE_FONT
        title_cell.alignment = Alignment(horizontal="center")

        ws.append([])

        headers = ["日期", "類型", "金額", "分類", "案號", "來源類型", "描述", "備註"]
        ws.append(headers)
        for col_idx, _ in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx)
            cell.font = _HEADER_FONT_WHITE
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = _BORDER

        total_income = Decimal("0")
        total_expense = Decimal("0")
        for item in items:
            amount = Decimal(str(item.amount)) if item.amount else Decimal("0")
            if item.entry_type == "income":
                total_income += amount
            else:
                total_expense += amount

            ws.append([
                str(item.entry_date) if item.entry_date else "",
                "收入" if item.entry_type == "income" else "支出",
                float(amount),
                item.category or "",
                item.case_code or "一般營運",
                item.source_type or "",
                item.description or "",
                item.notes or "",
            ])

        # 合計
        row_idx = len(items) + 4
        ws.cell(row=row_idx, column=1, value="合計").font = Font(bold=True)
        ws.cell(row=row_idx + 1, column=1, value="總收入")
        ws.cell(row=row_idx + 1, column=3, value=float(total_income)).number_format = _NUM_FORMAT
        ws.cell(row=row_idx + 2, column=1, value="總支出")
        ws.cell(row=row_idx + 2, column=3, value=float(total_expense)).number_format = _NUM_FORMAT
        ws.cell(row=row_idx + 3, column=1, value="淨額").font = Font(bold=True)
        ws.cell(row=row_idx + 3, column=3, value=float(total_income - total_expense))
        ws.cell(row=row_idx + 3, column=3).number_format = _NUM_FORMAT
        ws.cell(row=row_idx + 3, column=3).font = Font(bold=True)

        ws.cell(row=row_idx + 5, column=1, value=f"匯出時間: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        ws.cell(row=row_idx + 6, column=1, value=f"共 {len(items)} 筆 (總計 {total} 筆符合條件)")

        self._auto_column_width(ws, headers)

        for row in ws.iter_rows(min_row=4, max_row=row_idx - 1, min_col=3, max_col=3):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = _NUM_FORMAT
                    cell.alignment = Alignment(horizontal="right")

        return self._save_to_bytes(wb)

    def _build_title(
        self, base: str,
        date_from: Optional[date], date_to: Optional[date],
        case_code: Optional[str],
    ) -> str:
        parts = [base]
        if case_code:
            parts.append(f"案號: {case_code}")
        if date_from and date_to:
            parts.append(f"{date_from} ~ {date_to}")
        elif date_from:
            parts.append(f"{date_from} 起")
        elif date_to:
            parts.append(f"至 {date_to}")
        return " — ".join(parts)

    def _auto_column_width(self, ws, headers: list) -> None:
        for col_idx, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = max(len(header) * 2 + 4, 12)

    def _save_to_bytes(self, wb: Workbook) -> bytes:
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.read()
