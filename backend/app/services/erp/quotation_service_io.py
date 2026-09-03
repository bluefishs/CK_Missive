"""ERP 報價匯出/匯入服務

Version: 1.0.0
- v1.0.0: 從 quotation_service.py 拆分 (CSV/Excel export + import)
"""
import csv
import io
import logging
import re
import unicodedata
from typing import Optional, List
from decimal import Decimal

from sqlalchemy.ext.asyncio import AsyncSession

from app.repositories.erp import ERPQuotationRepository
from app.services.erp.quotation_service import ERPQuotationService
from app.services.erp.company_profit import get_company_profit_rate

logger = logging.getLogger(__name__)


class ERPQuotationIOService:
    """報價匯出/匯入服務 — CSV/Excel IO"""

    def __init__(self, db: AsyncSession):
        self.db = db
        self.repo = ERPQuotationRepository(db)

    # =========================================================================
    # CSV 匯出
    # =========================================================================

    async def export_csv(self, year: Optional[int] = None) -> str:
        """匯出報價為 CSV 字串 (含損益計算)"""
        items, _ = await self.repo.filter_quotations(
            year=year, skip=0, limit=9999,
        )

        # 公司留成比率整批取一次（2026-08-18）——
        # 匯出的毛利必須與畫面上的一致，否則對帳時會看到兩個數字
        # 而不知道該信哪一個。
        _rate = await get_company_profit_rate(self.db)

        output = io.StringIO()
        output.write("\ufeff")  # BOM for Excel
        writer = csv.writer(output)
        writer.writerow([
            "案號", "案名", "年度", "總價", "稅額",
            "外包費", "人事費", "管銷費", "其他成本",
            "毛利", "毛利率(%)", "狀態",
        ])

        for item in items:
            profit = ERPQuotationService.compute_profit(item, _rate)
            writer.writerow([
                item.case_code or "",
                item.case_name or "",
                item.year or "",
                item.total_price or "",
                item.tax_amount or "",
                item.outsourcing_fee or "",
                item.personnel_fee or "",
                item.overhead_fee or "",
                item.other_cost or "",
                profit["gross_profit"],
                profit["gross_margin"] or "",
                item.status or "",
            ])

        return output.getvalue()

    # =========================================================================
    # Excel 匯出
    # =========================================================================

    #: 匯出欄位＝owner 的「115報價單彙整總表／系統報價單」前 29 欄（2026-09-03）。
    #: 對齊的理由：總表是真值，匯出要能直接對照、回貼；此前匯出 14 欄（成本結構）與總表對不上，
    #: 每次對帳都得人工重排。成本欄（外包／人事／管銷）另走 `export`（JSON）。
    XLS_HEADERS = [
        "序號", "年度", "承辦同仁", "報價單編號", "是否成立", "報價日期", "客戶名稱", "案名", "工作地點",
        "報價金額", "稅內含", "稅額", "總價", "實收金額", "收款日期", "配合廠商", "支出金額",
        "聯絡人", "電話", "行動電話", "傳真", "統一編號", "E-mail", "地址", "備註", "發票日期", "印花",
        "PDF報價單", "完整案名(地點＋案名)",
        # 總表 v2 追加（發票明細比對）
        "發票號碼", "銷售額", "稅額(發票)", "發票金額",
        # 系統欄（總表沒有，對帳時要）
        "系統編號", "案號", "成案編號", "種類", "狀態",
    ]

    async def export_excel(self, year: Optional[int] = None) -> bytes:
        """匯出報價為 Excel（.xlsx），欄位＝總表格式（見 XLS_HEADERS）。一次 SQL 拉齊，不逐筆查。"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from sqlalchemy import text as _text

        def _roc(d):
            return f"{d.year - 1911}.{d.month:02d}.{d.day:02d}" if d else ""

        sql = """
        SELECT q.id, q.year, q.legacy_quotation_no, q.quotation_no, q.case_code, q.project_code, q.quote_kind, q.status,
               q.case_name, q.total_price, q.tax_amount, q.quoted_at, q.notes,
               COALESCE(c.client_agency, p.client_name) AS client_name,
               p.location,
               pv.contact_person, pv.phone, pv.tax_id, pv.email, pv.address,
               (SELECT string_agg(COALESCE(u.full_name,u.username), '、')
                  FROM project_user_assignments a JOIN users u ON u.id=a.user_id
                 WHERE a.case_code=q.case_code OR a.project_id=c.id) AS staff,
               (SELECT sum(b.payment_amount) FROM erp_billings b WHERE b.erp_quotation_id=q.id AND b.payment_status='paid') AS received,
               (SELECT max(b.payment_date) FROM erp_billings b WHERE b.erp_quotation_id=q.id AND b.payment_status='paid') AS pay_date,
               (SELECT sum(v.payable_amount) FROM erp_vendor_payables v WHERE v.erp_quotation_id=q.id) AS payable,
               (SELECT string_agg(v.vendor_name, '、') FROM erp_vendor_payables v WHERE v.erp_quotation_id=q.id) AS vendors,
               inv.invoice_number, inv.invoice_date, inv.amount AS inv_amt, inv.tax_amount AS inv_tax,
               (SELECT count(*) FROM pm_case_attachments att WHERE att.case_code=q.case_code AND att.doc_type='signed_quotation') AS pdf_n
        FROM erp_quotations q
        LEFT JOIN contract_projects c ON c.case_code=q.case_code
        LEFT JOIN pm_cases p ON p.case_code=q.case_code
        LEFT JOIN partner_vendors pv ON pv.id=p.client_vendor_id
        LEFT JOIN LATERAL (SELECT i.invoice_number, i.invoice_date, i.amount, i.tax_amount FROM erp_invoices i
                           WHERE i.erp_quotation_id=q.id AND i.invoice_number NOT LIKE 'XLS-%' ORDER BY i.invoice_date LIMIT 1) inv ON true
        WHERE q.deleted_at IS NULL {year_clause}
        ORDER BY q.year, q.legacy_quotation_no NULLS LAST, q.id
        """
        # asyncpg 對 `:year IS NULL OR q.year=:year` 推不出 $1 的型別（同名參數兩處）——有年度才加條件
        sql = sql.replace("{year_clause}", "AND q.year = :year" if year else "")
        rows = (await self.db.execute(_text(sql), {"year": year} if year else {})).mappings().all()

        wb = Workbook()
        ws = wb.active
        ws.title = "系統報價單"
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for col, h in enumerate(self.XLS_HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for n, r in enumerate(rows, 1):
            tp = float(r["total_price"] or 0)
            tax = float(r["tax_amount"] or 0)
            values = [
                n, r["year"], r["staff"] or "", r["legacy_quotation_no"] or r["quotation_no"] or "",
                "v" if r["project_code"] else "", _roc(r["quoted_at"].date() if r["quoted_at"] else None),
                r["client_name"] or "", r["case_name"] or "", r["location"] or "",
                round(tp - tax) if tp else "", "v" if tax else "", round(tax) if tax else "", round(tp) if tp else "",
                float(r["received"]) if r["received"] else "", _roc(r["pay_date"]),
                r["vendors"] or "", float(r["payable"]) if r["payable"] else "",
                r["contact_person"] or "", r["phone"] or "", "", "", r["tax_id"] or "", r["email"] or "", r["address"] or "",
                r["notes"] or "", _roc(r["invoice_date"]), "",
                (r["pdf_n"] or "") if r["pdf_n"] else "", r["case_name"] or "",
                r["invoice_number"] or "", (float(r["inv_amt"]) - float(r["inv_tax"] or 0)) if r["inv_amt"] else "",
                float(r["inv_tax"]) if r["inv_tax"] else "", float(r["inv_amt"]) if r["inv_amt"] else "",
                r["id"], r["case_code"] or "", r["project_code"] or "", r["quote_kind"] or "", r["status"] or "",
            ]
            for col, v in enumerate(values, 1):
                ws.cell(row=n + 1, column=col, value=v)
        for col in range(1, len(self.XLS_HEADERS) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 14
        ws.column_dimensions["H"].width = 40
        ws.column_dimensions["AC"].width = 40
        ws.freeze_panes = "B2"
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    # =========================================================================
    # 匯入範本
    # =========================================================================

    @classmethod
    def generate_import_template(cls) -> bytes:
        """匯入範本＝匯出同一份表頭（XLS_HEADERS，owner 的總表格式）。

        2026-09-03 之前範本是另一套 11 欄（案號／案名／年度／總價／成本…），與匯出的總表格式對不上；
        使用者拿匯出檔改完要匯入還得重排。現在三者一份表頭：匯出 → 改 → 匯入可往返。
        第 2 列是填寫說明（灰字），匯入時「報價單編號」空的列會被略過，所以說明列不會被吃進去。
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = Workbook()
        ws = wb.active
        ws.title = "系統報價單"
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        for col, h in enumerate(cls.XLS_HEADERS, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill = header_fill
            c.font = Font(bold=True, color="FFFFFF")
            c.alignment = Alignment(horizontal="center")
        hints = {
            "序號": "選填", "年度": "西元，如 2026", "承辦同仁": "姓名（對 users）", "報價單編號": "",
            "是否成立": "v＝成案", "報價日期": "民國 115.03.12 或西元", "客戶名稱": "對委託單位", "案名": "若有「完整案名」以它為準。報價單編號必填：既有編號＝更新、新編號＝新增；本說明列編號留空所以不會被匯入",
            "報價金額": "未稅", "稅額": "數字", "總價": "含稅", "實收金額": "有值＝已收", "收款日期": "有值＝已收",
            "發票日期": "民國或西元", "發票號碼": "AB12345678；「比對方式」含需確認者不匿入", "銷售額": "未稅", "稅額(發票)": "數字", "發票金額": "含稅",
            "完整案名(地點＋案名)": "優先於「案名」", "系統編號": "匯出帶出，匯入忽略", "案號": "匯入忽略", "成案編號": "匯入忽略", "種類": "匯入忽略", "狀態": "匯入忽略",
        }
        for col, h in enumerate(cls.XLS_HEADERS, 1):
            c = ws.cell(row=2, column=col, value=hints.get(h, ""))
            c.font = Font(italic=True, color="999999")
        for col in range(1, len(cls.XLS_HEADERS) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 14
        ws.column_dimensions["H"].width = 40
        ws.column_dimensions["AC"].width = 40
        ws.freeze_panes = "B3"
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    # =========================================================================
    # Excel 匯入
    # =========================================================================

    async def import_from_excel(self, file_bytes: bytes, user_id: Optional[int] = None) -> dict:
        """匯入報價 Excel — 用 case_code 做 upsert"""
        from app.services.base.excel_reader import load_workbook_any

        wb = load_workbook_any(file_bytes)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))

        created = 0
        updated = 0
        errors: list = []

        def _num(v) -> Decimal:
            if v is None:
                return Decimal("0")
            if isinstance(v, (int, float)):
                return Decimal(str(v))
            s = re.sub(r'[NT$￥,\s]', '', str(v).strip())
            return Decimal(s) if s else Decimal("0")

        def _str(v) -> Optional[str]:
            if v is None:
                return None
            return unicodedata.normalize('NFKC', str(v).strip()) or None

        for idx, row in enumerate(rows, start=2):
            try:
                if not row or len(row) < 3:
                    continue
                case_code = _str(row[0])
                case_name = _str(row[1])
                if not case_code or not case_name:
                    continue

                year_val = row[2]
                if isinstance(year_val, (int, float)):
                    year = int(year_val)
                else:
                    year = int(str(year_val).strip()) if year_val else None
                # 民國年自動轉西元
                if year and year < 1911:
                    year = year + 1911

                data = {
                    "case_code": case_code,
                    "case_name": case_name,
                    "year": year,
                    "total_price": _num(row[3]) if len(row) > 3 else Decimal("0"),
                    "tax_amount": _num(row[4]) if len(row) > 4 else Decimal("0"),
                    "outsourcing_fee": _num(row[5]) if len(row) > 5 else Decimal("0"),
                    "personnel_fee": _num(row[6]) if len(row) > 6 else Decimal("0"),
                    "overhead_fee": _num(row[7]) if len(row) > 7 else Decimal("0"),
                    "other_cost": _num(row[8]) if len(row) > 8 else Decimal("0"),
                    "status": _str(row[9]) or "draft" if len(row) > 9 else "draft",
                    "notes": _str(row[10]) if len(row) > 10 else None,
                }

                # Upsert by case_code
                existing = await self.repo.get_by_case_code(case_code)
                if existing:
                    update_data = {k: v for k, v in data.items() if k != "case_code" and v is not None}
                    await self.repo.update(existing.id, update_data)
                    updated += 1
                else:
                    data["created_by"] = user_id
                    await self.repo.create(data)
                    created += 1

            except Exception as e:
                errors.append({"row": idx, "error": str(e)})

        if created > 0 or updated > 0:
            await self.db.commit()

        return {
            "total_rows": len(rows),
            "created": created,
            "updated": updated,
            "errors": errors,
        }
