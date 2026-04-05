"""資產管理業務服務層"""
import logging
from typing import Optional, Tuple, List, Dict, Any
from decimal import Decimal

from sqlalchemy.ext.asyncio import AsyncSession

from app.extended.models.asset import Asset, AssetLog
from app.repositories.erp.asset_repository import AssetRepository, AssetLogRepository
from app.schemas.erp.asset import (
    AssetCreateRequest,
    AssetUpdateRequest,
    AssetListRequest,
    AssetLogCreateRequest,
    AssetLogListRequest,
)
from app.services.audit_mixin import AuditableServiceMixin

logger = logging.getLogger(__name__)


class AssetService(AuditableServiceMixin):
    """資產管理業務服務"""

    AUDIT_TABLE = "assets"

    def __init__(self, db: AsyncSession):
        self.db = db
        self.repo = AssetRepository(db)
        self.log_repo = AssetLogRepository(db)

    async def list_assets(self, params: AssetListRequest) -> Tuple[List[Asset], int]:
        """資產列表"""
        return await self.repo.list_assets(
            category=params.category,
            status=params.status,
            keyword=params.keyword,
            case_code=params.case_code,
            skip=params.skip,
            limit=params.limit,
        )

    async def get_asset(self, asset_id: int) -> Optional[Asset]:
        """取得資產詳情"""
        return await self.repo.get_asset(asset_id)

    async def create_asset(
        self, data: AssetCreateRequest, user_id: Optional[int] = None
    ) -> Asset:
        """建立資產 (asset_code 為空時自動生成)"""
        # 自動生成 asset_code (ADR-0013 Phase 1)
        if not data.asset_code:
            from datetime import date as _date
            from app.services.case_code_service import CaseCodeService
            code_svc = CaseCodeService(self.db)
            data.asset_code = await code_svc.generate_asset_code(
                year=_date.today().year,
                category=data.category or "equipment",
            )

        # Check duplicate code
        if await self.repo.check_code_exists(data.asset_code):
            raise ValueError(f"資產編號 {data.asset_code} 已存在")

        asset_data = data.model_dump()
        # Map schema field to ORM column
        if "asset_model" in asset_data:
            asset_data["model"] = asset_data.pop("asset_model")
        if data.current_value is None:
            asset_data["current_value"] = data.purchase_amount
        asset_data["created_by"] = user_id

        asset = await self.repo.create_asset(asset_data)
        await self.db.commit()

        await self.audit_create(asset.id, asset_data, user_id=user_id)

        return asset

    async def update_asset(
        self, params: AssetUpdateRequest, user_id: Optional[int] = None
    ) -> Optional[Asset]:
        """更新資產"""
        update_data = params.model_dump(exclude={"id"}, exclude_none=True)
        if "asset_model" in update_data:
            update_data["model"] = update_data.pop("asset_model")
        if not update_data:
            return await self.repo.get_asset(params.id)

        asset = await self.repo.update_asset(params.id, update_data)
        if asset:
            await self.db.commit()
            await self.audit_update(params.id, update_data, user_id=user_id)
        return asset

    async def delete_asset(self, asset_id: int, user_id: Optional[int] = None) -> bool:
        """刪除資產"""
        success = await self.repo.delete_asset(asset_id)
        if success:
            await self.db.commit()
            await self.audit_delete(asset_id, user_id=user_id)
        return success

    async def get_assets_by_invoice(self, expense_invoice_id: int) -> List[Asset]:
        """取得關聯到指定發票的所有資產"""
        return await self.repo.get_assets_by_invoice(expense_invoice_id)

    async def get_asset_with_relations(self, asset_id: int) -> Optional[Dict[str, Any]]:
        """取得資產完整詳情 (含關聯發票+案件)"""
        return await self.repo.get_asset_with_invoice(asset_id)

    async def get_stats(self) -> Dict[str, Any]:
        """取得資產統計"""
        return await self.repo.get_asset_stats()

    async def _get_project_code_map(self) -> Dict[str, str]:
        """case_code→project_code 映射表"""
        from sqlalchemy import select as sa_select
        from app.extended.models.erp import ERPQuotation
        result = await self.db.execute(
            sa_select(ERPQuotation.case_code, ERPQuotation.project_code)
            .where(ERPQuotation.project_code.isnot(None))
        )
        return {r[0]: r[1] for r in result.all()}

    async def _get_reverse_project_code_map(self) -> Dict[str, str]:
        """project_code→case_code 反查映射表"""
        from sqlalchemy import select as sa_select
        from app.extended.models.erp import ERPQuotation
        result = await self.db.execute(
            sa_select(ERPQuotation.project_code, ERPQuotation.case_code)
            .where(ERPQuotation.project_code.isnot(None))
        )
        return {r[0]: r[1] for r in result.all()}

    @staticmethod
    def _resolve_case_code(value: Optional[str], reverse_map: Dict[str, str]) -> Optional[str]:
        """解析使用者輸入的編號：若為 project_code 則反查 case_code，否則直接使用"""
        if not value:
            return None
        # 如果是 project_code 格式，反查 case_code
        if value in reverse_map:
            return reverse_map[value]
        # 否則當作 case_code 直接使用
        return value

    async def export_assets_excel(
        self,
        category: Optional[str] = None,
        status: Optional[str] = None,
    ) -> bytes:
        """匯出資產清單 Excel"""
        import io
        from openpyxl import Workbook

        items, _ = await self.repo.list_assets(
            category=category, status=status, skip=0, limit=10000
        )

        # 建立 case_code→project_code 映射
        code_map = await self._get_project_code_map()

        wb = Workbook()
        ws = wb.active
        ws.title = "資產清單"

        headers = [
            "資產編號", "名稱", "類別", "品牌", "型號", "序號",
            "購入日期", "購入金額", "目前價值", "狀態", "存放位置",
            "保管人", "成案編號", "備註",
        ]
        ws.append(headers)

        for item in items:
            ws.append([
                item.asset_code,
                item.name,
                item.category,
                item.brand,
                getattr(item, "model", None),
                item.serial_number,
                str(item.purchase_date) if item.purchase_date else "",
                float(item.purchase_amount or 0),
                float(item.current_value or 0),
                item.status,
                item.location,
                item.custodian,
                code_map.get(item.case_code, item.case_code) if item.case_code else "",
                item.notes,
            ])

        # Auto-width
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.read()

    async def import_assets_excel(self, file_bytes: bytes, user_id: Optional[int] = None) -> dict:
        """匯入資產清單 Excel — 用 asset_code 做 upsert (header 驅動匹配)"""
        from app.services.base.excel_reader import load_workbook_any

        wb = load_workbook_any(file_bytes)
        ws = wb.active

        # ---- Header 驅動欄位映射 ----
        header_row = [str(c.value or "").strip() for c in ws[1]]
        # 標準欄位名→內部 key 映射 (支持多種中英文名稱)
        HEADER_MAP: Dict[str, str] = {
            "資產編號": "asset_code", "asset_code": "asset_code",
            "名稱": "name", "name": "name",
            "類別": "category", "category": "category",
            "品牌": "brand", "brand": "brand",
            "型號": "model", "model": "model", "asset_model": "model",
            "序號": "serial_number", "serial_number": "serial_number",
            "購入日期": "purchase_date", "purchase_date": "purchase_date",
            "購入金額": "purchase_amount", "purchase_amount": "purchase_amount",
            "目前價值": "current_value", "current_value": "current_value",
            "狀態": "status", "status": "status",
            "存放位置": "location", "location": "location",
            "保管人": "custodian", "custodian": "custodian",
            "案件代碼": "case_code", "成案編號": "case_code",
            "case_code": "case_code", "project_code": "case_code",
            "備註": "notes", "notes": "notes",
        }
        import unicodedata
        col_map: Dict[str, int] = {}  # internal_key → column_index
        for i, h in enumerate(header_row):
            normalized = unicodedata.normalize('NFKC', h)
            key = HEADER_MAP.get(normalized)
            if key and key not in col_map:
                col_map[key] = i

        logger.info(f"資產匯入: 工作表='{ws.title}', header={header_row}, 映射={col_map}")

        rows = list(ws.iter_rows(min_row=2, values_only=True))

        created = 0
        updated = 0
        skipped = 0
        errors: List[Dict[str, Any]] = []

        logger.info(f"資產匯入: 資料列數={len(rows)}")

        # 建立 project_code→case_code 反查映射
        reverse_map = await self._get_reverse_project_code_map()

        CATEGORY_MAP = {
            "設備": "equipment", "車輛": "vehicle", "儀器": "instrument",
            "家具": "furniture", "其他": "other",
        }
        STATUS_MAP = {
            "使用中": "in_use", "維修中": "maintenance", "閒置": "idle",
            "已報廢": "disposed", "遺失": "lost",
        }

        import re

        def _nfkc(v) -> Optional[str]:
            if v is None: return None
            return unicodedata.normalize('NFKC', str(v).strip()) or None

        def _num(v) -> float:
            if v is None: return 0
            if isinstance(v, (int, float)): return float(v)
            s = re.sub(r'[NT$￥,\s]', '', str(v).strip())
            return float(s) if s else 0

        def _get(row_data, key: str):
            """根據 header 映射取得欄位值"""
            idx = col_map.get(key)
            if idx is None or idx >= len(row_data):
                return None
            return row_data[idx]

        # 必要欄位檢查
        if "asset_code" not in col_map or "name" not in col_map:
            logger.warning(f"資產匯入: header 缺少必填欄位 (asset_code/name), 映射={col_map}")
            wb.close()
            return {"total_rows": 0, "created": 0, "updated": 0, "skipped": 0,
                    "errors": [{"row": 1, "error": "Header 缺少必填欄位「資產編號」或「名稱」"}]}

        for idx, row in enumerate(rows, start=2):
            try:
                ac = _get(row, "asset_code")
                nm = _get(row, "name")
                if not ac or not nm:
                    skipped += 1
                    continue

                asset_code = _nfkc(ac) or ""
                raw_cat = _nfkc(_get(row, "category"))
                raw_status = _nfkc(_get(row, "status"))

                data: Dict[str, Any] = {
                    "asset_code": asset_code,
                    "name": _nfkc(nm) or "",
                    "category": CATEGORY_MAP.get(raw_cat or "", raw_cat or "equipment") if raw_cat else "equipment",
                    "brand": _nfkc(_get(row, "brand")),
                    "model": _nfkc(_get(row, "model")),
                    "serial_number": _nfkc(_get(row, "serial_number")),
                    "purchase_amount": _num(_get(row, "purchase_amount")),
                    "current_value": _num(_get(row, "current_value")) if _get(row, "current_value") else None,
                    "status": STATUS_MAP.get(raw_status or "", raw_status or "in_use") if raw_status else "in_use",
                    "location": _nfkc(_get(row, "location")),
                    "custodian": _nfkc(_get(row, "custodian")),
                    "case_code": self._resolve_case_code(_nfkc(_get(row, "case_code")), reverse_map),
                    "notes": _nfkc(_get(row, "notes")),
                }

                # Handle purchase_date
                raw_date = _get(row, "purchase_date")
                if raw_date:
                    from datetime import date as date_type, datetime as datetime_type
                    if isinstance(raw_date, (date_type, datetime_type)):
                        data["purchase_date"] = raw_date if isinstance(raw_date, date_type) else raw_date.date()
                    else:
                        try:
                            data["purchase_date"] = datetime_type.strptime(str(raw_date).strip(), "%Y-%m-%d").date()
                        except ValueError:
                            pass

                # Upsert by asset_code
                existing = await self.repo.get_by_code(asset_code)
                if existing:
                    update_data = {k: v for k, v in data.items() if k != "asset_code" and v is not None}
                    await self.repo.update_asset(existing.id, update_data)
                    updated += 1
                else:
                    data["created_by"] = user_id
                    if data.get("current_value") is None:
                        data["current_value"] = data.get("purchase_amount", 0)
                    await self.repo.create_asset(data)
                    created += 1

            except Exception as e:
                logger.warning(f"資產匯入 row {idx} 錯誤: {type(e).__name__}: {e}")
                errors.append({"row": idx, "error": str(e)})

        await self.db.commit()
        wb.close()

        logger.info(f"資產匯入完成: total={len(rows)}, created={created}, updated={updated}, skipped={skipped}, errors={len(errors)}")
        return {
            "total_rows": len(rows),
            "created": created,
            "updated": updated,
            "skipped": skipped,
            "errors": errors,
        }

    async def batch_inventory(
        self,
        asset_ids: List[int],
        operator: str,
        notes: Optional[str] = None,
        user_id: Optional[int] = None,
    ) -> dict:
        """批次盤點 — 為多個資產建立 inspect 行為紀錄"""
        from datetime import date as date_type

        today = date_type.today()
        created = 0
        not_found: List[int] = []

        for asset_id in asset_ids:
            asset = await self.repo.get_asset(asset_id)
            if not asset:
                not_found.append(asset_id)
                continue

            log_data = {
                "asset_id": asset_id,
                "action": "inspect",
                "action_date": today,
                "description": f"盤點確認 — {asset.name}",
                "operator": operator,
                "notes": notes,
                "created_by": user_id,
            }
            await self.log_repo.create_asset_log(log_data)
            created += 1

        await self.db.commit()
        return {
            "total": len(asset_ids),
            "inspected": created,
            "not_found": not_found,
            "date": str(today),
            "operator": operator,
        }

    async def export_inventory_report(self) -> bytes:
        """匯出盤點報表 Excel — 所有資產 + 最近盤點日期"""
        import io
        from openpyxl import Workbook
        from sqlalchemy import select, func

        # Get all assets
        items, _ = await self.repo.list_assets(skip=0, limit=10000)

        # Get latest inspect log per asset
        latest_inspect = (
            select(
                AssetLog.asset_id,
                func.max(AssetLog.action_date).label("last_inspect"),
            )
            .where(AssetLog.action == "inspect")
            .group_by(AssetLog.asset_id)
        ).subquery()

        result = await self.db.execute(
            select(latest_inspect.c.asset_id, latest_inspect.c.last_inspect)
        )
        inspect_map = {r.asset_id: r.last_inspect for r in result.all()}

        wb = Workbook()
        ws = wb.active
        ws.title = "盤點報表"

        headers = [
            "資產編號", "名稱", "類別", "品牌", "型號",
            "狀態", "存放位置", "保管人", "購入金額", "最近盤點日",
        ]
        ws.append(headers)

        for item in items:
            last = inspect_map.get(item.id)
            ws.append([
                item.asset_code,
                item.name,
                item.category,
                item.brand,
                getattr(item, "model", None),
                item.status,
                item.location,
                item.custodian,
                float(item.purchase_amount or 0),
                str(last) if last else "未盤點",
            ])

        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.read()

    def generate_import_template(self) -> bytes:
        """產生資產匯入範本 Excel"""
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "資產匯入範本"

        headers = [
            "資產編號", "名稱", "類別", "品牌", "型號", "序號",
            "購入日期", "購入金額", "目前價值", "狀態", "存放位置",
            "保管人", "成案編號", "備註",
        ]

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        # Example rows (成案編號欄位支持 project_code 或 case_code)
        examples = [
            ["AST-2026-001", "全測站 ES-105", "儀器", "Topcon", "ES-105", "SN20250001",
             "2025-06-15", 280000, 280000, "使用中", "公司倉庫", "王技師", "", ""],
            ["AST-2026-002", "GPS RTK", "儀器", "Trimble", "R12i", "SN20250002",
             "2025-01-10", 450000, 400000, "使用中", "工地A", "李技師", "CK2025_PJ_01_001", ""],
        ]
        for row_data in examples:
            ws.append(row_data)

        # Notes sheet
        ws2 = wb.create_sheet("說明")
        ws2.append(["欄位", "說明", "必填", "可選值"])
        notes = [
            ["資產編號", "唯一識別碼", "是", "自訂格式"],
            ["名稱", "資產名稱", "是", ""],
            ["類別", "設備類型", "否", "設備/車輛/儀器/家具/其他"],
            ["狀態", "使用狀態", "否", "使用中/維修中/閒置/已報廢/遺失"],
            ["購入日期", "格式 YYYY-MM-DD", "否", ""],
            ["購入金額", "數字", "否", ""],
        ]
        for n in notes:
            ws2.append(n)

        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    async def list_logs(self, params: AssetLogListRequest) -> Tuple[List[AssetLog], int]:
        """取得資產行為紀錄列表"""
        return await self.log_repo.list_asset_logs(
            asset_id=params.asset_id,
            action=params.action,
            skip=params.skip,
            limit=params.limit,
        )

    async def create_log(
        self, data: AssetLogCreateRequest, user_id: Optional[int] = None
    ) -> AssetLog:
        """建立行為紀錄"""
        # Verify asset exists
        asset = await self.repo.get_asset(data.asset_id)
        if not asset:
            raise ValueError(f"資產 ID {data.asset_id} 不存在")

        log_data = data.model_dump()
        log_data["created_by"] = user_id

        # Side effects based on action
        if data.action == "transfer" and data.to_location:
            await self.repo.update_asset(data.asset_id, {"location": data.to_location})
        elif data.action == "dispose":
            await self.repo.update_asset(data.asset_id, {"status": "disposed"})
        elif data.action == "repair":
            # Only set maintenance if currently in_use
            if asset.status == "in_use":
                await self.repo.update_asset(data.asset_id, {"status": "maintenance"})

        log_entry = await self.log_repo.create_asset_log(log_data)
        await self.db.commit()
        return log_entry
