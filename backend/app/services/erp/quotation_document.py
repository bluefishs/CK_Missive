"""報價單正式文件輸出 —— 取料層（版面待 owner 提供範本）。

owner 2026-08-17：
  ① 「新增報價單要可輸出正式文件，非僅資料列表用途」
  ② 「明日提供報價單範本與格式再行評估設計，**先完成整體流程與規劃**」
  ③ 「目前報價單採用 **xls** 格式」

# 為什麼這個檔只有取料、沒有版面

我第一版把版面寫成 docx（自己編了抬頭、欄位順序、簽署欄）。
在 owner 說出 ②③ 之後那份版面必然要丟掉 —— 而丟掉的不只是排版，
是「我以為報價單長什麼樣」的一整套假設（我連公司名都是猜的）。

所以拆成兩層，只做現在做得準的那一層：

    _gather()   取料 —— 與範本完全無關，範本換了它不用改
    render_*()  版面 —— 待範本，明日補

這樣明天拿到那份 xls 時要做的是「把值填進既有格子」，
而不是「把我編的版面改成你們的版面」。

# 格式：xls（openpyxl）不是 docx

owner 明講目前用 xls。`openpyxl` 專案已在 5 處使用（asset／expense／
finance 匯出），沿用同一套；`python-docx` 目前只用來**讀**附件、
從未用來產生文件，導入寫入用法會多一條沒有既有實踐的路。

⚠️ 若那份範本是 **.xls（BIFF，Excel 97-2003）** 而非 .xlsx，
openpyxl **讀不了也寫不了** —— 需先確認副檔名實際格式，
別假設「xls」是口語泛稱。這件事明天拿到檔案第一步就要驗，
因為它決定用 openpyxl 還是得另找套件（而後者要加依賴）。

# 取料的取捨（這些與範本無關，已定）

* 委託單位／工作地點／類別**沿用 `quotation_service` 同一條取法**
  （承攬案件 `client_agency` 優先、其次 PM `client_name`）——
  不另寫一份，否則文件上的委託單位會與畫面上的各自演化。
* 明細小計與報價總價**不強制相等**：議價後可能只調總價而未逐項回改，
  兩個數字都給出去，讓落差在版面上看得見，而不是挑一個顯示。
* 金額 None 與 0 分開回傳：「沒填」印成 0 會被讀成免費。
"""
from __future__ import annotations

import io
import logging
from datetime import date, datetime
from pathlib import Path
from decimal import Decimal
from typing import Any, Optional

from sqlalchemy import select, text
from sqlalchemy.ext.asyncio import AsyncSession

# ⚠️ `ERPQuotation` 在 `models.erp` 不在 `models.invoice`（我第一版寫錯了）。
# 這個錯誤的形狀正是 v6.56 記錄的病灶家族「模組匯入即失敗，而沒有人在匯入它」：
# py_compile 過、型別檢查過、測試不碰它，要等真的有人打匯出端點才會爆。
# 抓到它的是 daily step 12 `module_import_sweep`（真的把每個模組匯入一次）
# —— 本次是我自己手動 import 時撞到，比等到明天早一天。
from app.extended.models.erp import ERPQuotation
from app.services.contract.case_code import CaseCodeService

logger = logging.getLogger(__name__)

ZERO = Decimal("0")

# 報價有效期限（天）。業界慣例 30 天；先寫死是因為目前沒有任何地方在存它，
# 憑空開一個設定欄位只會多一個沒人維護的數字。範本若有這一格再對齊。
VALID_DAYS = 30


def roc_date(d: Optional[date | datetime]) -> str:
    """民國年月日 —— 客戶手上其他文件都是這樣寫的。"""
    if not d:
        return ""
    dd = d.date() if isinstance(d, datetime) else d
    return f"{dd.year - 1911}年{dd.month:02d}月{dd.day:02d}日"


class QuotationDocumentService:
    """把一張報價整理成「可填入範本」的資料包。

    刻意回 dict 而不是直接回檔案 bytes：
    範本還沒到，而取料是現在就能定案並被測試的部分。
    """

    def __init__(self, db: AsyncSession):
        self.db = db

    async def gather(self, quotation_id: int) -> dict[str, Any]:
        """收齊一張報價單所需的全部欄位。

        Returns 的 key 皆為業務語意名（`display_no`／`client_name`／`items`…），
        **不含任何版面概念**（沒有 row/col、沒有字型）——
        版面在 renderer 那一層，換範本不動這裡。
        """
        q = (await self.db.execute(
            select(ERPQuotation).where(
                ERPQuotation.id == quotation_id,
                ERPQuotation.deleted_at.is_(None),
            )
        )).scalar_one_or_none()
        if not q:
            raise ValueError(f"找不到報價 {quotation_id}")

        row = (await self.db.execute(text("""
            SELECT COALESCE(cp.client_agency, pm.client_name) AS client_name,
                   cp.location,
                   cp.agency_contact_person,
                   ga.tax_id          AS client_tax_id,
                   ga.phone           AS client_phone,
                   ga.address         AS client_address,
                   pc.contact_name    AS contact_name,
                   pc.phone           AS contact_phone,
                   pc.mobile          AS contact_mobile,
                   pc.email           AS contact_email,
                   su.staff_name      AS staff_name,
                   su.staff_email     AS staff_email,
                   COALESCE(cp.category, pm.category)         AS category
              FROM (SELECT :cc AS cc) x
              -- ⚠️ contract_projects 與 pm_cases **都沒有 deleted_at**
              -- （只有 erp_quotations 有）。第一版我照著習慣加了
              -- `AND cp.deleted_at IS NULL`，實跑當場 UndefinedColumn ——
              -- 而那條 SQL 走的是文件輸出這條偶爾才跑的路徑，
              -- 若沒有實跑就交付，症狀會是「按匯出就 500」而測試全綠。
              LEFT JOIN contract_projects cp ON cp.case_code = x.cc
              LEFT JOIN pm_cases pm          ON pm.case_code = x.cc
              -- 2026-08-18：範本（`app/templates/quotation_template.xlsx`）
              -- 要的欄位比原本多，逐一接上真實來源，不留空格：
              --   統一編號／聯絡電話／傳真／地址 → 機關主檔
              --   聯絡人手機／E-mail              → 該案的機關承辦
              --   服務人員姓名／E-mail            → 專案負責人（is_primary 優先）
              LEFT JOIN government_agencies ga ON ga.id = cp.client_agency_id
              LEFT JOIN LATERAL (
                  SELECT pc.contact_name, pc.phone, pc.mobile, pc.email
                    FROM project_agency_contacts pc
                   WHERE pc.project_id = cp.id
                   ORDER BY pc.is_primary DESC NULLS LAST, pc.id
                   LIMIT 1
              ) pc ON TRUE
              LEFT JOIN LATERAL (
                  SELECT COALESCE(u.full_name, u.username) AS staff_name, u.email AS staff_email
                    FROM project_user_assignments pa
                    LEFT JOIN users au ON au.id = pa.user_id
                    -- ADR-0025：以 canonical 人為準，分身帳號不得顯示成另一個人
                    LEFT JOIN users u ON u.id = COALESCE(au.canonical_user_id, au.id)
                   -- 2026-08-21：也吃 `pa.case_code`。
                   --
                   -- 原本只比對 `pa.project_id = cp.id` —— 那要求案件已經成案
                   -- （有 contract_project）。而從彙整表匯入的 179 件邀標案件
                   -- 多數還沒成案，於是正式報價單的「服務人員」永遠是空白，
                   -- 即使承辦同仁明明已經指派好了。
                   --
                   -- `project_user_assignments` 本來就有 case_code 欄位
                   -- （`/project-staff/case/{case_code}/list` 端點註解寫著
                   --  「支援未成案的 PM 案件」）—— 資料一直都在，是這條 JOIN
                   -- 只認其中一條路。
                   WHERE pa.project_id = cp.id OR pa.case_code = x.cc
                   ORDER BY pa.is_primary DESC NULLS LAST, pa.id
                   LIMIT 1
              ) su ON TRUE
             LIMIT 1
        """), {"cc": q.case_code})).mappings().first() or {}

        items = (await self.db.execute(text("""
            SELECT item_name, spec, unit, qty, unit_price, amount, notes
              FROM erp_quotation_items
             WHERE quotation_id = :qid
             ORDER BY sort_order, id
        """), {"qid": quotation_id})).mappings().all()

        subtotal = sum((Decimal(str(r["amount"] or 0)) for r in items), ZERO)
        total = Decimal(str(q.total_price)) if q.total_price is not None else None
        tax = Decimal(str(q.tax_amount or 0))
        quoted = q.quoted_at or q.created_at

        return {
            "quotation_id": q.id,
            # 含版次後綴的對外單號（QT2026_018-2）；未編號時回「未編號」，
            # 不回空字串 —— 正式文件上留白會被當成漏印。
            #
            # 2026-08-20：沒有系統單號時**改用舊案號**（B114-A016-3）。
            # 179 張從彙整表匯入的報價單本來就沒有系統單號，但它們有舊案號 ——
            # 而那正是客戶手上那張紙、以及回簽 PDF 檔名用的編號。
            # 印「未編號」不只難看，是**印了一個比實際更少資訊的值**。
            "display_no": (
                CaseCodeService.format_quotation_no(q.quotation_no, q.revision or 1)
                if q.quotation_no
                else (q.legacy_quotation_no or CaseCodeService.format_quotation_no(None, 1))
            ),
            "quotation_no": q.quotation_no,
            "revision": q.revision or 1,
            "case_code": q.case_code,
            "case_name": q.case_name,
            "year": q.year,
            "quoted_date_roc": roc_date(quoted),
            "quoted_date": quoted,
            "valid_days": VALID_DAYS,
            "client_name": row.get("client_name"),
            "location": row.get("location"),
            "contact_person": row.get("contact_name") or row.get("agency_contact_person"),
            # 以下 2026-08-18 新增 —— 範本上有格子的欄位一律給值，
            # 取不到就給 None，由 renderer 決定留白還是印「—」。
            "client_tax_id": row.get("client_tax_id"),
            "client_phone": row.get("client_phone"),
            # 範本有「傳真號碼」格子，但 `government_agencies` **沒有 fax 欄位**。
            # 不為了填滿版面而發明資料來源 —— 留白比印一個錯的號碼好。
            "client_fax": None,
            "client_address": row.get("client_address"),
            "contact_phone": row.get("contact_phone"),
            "contact_mobile": row.get("contact_mobile"),
            "contact_email": row.get("contact_email"),
            "staff_name": row.get("staff_name"),
            "staff_email": row.get("staff_email"),
            # '01' 委辦招標（標案，無明細）／'02' 承攬報價（逐項單價）
            "category": row.get("category") or "",
            "items": [dict(r) for r in items],
            "items_subtotal": subtotal,
            "tax_amount": tax,
            "total_price": total,
            # 兩者不一致時 renderer 應在文件上標出，不得只挑一個顯示
            "amount_mismatch": (
                total is not None and abs(subtotal - total) > Decimal("0.01")
                and len(items) > 0
            ),
            "notes": q.notes,
            "has_items": len(items) > 0,
        }

    def suggest_filename(self, data: dict[str, Any], ext: str = "xlsx") -> str:
        """建議下載檔名。ext 待範本確認（.xls 與 .xlsx 是兩種格式）。"""
        name = (data.get("case_name") or data.get("case_code") or "報價單").strip()
        for ch in '\\/:*?"<>|\r\n\t':
            name = name.replace(ch, "_")
        return f"報價單_{data['display_no']}_{name[:40]}.{ext}"

    # ── 版面層（2026-08-18 owner 提供範本後實作）──────────────────────
    #
    # 範本＝`app/templates/quotation_template.xlsx`，取自
    # `Z:\03.專案管控專區\01.報價單紀錄\報價單_B115-D004-0_….xlsx`。
    #
    # **以範本為底填值，不用程式重畫版面。** 理由：
    #   · 範本內含 2 張圖片（公司抬頭）、21 組合併儲存格、框線與字型
    #   · 合計公式已寫在裡面（`E26=SUM(F16:F25)`、稅額 5%、總計）
    #   · 日後版面要改，是換一個檔案，不是改程式
    #
    # 動手前實測過 openpyxl 讀寫的往返保真：
    # 圖片 2→2、合併格 21→21、公式保留、字型/框線/欄寬/列印範圍皆存活。
    # （若沒驗就寫，最可能的失敗是「檔案產生了、打開發現 logo 不見了」。）

    #: 範本各欄位的位置。**寫在同一處**，範本改版時只改這張表。
    #:
    #: 版面規律（實測範本得出，不是猜的）：
    #:   **標籤在 A 欄與 C:D 合併格；值在 B 欄與 E 欄。**
    #:
    #: ⚠️ 我第一版把統編／工程編號／手機／E-mail 的值放在 D 欄 ——
    #: 而 `C7:D7` 是合併格且裡面裝的是**標籤**（「統一編號：」），
    #: openpyxl 直接 `AttributeError: MergedCell ... read-only`。
    #: 實跑當場擋下；若只讀程式碼不跑，這會變成「按匯出就 500」。
    #: 對照 `E12=賴柏霖`／`E13=email` 才確認值欄位在 E。
    CELLS = {
        "quotation_no": "F4",       # 報價單編號
        "quoted_date": "F5",        # 報價日期（民國）
        "client_name": "B7",        # 客戶名稱
        "client_tax_id": "E7",      # 統一編號（C7:D7 是標籤）
        "contact_person": "B8",     # 聯絡人
        "case_code": "E8",          # 工程編號 ← 用內部案號
        "contact_phone": "B9",      # 聯絡電話
        "contact_mobile": "E9",     # 手機號碼
        "client_fax": "B10",        # 傳真號碼（無資料來源，留白）
        "contact_email": "E10",     # E-mail
        "client_address": "B11",    # 聯絡地址
        "case_name": "B12",         # 計畫名稱
        "staff_name": "E12",        # 服務人員
        "location": "B13",          # 工作地點
        "staff_email": "E13",       # 服務人員 E-mail
        "notes": "B29",             # 備註（2026-08-29 由 B21 下移，見下方說明）
    }

    #: 明細列範圍。
    #:
    #: ⚠️ 2026-08-20 由 25 改為 20。範本的 `SUM(F16:F25)` 涵蓋 10 列沒錯，
    #: **但第 21 列是範本的備註列**（`A21="備註："` / `B21=備註內容`）——
    #: 範本樣本只有 3 項明細，作者就把備註放在那裡了。
    #: 於是那 10 項容量是假的，實際可用只有 5 項。
    #:
    #: ⭐ 2026-08-29 恢復為 25 —— **備註改印在合計之後的第 29 列**。
    #:
    #: 外部評估建議用 `ws.insert_rows()` 動態插列突破上限。**沒有採用**：
    #: openpyxl 的 `insert_rows` 不搬移合併儲存格與圖片、也不更新公式參照，
    #: 而本範本有 **2 張圖片、12 組合併儲存格**、合計三式引用 F16:F25。
    #: 那個作法會在客戶收到的檔案上壞掉，而且不會報錯。
    #:
    #: 實查範本後發現根本不需要插列：**列 22~25 本來就是空的**，
    #: 卡住的只有備註列的位置。把備註移到列 29（合計 28 與生效條款 30 之間，
    #: 原本是空列，語意上也該在那裡）⇒ 明細 5 項 → **10 項**，
    #: 版面、公式、合併儲存格一律不動。
    #:
    #: ⚠️ 超過 10 項仍會拋錯（不靜靜截斷）。要再更多就得真的改範本版面，
    #: 那要 owner 決定 —— 不在程式裡假裝有更多列。
    ITEM_FIRST_ROW = 16
    ITEM_LAST_ROW = 25
    #: 備註列 —— 不得被明細或清空迴圈碰到。
    #: 2026-08-29 由 21 移到 29（合計之後）；範本的 A29/B29 原為空白，
    #: 故 `_write_notes_label` 會補上「備註：」標籤並沿用 A21 的字型。
    NOTES_ROW = 29
    #: 範本**原本**的備註列。搬走之後它落在明細區裡（16~25），
    #: 兩個用途：① 新備註標籤沿用它的字型（那是 owner 提供的版面，不自己挑）
    #: ② 範本殘留的舊備註要清掉，否則明細不足 6 項時該列會同時出現空明細與舊備註。
    #: ⚠️ 不是「還能再用一次的座標」—— 下次再搬列時這三個常數要一起看。
    LEGACY_NOTES_ROW = 21

    #: 項次的中文數字（範本用「一、二、三、」）
    _CN = "一二三四五六七八九十"

    @staticmethod
    def _set(ws, coord: str, value: Any) -> None:
        """寫值，並清掉範本殘留在該格的超連結。

        ⚠️ openpyxl 對「有 hyperlink 但 value=None」的儲存格，存檔時會把
        hyperlink 的 target 當成顯示值寫出去 —— 於是「清空一個欄位」的結果
        是印出 `mailto:xxx@gmail.com`。

        範本 E13（服務人員 E-mail）就有這樣一個殘留連結，指向範本原主的信箱；
        在服務人員取不到值時，它會被印在**每一份**送給客戶的報價單上。
        2026-08-18 實測：`ws['E13']=None` 存檔後讀回，value 變成
        `'mailto:david790707@gmail.com'`。清值必須連 hyperlink 一起清。
        """
        cell = ws[coord]
        cell.value = value
        cell.hyperlink = None

    def render_xlsx(self, data: dict[str, Any]) -> bytes:
        """把取料結果填進範本，回傳 xlsx bytes。"""
        from copy import copy
        from openpyxl import load_workbook

        tpl = Path(__file__).resolve().parents[2] / "templates" / "quotation_template.xlsx"
        if not tpl.exists():
            # 明確失敗而不是產生一個沒有版面的檔案 ——
            # 「檔案下載成功但長得不對」比下載失敗更難查。
            raise FileNotFoundError(f"找不到報價單範本：{tpl}")

        wb = load_workbook(tpl)
        ws = wb["報價單"]

        # ── 表頭與客戶資訊 ──
        for key, cell in self.CELLS.items():
            if key in ("case_code", "quoted_date", "quotation_no"):
                continue
            self._set(ws, cell, data.get(key) or None)

        self._set(ws, self.CELLS["quotation_no"], data["display_no"])
        self._set(ws, self.CELLS["quoted_date"], data.get("quoted_date_roc") or None)
        # 工程編號用內部案號：客戶收到的單子上有它，回頭對帳才對得起來
        self._set(ws, self.CELLS["case_code"], data.get("case_code") or None)

        # ── 明細 ──
        items = data.get("items") or []
        if len(items) > (self.ITEM_LAST_ROW - self.ITEM_FIRST_ROW + 1):
            # 不靜靜截斷 —— 少印幾項的報價單會被當成完整報價送出去。
            cap = self.ITEM_LAST_ROW - self.ITEM_FIRST_ROW + 1
            raise ValueError(
                f"報價明細 {len(items)} 項超過範本可容納的 {cap} 項。"
                f"範本第 {self.NOTES_ROW} 列是備註列，明細只能用第 "
                f"{self.ITEM_FIRST_ROW}~{self.ITEM_LAST_ROW} 列；"
                "要更多項目請調整範本（把備註移到合計之後、擴充列數並同步 SUM 範圍），"
                "或把工項合併後把細節寫進備註。"
            )

        for i, it in enumerate(items):
            r = self.ITEM_FIRST_ROW + i
            ws[f"A{r}"] = f"{self._CN[i]}、" if i < len(self._CN) else f"{i + 1}、"
            ws[f"B{r}"] = it.get("item_name") or ""
            ws[f"C{r}"] = float(it.get("qty") or 0)
            ws[f"D{r}"] = it.get("unit") or ""
            ws[f"E{r}"] = float(it.get("unit_price") or 0)
            # F 欄寫**公式**不是數值 —— 客戶改數量時金額自己跟著動。
            #
            # ⚠️ 必須每列都寫：範本只在 16/17/18 有公式（樣本剛好 3 項），
            # 第 4 項以後若不補，複價欄會是空白而合計卻少算 ——
            # 那種錯不會報錯，只會讓總價比明細少。
            ws[f"F{r}"] = f"=E{r}*C{r}"
            ws[f"G{r}"] = it.get("notes") or ""

        # 清掉範本殘留的範例資料（範本本身是一份填好的真實報價單）
        #
        # ⚠️ **F 欄也要清**。第一版把 F 排除在外（想保留公式），
        # 實際開檔驗證才發現：範本樣本有 3 項，於是 F16~F18 的
        # `=E*C` 公式留在沒有明細的列上 → 那三格顯示 0，
        # 而 F16 那一列還同時印著「本案為委辦招標案」的說明文字。
        #
        # 這是「只看 bytes 不開檔」會漏掉的一類 —— 檔案產生了、
        # 大小正常、不報錯，打開才看得出版面不對。
        for r in range(self.ITEM_FIRST_ROW + len(items), self.ITEM_LAST_ROW + 1):
            for col in ("A", "B", "C", "D", "E", "F", "G"):
                ws[f"{col}{r}"] = None

        # ── 備註標籤（2026-08-29 備註由列 21 下移到列 29 之後才需要）──
        #
        # 範本的「備註：」標籤原本在 A21，而 A29 是空白格。
        # 只搬內容不搬標籤的話，客戶收到的單子上會有一段沒有抬頭的文字。
        # 字型沿用 A21（那是 owner 提供的版面，不自己挑）。
        #
        # ⚠️ 有備註才印標籤 —— 沒有備註卻印一個孤零零的「備註：」，
        # 比不印更像出錯。
        if data.get("notes"):
            label = ws[f"A{self.NOTES_ROW}"]
            label.value = "備註："
            src = ws[f"A{self.LEGACY_NOTES_ROW}"]
            if src.has_style:
                label.font = copy(src.font)
                label.alignment = copy(src.alignment)

        # 舊備註列現在屬於明細區，範本殘留的「備註：」/內容要清掉 ——
        # 否則明細不足 6 項時，該列會同時出現空明細與舊備註。
        # 條件＝它落在「明細沒填到的那一段」；有填到就已被明細覆寫，不能再清。
        if self.LEGACY_NOTES_ROW >= self.ITEM_FIRST_ROW + len(items):
            for col in ("A", "B"):
                cell = ws[f"{col}{self.LEGACY_NOTES_ROW}"]
                if isinstance(cell.value, str) and (
                    cell.value.startswith("備註") or col == "B"
                ):
                    cell.value = None

        # ── 沒有明細時（標案類）──
        #
        # 合計三格是公式（SUM/稅額/總計），沒有明細時它們會算出 0 ——
        # 而「總計 0 元」印在報價單上是錯的。改為寫入實際總價，
        # **覆蓋公式**：這一類本來就沒有逐項可加總。
        if not items:
            total = data.get("total_price")
            tax = data.get("tax_amount") or ZERO
            if total is not None:
                ws["E26"] = float(total - tax)
                ws["E27"] = float(tax)
                ws["E28"] = float(total)
            else:
                ws["E26"] = ws["E27"] = ws["E28"] = None
            if data.get("category") == "01":
                ws["B16"] = "本案為委辦招標案，依招標文件所列項目辦理"

        # 明細小計與報價總價不一致時**標在文件上**，不是挑一個顯示
        if data.get("amount_mismatch"):
            cur = ws["B21"].value or ""
            ws["B21"] = (
                f"{cur}\n※ 系統提醒：明細小計與報價總額不一致，請確認後再對外發出。"
            ).strip()

        # ── 列印縮放：讓 PDF 是「一頁寬」──
        #
        # 範本的 `fitToWidth`／`scale` 全是 None（沒設過縮放），而 A~G 欄總寬
        # 超過 A4 直式 —— 2026-08-19 實測轉出來是 **4 頁**：橫向被切成兩半，
        # 右半頁單獨成頁。一張報價單印成 4 頁送給客戶是不能用的。
        #
        # 設在這裡而不是改範本檔：範本是 owner 提供的原件，動它等於在二進位
        # 檔案裡藏一個看不見的變更；寫在程式裡，每次輸出都保證正確且看得懂。
        from openpyxl.worksheet.properties import PageSetupProperties

        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0  # 0＝高度不限，長明細自然往下延頁

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ── PDF ───────────────────────────────────────────────────────────
    #
    # owner 2026-08-18：「報價單要能輸出 pdf」。
    #
    # 轉換交給 LibreOffice，**不在程式碼裡重畫版面** —— 版面的唯一來源是
    # `templates/quotation_template.xlsx`。若用 reportlab 重畫，xlsx 與 PDF
    # 會各有一份版面，改抬頭要改兩處，而兩份不一致時沒有任何一方會報錯。

    #: soffice 轉檔上限。超過就當它卡住 —— 一個永遠不回來的轉檔，
    #: 症狀會是「按匯出之後什麼都沒發生」，比明確失敗更難查。
    PDF_TIMEOUT_SEC = 120

    @classmethod
    def render_pdf(cls, xlsx_bytes: bytes) -> bytes:
        """把 xlsx 轉成 PDF；失敗一律 raise，不退回 xlsx 假裝成功。"""
        import shutil
        import subprocess
        import tempfile
        import uuid

        soffice = shutil.which("soffice") or shutil.which("libreoffice")
        if soffice is None:
            # 明確講出缺什麼。靜默退回 xlsx 會讓使用者拿到副檔名是 .pdf
            # 但其實是 xlsx 的檔案 —— 那種檔案打不開，而且看不出原因。
            raise RuntimeError(
                "容器內找不到 LibreOffice（soffice）——"
                "PDF 轉換需要它，見 backend/Dockerfile 的 libreoffice-calc"
            )

        with tempfile.TemporaryDirectory() as tmp:
            src = Path(tmp) / "quotation.xlsx"
            src.write_bytes(xlsx_bytes)
            # ⚠️ `-env:UserInstallation` 是必要的，不是保險：
            # soffice 預設要在 $HOME 建設定檔，而容器內跑 job 的身分未必有
            # 可寫的 HOME。2026-08-15 就有一次 `pip-audit` 因為
            # `PermissionError: /nonexistent` 而**從來沒有跑起來過**，
            # 且三個 except 都只寫 logger.debug ⇒ 看起來像「沒有問題」。
            # 這裡指定一個一定可寫的 profile 目錄，從源頭避開同一件事。
            profile = f"file://{tmp}/lo_{uuid.uuid4().hex[:8]}"
            proc = subprocess.run(
                [soffice, "--headless", "--norestore", f"-env:UserInstallation={profile}",
                 "--convert-to", "pdf", "--outdir", tmp, str(src)],
                capture_output=True, timeout=cls.PDF_TIMEOUT_SEC,
            )
            out = Path(tmp) / "quotation.pdf"
            if not out.exists():
                # soffice 失敗時退出碼常常仍是 0 —— 所以判準是「檔案在不在」
                # 而不是回傳碼（又一次「成功訊號不等於做了事」）。
                raise RuntimeError(
                    "LibreOffice 轉檔沒有產生 PDF"
                    f"（rc={proc.returncode}）：{(proc.stderr or b'')[:300].decode('utf-8', 'replace')}"
                )
            content = out.read_bytes()

        # 產出必須真的是 PDF。只驗「檔案存在」等於沒驗。
        if not content.startswith(b"%PDF"):
            raise RuntimeError(f"轉檔產物不是 PDF（開頭 {content[:8]!r}）")
        return content

    # ── 自動存檔 ──────────────────────────────────────────────────────
    #
    # owner 2026-08-18：「並且自動納入系統存檔」，策略選「只保留最新一份」。
    #
    # 落點沿用既有的 `pm_case_attachments`（以 `case_code` 關聯，而報價單的
    # case_code 100% 有值）—— **不另造一套報價單專用附件表**。路徑與命名
    # 沿用 `api/endpoints/pm/attachments.py` 的慣例，讓既有的列表／下載／
    # 刪除三個端點直接就能用，不必為這批檔案再寫一份。

    #: 檔案落點（與 pm/attachments.py 同一個 env，不另立第二個設定）
    ARCHIVE_ROOT_ENV = "PM_ATTACHMENT_DIR"
    ARCHIVE_ROOT_DEFAULT = "uploads/pm_attachments"

    async def archive(
        self, data: dict[str, Any], content: bytes, ext: str, user_id: Optional[int]
    ) -> dict[str, Any]:
        """把輸出的報價單文件存進系統，**覆蓋同一張報價單的舊檔**。

        檔名固定為 `報價單_<報價單編號>.<ext>`（不帶時間戳）—— 覆蓋策略要靠
        它才找得到上一份；若檔名帶時間戳，每次輸出都會變成新的一筆，
        那是「保留版本」而不是 owner 選的「只保留最新」。
        """
        import hashlib
        import os

        from app.extended.models.pm import PMCaseAttachment

        case_code = data.get("case_code")
        if not case_code:
            # 沒有 case_code 就沒有掛載點。這裡不靜靜跳過 ——
            # 「文件下載成功但沒有存進系統」正是使用者最不會發現的那種失敗。
            raise ValueError("報價單缺少 case_code，無法存檔（附件以 case_code 關聯）")

        display_no = data.get("display_no") or f"Q{data.get('quotation_id')}"
        file_name = f"報價單_{display_no}.{ext}"
        root = os.environ.get(self.ARCHIVE_ROOT_ENV, self.ARCHIVE_ROOT_DEFAULT)
        dir_path = os.path.join(root, str(case_code), datetime.now().strftime("%Y%m"))
        os.makedirs(dir_path, exist_ok=True)
        # 正斜線寫入：2026-05-27 有一批 `file_path` 存成 Windows 反斜線，
        # 進 Linux 容器後 `os.path.exists` 一律 false（L49.3）。
        full_path = os.path.join(dir_path, file_name).replace("\\", "/")

        # 先清掉同一張報價單的舊紀錄與舊實體檔（owner 選「只保留最新一份」）
        old = (await self.db.execute(
            select(PMCaseAttachment).where(
                PMCaseAttachment.case_code == case_code,
                PMCaseAttachment.file_name == file_name,
            )
        )).scalars().all()
        for att in old:
            prev = (att.file_path or "").replace("\\", os.sep)
            if prev and os.path.exists(prev) and os.path.abspath(prev) != os.path.abspath(full_path):
                try:
                    os.remove(prev)
                except OSError:
                    # 舊檔刪不掉不該擋住新檔存檔，但要留下痕跡而不是靜靜吞掉
                    logger.warning("報價單存檔：舊檔刪除失敗 path=%s", prev)
            await self.db.delete(att)

        with open(full_path, "wb") as f:
            f.write(content)

        att = PMCaseAttachment(
            case_code=case_code,
            file_name=file_name,
            file_path=full_path,
            file_size=len(content),
            mime_type=(
                "application/pdf" if ext == "pdf"
                else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
            original_name=file_name,
            checksum=hashlib.sha256(content).hexdigest(),
            uploaded_by=user_id,
            # 2026-09-04：模型註解與前端標籤都說系統輸出會標 generated_quotation，
            # 但這裡從沒寫過 ⇒ 附件列表把它們當「未分類」，報價單分頁也篩不出來。
            doc_type="generated_quotation",
        )
        self.db.add(att)
        await self.db.commit()
        return {"file_name": file_name, "file_path": full_path, "replaced": len(old)}
