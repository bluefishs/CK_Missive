"""既有報價單彙整 XLS 匯入（個人管理時期的資料統整進系統）。

owner 2026-08-19：
  ①「新統整匯入與比對更新，對之前提及個人管理導致公司無法統整，
     故舊有資料仍有保留『案號如 B114-B002』，而成案編號 CK2026_01_01_007
     為系統統一碼機制」
  ②「新增與更新整合為一個按鍵鈕」
  ③「若線上產出報價單未完全上線前，如何匯入與管理既有 XLS 為目前階段重點」

# 資料現況（2026-08-19 實測，不是估計）

    114報價單彙整.xlsx   208 列（跨 5 個工作表：原始／老闆／慶忠／元宏／其他）
    115報價單彙整.xlsx    69 列
    合計                277 列，其中「是否成立=v」274 筆
    系統既有 erp_quotations                77 張

以案名比對：可對上 90 筆、近似 8 筆、**系統裡沒有 179 筆**。
所以匯入必須支援新增，不能只更新 —— 既有的 `pm/cases/import-xlsx`
是「拿第一欄當 PM 案件 id 去更新」，找不到就報錯，不適用於這裡。

# 為什麼用 legacy_quotation_no 當比對鍵

它是 XLS、紙本、回簽 PDF 檔名三者**共同**的識別（回簽檔名長這樣：
`回簽報價單_B115-C013-0_朱冠綸_….pdf`）。案名會被改寫、客戶名有簡稱，
只有這組編號穩定。

比對鍵一旦選錯，「更新」就會變成「重複新增」——所以刻意不用案名。

# 一個入口，不分新增/更新

依 `legacy_quotation_no` 決定 upsert：有就更新、沒有就新增。
使用者不需要先知道哪些是新的（他也無從知道 277 筆裡哪些已在系統）。

# 預覽先於寫入

`dry_run=True` 時只算不寫。**第一次匯入 277 筆業務資料**，
沒有預覽就按下去，錯了要靠備份還原 —— 而預覽的成本只是多一次點擊。
"""
from __future__ import annotations

import io
import logging
import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any, Optional

from sqlalchemy import select, text
from sqlalchemy.ext.asyncio import AsyncSession

from app.extended.models.erp import ERPQuotation
from app.services.contract.case_code import CaseCodeService

logger = logging.getLogger(__name__)

#: 表頭對應。XLS 的「報價金額 」尾端有空白 —— 表頭一律去空白後比對，
#: 不要求人去改檔案。
#:
#: ⚠️ 同一個概念在兩份檔案裡叫不同名字：115 是「發票日期」、114 是「發票日」。
#: 兩個都收 —— 要求人先統一表頭才能匯入，等於把工作推回去給填表的人。
HEADER_MAP = {
    "報價單編號": "legacy_no",
    "是否成立": "established",
    "報價日期": "quoted_date",
    "客戶名稱": "client_name",
    "案名": "case_name",
    "工作地點": "location",
    "報價金額": "total_price",
    "稅內含": "tax_included",
    "稅額": "tax_amount",
    "總價": "grand_total",
    # 以下是財務對帳的關鍵，第一版漏收了 ——
    # 匯入了卻少這些欄位，等於還是得回去翻 Excel。
    "實收金額": "received_amount",
    "收款日期": "received_date",
    "配合廠商": "partner_vendor",
    "支出金額": "expense_amount",
    "聯絡人": "contact_person",
    "電話": "contact_phone",
    "行動電話": "contact_mobile",
    "傳真": "contact_fax",
    "統一編號": "client_tax_id",
    "E-mail": "contact_email",
    "地址": "client_address",
    "備註": "remark",
    "發票日期": "invoice_date",
    "發票日": "invoice_date",      # 114 年度五張工作表用的是這個名字
    "印花": "stamp_duty",
    # 總表 v2（2026-09-03）：完整案名優先；發票明細比對後追加的四欄＋比對方式（「需確認」的不進系統）
    "完整案名(地點＋案名)": "full_case_name",
    "發票號碼": "invoice_no",
    "銷售額": "invoice_sales",
    "稅額(發票)": "invoice_tax",
    "發票金額": "invoice_amount",
    "比對方式": "match_method",
}

#: 這些欄位系統目前**沒有對應的結構化位置**，先原樣保進 notes。
#:
#: 為什麼不現在就開欄位：實收金額／收款日期屬應收、支出金額／配合廠商屬應付，
#: 它們各自該落在 `erp_billings`／`erp_vendor_payables` 而不是報價單上；
#: 而要建那些關聯，得先確定每一列對應到哪一張應收單 —— 那是下一階段的事。
#: **現在最重要的是資料不要遺失**，先原樣帶進來，結構化之後再從 notes 抽。
_NOTES_FIELDS = [
    ("工作地點", "location"), ("客戶", "client_name"), ("統編", "client_tax_id"),
    ("聯絡人", "contact_person"), ("電話", "contact_phone"), ("行動", "contact_mobile"),
    ("傳真", "contact_fax"), ("E-mail", "contact_email"), ("地址", "client_address"),
    ("總價", "grand_total"), ("實收金額", "received_amount"), ("收款日期", "received_date"),
    ("配合廠商", "partner_vendor"), ("支出金額", "expense_amount"),
    ("發票日", "invoice_date"), ("印花", "stamp_duty"), ("原始備註", "remark"),
]


def _clean_header(v: Any) -> str:
    return "".join(str(v or "").split())


def _to_decimal(v: Any) -> Optional[Decimal]:
    """金額：吃得下 45000 / '45,000' / '45000元' / None。"""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float, Decimal)):
        return Decimal(str(v))
    s = re.sub(r"[^\d.\-]", "", str(v))
    if not s or s in ("-", "."):
        return None
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def _roc_to_date(v: Any) -> Optional[date]:
    """報價日期是民國格式 `114.02.03`；也吃 Excel 原生日期。

    ⚠️ 不要用 `int(y) + 1911` 之前先判斷位數 —— `114` 是民國、
    `2025` 是西元，兩者都可能出現在同一欄（不同人填的）。
    """
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip().replace("/", ".").replace("-", ".")
    m = re.match(r"^(\d{2,4})\.(\d{1,2})\.(\d{1,2})$", s)
    if not m:
        return None
    y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
    if y < 1911:  # 民國
        y += 1911
    try:
        return date(y, mo, d)
    except ValueError:
        return None


def _year_from_legacy(legacy_no: str, quoted: Optional[date]) -> Optional[int]:
    """年度優先從舊案號取（`B115-C013-0` → 民國 115 → 2026）。

    用編號而不是報價日期：跨年度的案子（12 月報價、1 月成立）
    在編號上仍屬原年度，而那才是彙整表分年的依據。
    """
    m = re.match(r"^[A-Za-z]?(\d{3})[-_]", str(legacy_no or "").strip())
    if m:
        return int(m.group(1)) + 1911
    return quoted.year if quoted else None


#: 「是否成立」的肯定值。空白／其他一律視為未成立（不猜）。
_ESTABLISHED = {"v", "V", "y", "Y", "yes", "是", "✓", "○", "1", 1, True}


#: 工作表名稱的別名 —— 不是姓名、但 owner 指認過對應的人。
#:
#: owner 2026-08-20：「老闆 董事長 張坤樹」。
#: 只放**人講過**的對應，不自行推測（「其他」「原始」「工作表1」不在此列，
#: 它們本來就不是人）。這一份會隨檔案而變，所以放在這裡而不是寫進資料庫 ——
#: 下一份彙整表若換了工作表命名，改的是這一行。
#: 舊報價編號的**承辦代碼** → 人。
#:
#: ⚠️ owner 2026-08-29：「114、115代碼解析 **已多次提出** A坤樹 B慶忠 C元宏 D廷睿」
#: ＋「Y也指定慶忠」。這組對應被口頭提供過不只一次，而系統一直沒有記下來 ——
#: 寫在這裡就是為了**不用再有人講第三次**。
#:
#: 編號格式 `B115-A001-0A`：第一段 `B115` 是年度前綴（那個 B **不是人**），
#: 人的代碼是**第二段開頭那個字母**。
#:
#: 這比工作表名稱可靠：115 檔只有一個「工作表1」，但每一列的編號都帶著代碼。
#: 2026-08-29 依此回填 115 件案號（張坤樹 5／洪慶忠 61／邱元宏 49），
#: 報價單有承辦者由 135 張升到 250 張。
_LEGACY_CODE_TO_NAME = {
    "A": "張坤樹",
    "B": "洪慶忠",
    "C": "邱元宏",
    "D": "曾廷睿",
    "Y": "洪慶忠",
}

#: 從舊編號取承辦代碼
_LEGACY_STAFF_CODE_RE = re.compile(r"^[A-Z][0-9]{3}-([A-Z])")

_SHEET_ALIASES = {
    "老闆": "張坤樹",
}


def _derive_case_code(legacy_no: str) -> str:
    """case_code 就是**完整的舊案號** —— 不去尾碼。

    ## 為什麼不去 `-N`

    我第一版把 `-N` 當版次去掉（`B114-B004-1` → `B114-B004`），實測推翻：

      * `pm_cases` 裡本來就同時存在 `B114-B026`（平鎮區土地協議市價查估）
        與 `B114-B026-2`（翠64透地雷達作業）—— **兩個不同案件**；
      * 03-17 匯入的 70 件中 68 件不帶子號，而彙整表用的是帶子號的體系；
      * 去掉之後 4 組被硬掛在一起（案名完全不同、金額比值 0.49／0.54），
        另 36 組重複建立（金額正好差 5% ＝含稅／未稅，案名近似）。

    **子號是子案，不是版次。**

    ⚠️ 而 `signed_quotation_import.normalize_legacy_no`（回簽 PDF 掛檔）**刻意相反**，
    它忽略尾碼 —— 因為那一側「掛不上」的代價高於「掛錯」，而這一側正好相反。
    **同一個欄位兩種讀法，都是對的。不要統一。**
    已登記 `docs/architecture/TIER3_INTENTIONAL_DIVERGENCE_REGISTRY.md` §10。

    ## 那「既有那張與匯入這張是不是同一張」呢

    不由編號規則決定 —— 那要看案名與金額語意（36 組看起來是，4 組明顯不是），
    屬於人的判斷。系統讓每一筆有自己的身分，合併與否留給人。

    ## 為什麼不產新號

    2026-08-20 第一版用 `generate_case_code()`，結果 180 筆**全部拿到同一個
    `CK2025_FN_02_001`** —— 產號器查 DB 現有最大流水號，而同一交易內
    前面那些還沒 commit。那批已回滾。
    """
    return (legacy_no or "").strip()

def parse_workbook(content: bytes) -> list[dict[str, Any]]:
    """解析彙整檔；**全部工作表**都讀（114 年度分成 5 個表）。"""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    rows: list[dict[str, Any]] = []
    try:
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            it = ws.iter_rows(values_only=True)
            try:
                header = [_clean_header(c) for c in next(it)]
            except StopIteration:
                continue
            idx = {HEADER_MAP[h]: i for i, h in enumerate(header) if h in HEADER_MAP}
            if "legacy_no" not in idx:
                continue  # 這張表不是報價彙整（例如統計表）
            for raw in it:
                if not raw:
                    continue
                legacy = str(raw[idx["legacy_no"]] or "").strip()
                if not legacy:
                    continue
                def g(k):
                    return raw[idx[k]] if k in idx and idx[k] < len(raw) else None
                quoted = _roc_to_date(g("quoted_date"))
                rec = {
                    "sheet": sheet,
                    "legacy_no": legacy,
                    "established": g("established") in _ESTABLISHED,
                    "quoted_date": quoted,
                    "client_name": str(g("client_name") or "").strip() or None,
                    # 完整案名（地點＋案名）優先——「建物第一次測量」那種泛名對 70 個案，不具識別度
                    "case_name": (str(g("full_case_name") or "").strip() or str(g("case_name") or "").strip() or None),
                    "location": str(g("location") or "").strip() or None,
                    "total_price": _to_decimal(g("total_price")),
                    "tax_amount": _to_decimal(g("tax_amount")),
                    "year": _year_from_legacy(legacy, quoted),
                }
                # 其餘欄位原樣帶出（金額類轉 Decimal、日期類轉 date），
                # 目前沒有結構化位置的會進 notes —— 見 _NOTES_FIELDS。
                for key in ("grand_total", "received_amount", "expense_amount", "stamp_duty"):
                    rec[key] = _to_decimal(g(key))
                for key in ("received_date", "invoice_date"):
                    rec[key] = _roc_to_date(g(key))
                # v2 發票欄
                rec["invoice_no"] = (str(g("invoice_no") or "").strip() or None)
                rec["match_method"] = (str(g("match_method") or "").strip() or None)
                for key in ("invoice_sales", "invoice_tax", "invoice_amount"):
                    rec[key] = _to_decimal(g(key))
                for key in ("partner_vendor", "contact_person", "contact_phone",
                            "contact_mobile", "contact_fax", "client_tax_id",
                            "contact_email", "client_address", "remark", "tax_included"):
                    v = g(key)
                    rec[key] = str(v).strip() if v not in (None, "") else None
                rows.append(rec)
    finally:
        wb.close()
    return rows


def _build_notes(r: dict[str, Any]) -> str:
    """把還沒有結構化位置的欄位原樣保進備註。

    **不丟棄任何有值的欄位** —— 匯入的目的就是讓資料離開 Excel，
    落地時少一欄，使用者就還是得回去翻檔案。
    """
    parts = [f"由「{r.get('sheet')}」工作表匯入（舊案號 {r.get('legacy_no')}）"]
    for label, key in _NOTES_FIELDS:
        v = r.get(key)
        if v not in (None, ""):
            parts.append(f"{label}：{v}")
    return "；".join(parts)


class QuotationLegacyImportService:
    """一個入口做 upsert：有就更新、沒有就新增。"""

    def __init__(self, db: AsyncSession):
        self.db = db
        self.code_service = CaseCodeService(db)


    # ------------------------------------------------------------------
    # 邀標案件（pm_cases）補建
    # ------------------------------------------------------------------
    #
    # owner 2026-08-20：「/pm/cases 無 115 年度報價單紀錄」「2026 僅 4 筆紀錄」。
    #
    # 查證：`pm_cases` 的 70 件 `B114-B001~B070` 是 **2026-03-17 一次性匯入**
    # 建立的 —— 也就是說「一個報價案號 = 一個邀標案件」本來就是這個系統的做法。
    # 而 115 年度有 68 個案號只有報價單、沒有案件 ⇒ **那是缺的，不是刻意不建**。
    #
    # ⚠️ 我一度套用 2026-08-19 的「不擅自補建 pm_cases」判準而沒有建 ——
    #    那是**誤用**：那條講的是「手動建承攬案件時，憑空造一筆從未經過邀標的
    #    PM 案件」；而這裡的報價單**本身就是邀標紀錄**，補的是已經發生的事實。
    #
    # 欄位比照 03-17 那批（案名／年度／類別／客戶／金額／狀態／地點／作業日期），
    # 資料全部來自彙整表，不發明任何值。
    async def _ensure_pm_cases(
        self, rows: list[dict[str, Any]], dry_run: bool = False,
        code_of: Optional[dict[str, Optional[str]]] = None,
    ) -> int:
        """為新報價單補建對應的邀標案件；已存在則不動。回傳補建數。

        2026-08-28 step 6（A32 之後才能做的那一步）：案件身分不再由
        `_derive_case_code(legacy_no)` 決定 —— A32 已把存量 pm_cases 全轉為
        CK 建案案號，再用舊編號去比對必然「找不到 ⇒ 重複建案」。
        改由呼叫端傳入 `code_of`（legacy_no → case_code）：
          · 既有報價單（以 legacy_quotation_no 對到）→ 用它現在的 case_code
          · 真正的新業務 → 寫入時由產號器發 CK 碼；dry-run 時值為 None
            （代表「必建」，數字不預先消耗流水號）
        未傳 code_of 時退回舊行為（僅供 legacy 情境）。
        """
        from app.extended.models.pm import PMCase

        wanted: dict[str, dict[str, Any]] = {}
        pending_new: list[dict[str, Any]] = []
        for r in rows:
            ln = r["legacy_no"]
            if code_of is not None and ln in code_of:
                code = code_of[ln]
                if code is None:
                    # dry-run 的新業務：案號寫入時才產，必然是待補建
                    pending_new.append(r)
                    continue
            else:
                code = _derive_case_code(ln)
            if not code:
                continue
            # 2026-09-03：只對 PM 制案號補建。既有報價單若掛在舊制／GN 制案號上（01 標案、直接建的承攬案），
            # 它本來就沒有邀標階段，補一個 PM 案是憑空造（08-18 判過）——匯出→匯入往返實測造了 21 個殼。
            if "_PM_" not in code:
                continue
            # 同一案號多版報價 → 只建一件，取**最早報價**那筆當案件基本資料
            cur = wanted.get(code)
            if cur is None or (r.get("quoted_date") and cur.get("quoted_date")
                               and r["quoted_date"] < cur["quoted_date"]):
                wanted[code] = r
        if not wanted and not pending_new:
            return 0

        existing = set()
        if wanted:
            existing = {
                row[0] for row in (await self.db.execute(
                    select(PMCase.case_code).where(PMCase.case_code.in_(list(wanted)))
                )).all()
            }
        missing = {k: v for k, v in wanted.items() if k not in existing}

        # 2026-08-27：這裡只比對 **完整 case_code**，而 08-20 那次匯入正是這樣
        # 建出 26 件分身：既有 `B114-B003`，彙整表帶的是 `B114-B003-0`，
        # 兩者 case_code 不同 ⇒ 判為「不存在」⇒ 再建一件，**而且不出聲**。
        # 事後查證：那 26 組兩側**案名完全相同**，且有碼那一側全部都有金流
        # （錢記在原始那筆上），⇒ 新建的那 26 件是分身，不是業務缺口。
        #
        # ⚠️ **刻意不自動合併**：本檔 `_derive_case_code` 的判斷是
        # 「子號是子案，不是版次」，而那是對的 —— 實測有 4 組同 base 同版次形態
        # 但案名完全不同（`B114-B026` 平鎮區查估 vs `B114-B026-0` 永翠76透地雷達）。
        # 合併與否要看案名與金額語意，屬人的判斷。
        #
        # 但**沉默地再建一次**不是「留給人判斷」，是不給人判斷的機會。
        # 所以這裡只做一件事：把「同 base 且**案名完全相同**、只差版次」的挑出來說。
        self.dup_candidates: list[dict[str, str]] = []
        if missing:
            bases = {re.sub(r"-[0-9]+$", "", k) for k in missing}
            rs = (await self.db.execute(text(
                "SELECT case_code, case_name FROM pm_cases "
                "WHERE regexp_replace(case_code, '-[0-9]+$', '') = ANY(:bs)"
            ), {"bs": list(bases)})).all()
            by_base: dict[str, list[tuple[str, str]]] = {}
            for cc, cn in rs:
                by_base.setdefault(re.sub(r"-[0-9]+$", "", cc), []).append((cc, cn))
            for code, r in missing.items():
                base = re.sub(r"-[0-9]+$", "", code)
                for cc, cn in by_base.get(base, []):
                    if cn and cn == r.get("case_name"):
                        self.dup_candidates.append(
                            {"new_case_code": code, "existing_case_code": cc,
                             "case_name": cn, "matched_by": "版次後綴"})
                        break

            # ⚠️ 2026-08-31：上面那段**只認 `-N` 版次後綴**（`B114-B003` vs
            # `B114-B003-0`），對 CK 制的案號完全無效 —— `CK2025_PM_02_010` 與
            # `CK2025_PM_02_112` 去掉 `-N` 之後 base 仍然不同 ⇒ 判為不存在 ⇒
            # **靜靜再建一件**，正是 08-20 那 26 件分身的同一個形狀換了編碼體系。
            #
            # 當天實測存量：39 個 PM 案與既有承攬案「同名＋同年＋同委託單位」，
            # 其中 16 組金額恰為 ×1.05（未稅／含稅的同一件工作）、5 組金額完全相同。
            # 那批全部卡在「已承攬但無法成案」——因為成案的防重擋住了它們。
            #
            # ⇒ 這裡改用**與 `promote_to_project` 相同的判準**（同名＋同年＋
            # 同委託單位）再掃一次。兩條路徑用同一條規則，否則擋住一邊等於沒擋。
            #
            # 一樣**只報不合併**：同名同年同客戶可能真的是兩件工作
            # （同一客戶同年做兩次透地雷達完全合理），合併屬人的判斷。
            # 但沉默地再建一次，不是「留給人判斷」，是不給人判斷的機會。
            already = {d["new_case_code"] for d in self.dup_candidates}
            todo = {k: v for k, v in missing.items() if k not in already}
            if todo:
                rs2 = (await self.db.execute(text(
                    "SELECT case_code, case_name, year, COALESCE(client_name,'') "
                    "FROM pm_cases WHERE case_name = ANY(:ns)"
                ), {"ns": [v.get("case_name") for v in todo.values() if v.get("case_name")]})).all()
                by_key: dict[tuple, str] = {}
                for cc, cn, yr, cl in rs2:
                    by_key.setdefault(((cn or "").strip(), yr, (cl or "").strip()), cc)
                for code, r in todo.items():
                    nm = (r.get("case_name") or "").strip()
                    if not nm:
                        continue
                    cl = (r.get("client_name") or "").strip()
                    hit = by_key.get((nm, r.get("year"), cl))
                    if hit and hit != code:
                        self.dup_candidates.append(
                            {"new_case_code": code, "existing_case_code": hit,
                             "case_name": nm, "matched_by": "同名＋同年＋同委託單位"})

        if dry_run:
            # 預告會成案幾件（寫入時走正式 promote），讓 preview 看得到不可逆動作的規模
            self.will_promote = sum(1 for v in missing.values() if v.get("established")) + \
                sum(1 for r in pending_new if r.get("established"))
            return len(missing) + len(pending_new)
        # 寫入模式下 code_of 必須完整（新業務的號已先產好）—— pending_new
        # 非空代表呼叫端漏了產號，出聲而不是靜靜少建
        if pending_new:
            raise ValueError(
                f"{len(pending_new)} 筆新業務沒有建案案號 —— 寫入前必須先產號")
        if not missing:
            return 0

        # 客戶名稱 → 既有廠商 id。對不到就只留文字（不自動建廠商 ——
        # 那會把彙整表裡的簡稱、筆誤全部變成新廠商，而 owner 正在處理
        # 廠商重複的問題「勤典工程行／勤典測量工程行」）。
        names = {v["client_name"] for v in missing.values() if v.get("client_name")}
        vendor_id: dict[str, int] = {}
        if names:
            rs = (await self.db.execute(text(
                "SELECT id, vendor_name FROM partner_vendors WHERE vendor_name = ANY(:ns)"
            ), {"ns": list(names)})).all()
            vendor_id = {row[1]: row[0] for row in rs}

        for code, r in missing.items():
            self.db.add(PMCase(
                case_code=code,
                case_name=r["case_name"],
                year=r["year"] or date.today().year,
                # 邀標報價 = 02。01 是委辦招標（政府標案），那類不走報價單流程。
                category="02",
                client_name=r.get("client_name"),
                client_vendor_id=vendor_id.get(r.get("client_name") or ""),
                contract_amount=r.get("total_price"),
                # 「是否成立=v」＝客戶接受了這張報價 ⇒ 已承攬；否則仍在評估。
                # 不寫 in_progress —— 邀標案件沒有那個狀態（2026-08-16 owner 已定）。
                status="contracted" if r.get("established") else "planning",
                start_date=r.get("quoted_date"),
                location=r.get("location"),
                notes=f"由報價單彙整匯入（舊案號 {r['legacy_no']}）",
            ))

        # 2026-09-04 金流複查：此前「已成立」的列只寫 status=contracted、**不建承攬案** ——
        # 09-03 那次匯入留下 16 筆 PM 案標已承攬而承攬案列表看不到、報價單沒有 project_code、
        # 損益摘要把它們當未成案、掛在上面的請款在成案口徑裡消失。
        # 成立＝已承攬＝要有承攬案（owner 09-02：XLS 為真值；成案即應收）。走正式 promote_to_project：
        # 它會擋同名承攬案（重複建案）與缺金額，擋住的**列出來**（promote_failures），不吞、也不降回 planning
        # （總表說成立，那就是成立；是不是重複要人判）。⚠️ promote 內部會 commit（L139），本函式只在寫入模式到這裡。
        await self.db.flush()
        self.promote_failures: list[dict[str, str]] = []
        self.promoted_count = 0
        for code, r in missing.items():
            if not r.get("established"):
                continue
            try:
                await self.code_service.promote_to_project(code)
                self.promoted_count += 1
            except ValueError as e:
                self.promote_failures.append({"case_code": code, "case_name": r.get("case_name") or "",
                                              "reason": str(e)[:200]})
                logger.warning("匯入成案被擋 case_code=%s：%s", code, str(e)[:200])
        return len(missing)


    # ------------------------------------------------------------------
    # 承辦同仁：從來源工作表名稱對應
    # ------------------------------------------------------------------
    #
    # owner 2026-08-20：「報價單要對應使用者以利自我案件維護管理」
    # （並指向 `/pm/cases/305?tab=staff` —— 也就是既有的承辦同仁機制）。
    #
    # 114 彙整表分成 5 個工作表，而其中三個**是人名**：原始／老闆／慶忠／元宏／其他。
    # 那不是分類，是「這個案子誰在跑」—— 資訊一直都在檔案裡，
    # 我第一版匯入只把它寫進備註文字，等於丟掉了。
    #
    # ⚠️ 必須用**原始列**（未去重）：去重保留的是先遇到的那份，而工作表順序
    # 「原始」排第一 ⇒ 去重後 108 張都掛在「原始」名下，人名全部消失。
    #
    # 比對方式：工作表名是否**包含**在使用者姓名裡（「慶忠」⊂「洪慶忠」）。
    # 對不到就**不指派**並列出來 —— 「老闆」「其他」「原始」「工作表1」都對不到，
    # 而「老闆是誰」是我猜不得的事（owner 提過張坤樹是董事長，但那是他說的，
    # 不是這份檔案說的）。
    async def _assign_staff_from_sheets(
        self, raw_rows: list[dict[str, Any]], dry_run: bool = False,
    ) -> dict[str, Any]:
        """依來源工作表名稱指派承辦同仁。冪等：已存在的指派不重複建立。"""
        from collections import defaultdict

        sheets_by_case: dict[str, set[str]] = defaultdict(set)
        codes_by_case: dict[str, set[str]] = defaultdict(set)
        # 2026-09-03：案號先問資料庫（legacy／QT 號 → 報價單現在的 case_code），推導只當後備；
        # 推不出 CK 制的一律不建 —— 此前 `_derive_case_code` 推不出來就把舊案號當 case_code 寫進指派表，
        # 往返匯入一次造了 103 筆 `B115-C001a-0` 這種鍵（09-02 附件掛錯鍵同族）。
        _lns = [r.get("legacy_no") for r in raw_rows if r.get("legacy_no")]
        _map: dict[str, str] = {}
        if _lns:
            for _ln, _cc in (await self.db.execute(text(
                "SELECT COALESCE(legacy_quotation_no, quotation_no), case_code FROM erp_quotations "
                "WHERE deleted_at IS NULL AND (legacy_quotation_no = ANY(:l) OR quotation_no = ANY(:l))"
            ), {"l": _lns})).all():
                _map[_ln] = _cc
        for r in raw_rows:
            ln = r.get("legacy_no") or ""
            if not any(ch.isdigit() for ch in ln):
                continue
            _cc = _map.get(ln) or _derive_case_code(ln)
            if not _cc or not str(_cc).startswith("CK"):
                continue
            sheets_by_case[_cc].add(r.get("sheet") or "")
            codes_by_case[_cc].add(ln)

        users = (await self.db.execute(text(
            "SELECT id, COALESCE(full_name, username) AS nm FROM users"
            " WHERE is_active AND canonical_user_id IS NULL"
        ))).all()

        def match(sheet: str):
            # owner 指認過的別名優先（「老闆」→ 張坤樹），其次才是姓名包含比對
            target = _SHEET_ALIASES.get(sheet, sheet)
            for uid, nm in users:
                if target and nm and target in nm:
                    return uid, nm
            return None, None

        wanted: list[tuple[str, int]] = []
        unmatched: dict[str, int] = {}
        # 代碼優先於工作表名稱：115 檔只有一個「工作表1」（對不到任何人），
        # 但它每一列的編號都帶著承辦代碼。先前只讀工作表名，於是整批 115 沒有承辦。
        def match_code(case_code: str):
            for ln in codes_by_case.get(case_code, ()):
                m = _LEGACY_STAFF_CODE_RE.match(ln)
                if not m:
                    continue
                nm_want = _LEGACY_CODE_TO_NAME.get(m.group(1))
                if not nm_want:
                    continue
                for uid, nm in users:
                    if nm and nm_want in nm:
                        return uid
            return None

        for case_code, sheet_set in sheets_by_case.items():
            uid = match_code(case_code)
            if uid:
                wanted.append((case_code, uid))
                continue
            for sh in sheet_set:
                uid, _ = match(sh)
                if uid:
                    wanted.append((case_code, uid))
                else:
                    unmatched[sh] = unmatched.get(sh, 0) + 1
        if not wanted:
            return {"assigned": 0, "unmatched_sheets": unmatched}

        existing = {
            (row[0], row[1]) for row in (await self.db.execute(text(
                # 2026-09-03：指派有兩條綁法（case_code／project_id→承攬案），查重要兩條都認——
                # 只認 case_code 讓成案後改掛 project_id 的那筆被判「不存在」而重建（同族第十一處）
                "SELECT a.case_code, a.user_id FROM project_user_assignments a WHERE a.case_code = ANY(:cs)"
                " UNION SELECT c.case_code, a.user_id FROM project_user_assignments a"
                " JOIN contract_projects c ON c.id = a.project_id WHERE c.case_code = ANY(:cs)"
            ), {"cs": list({c for c, _ in wanted})})).all()
        }
        todo = [(c, u) for c, u in dict.fromkeys(wanted) if (c, u) not in existing]
        if dry_run or not todo:
            return {"assigned": len(todo), "unmatched_sheets": unmatched}

        name_of = {uid: nm for uid, nm in users}
        for case_code, uid in todo:
            await self.db.execute(text(
                "INSERT INTO project_user_assignments"
                " (case_code, user_id, staff_name, role, is_primary, status, notes,"
                "  assignment_date, created_at, updated_at)"
                " VALUES (:cc, :uid, :nm, :role, false, 'active', :notes,"
                "         CURRENT_DATE, NOW(), NOW())"
            ), {
                "cc": case_code, "uid": uid, "nm": name_of.get(uid),
                # 既有選項（計畫主持／計畫協同／專案PM／職安主管）取「專案PM」——
                # 彙整表只說了「這是誰的案子」，沒說職責層級，不多猜一層。
                "role": "專案PM",
                "notes": "由報價單彙整匯入（來源工作表即承辦人）",
            })
        return {"assigned": len(todo), "unmatched_sheets": unmatched}

    async def _sync_finance(self, rows: list[dict[str, Any]]) -> dict[str, int]:
        """匯入列 → 請款／發票（2026-09-03）。

        規則與 09-02／09-03 兩次總表匯入相同：
        - 成立且有總額 ⇒ `ensure_first_period`（一次請領、pending；已有請款就不動）
        - 有收款日期 ⇒ 第一筆請款標 paid（payment_date／payment_amount＝實收或總價）
        - 有發票號碼且比對方式非「需確認」⇒ 有佔位（XLS-）就補真號，沒有發票就建一張綁第一筆請款
        全部只補空、不覆蓋人填的值。
        """
        from sqlalchemy import text as _t
        from app.services.erp.billing_service import ERPBillingService
        out = {"first_period": 0, "paid": 0, "invoice_created": 0, "invoice_updated": 0, "skipped_unconfirmed": 0}
        bsvc = ERPBillingService(self.db)
        inv_pat = re.compile(r"^[A-Z]{2}[0-9]{8}$")
        for r in rows:
            q = await self.db.scalar(_t("SELECT id FROM erp_quotations WHERE (legacy_quotation_no=:l OR quotation_no=:l) AND deleted_at IS NULL LIMIT 1"), {"l": r["legacy_no"]})
            if not q:
                continue
            if r.get("established") and (r.get("total_price") or 0) > 0:
                if await bsvc.ensure_first_period(int(q), reason="總表匯入"):
                    out["first_period"] += 1
            if r.get("received_date"):
                res = await self.db.execute(_t(
                    "UPDATE erp_billings SET payment_status='paid', payment_date=:d, payment_amount=COALESCE(payment_amount, :a), updated_at=now() "
                    "WHERE id=(SELECT id FROM erp_billings WHERE erp_quotation_id=:q ORDER BY billing_date LIMIT 1) AND payment_status<>'paid'"
                ), {"d": r["received_date"], "a": r.get("received_amount") or r.get("total_price"), "q": int(q)})
                out["paid"] += res.rowcount or 0
            inv_no = r.get("invoice_no")
            if inv_no and "需確認" in (r.get("match_method") or ""):
                out["skipped_unconfirmed"] += 1
                inv_no = None
            if inv_no and inv_pat.match(inv_no):
                dup = await self.db.scalar(_t("SELECT 1 FROM erp_invoices WHERE invoice_number=:n AND erp_quotation_id<>:q LIMIT 1"), {"n": inv_no, "q": int(q)})
                if dup:
                    continue
                res = await self.db.execute(_t(
                    "UPDATE erp_invoices SET invoice_number=:n, invoice_date=COALESCE(CAST(:d AS date), invoice_date), amount=COALESCE(CAST(:amt AS numeric), amount), tax_amount=COALESCE(CAST(:tax AS numeric), tax_amount), updated_at=now() "
                    "WHERE erp_quotation_id=:q AND invoice_number LIKE 'XLS-%'"
                ), {"n": inv_no, "d": r.get("invoice_date"), "amt": r.get("invoice_amount"), "tax": r.get("invoice_tax"), "q": int(q)})
                if res.rowcount:
                    out["invoice_updated"] += res.rowcount
                else:
                    has = await self.db.scalar(_t("SELECT 1 FROM erp_invoices WHERE erp_quotation_id=:q LIMIT 1"), {"q": int(q)})
                    if not has and r.get("invoice_date"):
                        await self.db.execute(_t(
                            "INSERT INTO erp_invoices (erp_quotation_id, invoice_number, invoice_date, amount, tax_amount, invoice_type, status, billing_id, notes, source, created_at, updated_at) "
                            "VALUES (:q, :n, :d, :amt, :tax, 'sales', 'issued', (SELECT id FROM erp_billings WHERE erp_quotation_id=:q ORDER BY billing_date LIMIT 1), '由報價單彙整匯入（發票明細）', 'xls_import', now(), now())"
                        ), {"q": int(q), "n": inv_no, "d": r["invoice_date"], "amt": r.get("invoice_amount") or r.get("total_price"), "tax": r.get("invoice_tax") or 0})
                        out["invoice_created"] += 1
        await self.db.commit()
        return out

    @staticmethod
    def _legacy_base(ln: Optional[str]) -> Optional[str]:
        """取估價單編號的 base（去掉版次後綴）。

            B114-B048-1  →  B114-B048
            B114-B048    →  B114-B048
            B114-C033a-0 →  B114-C033a   ← **子碼 a/b/c 保留**

        ⚠️ 子碼刻意留在 base 裡：`C033a` 與 `C033b` 是**同一份估價單的分項子單**
        （實測 `C014a` 68,000 ＋ `C014b` 188,090，各涵蓋不同部分），
        把它們當成同一張會把兩筆真實金額併成一筆。
        版次（尾端 `-N`）才是同一張單的不同版本。
        """
        m = re.match(r"^([A-Z]\d+-[A-Z]\d+[a-z]*)(?:-(.*))?$", (ln or "").strip())
        return m.group(1) if m else None

    async def _detect_revision_dups(self, rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
        """檔案裡的編號，其 base 是否已經以**別的版次**存在於資料庫。

        回傳候選清單（只報不動）。取不到 base 的編號（格式異常）直接略過 ——
        猜不出來就不猜，那類編號會出現在 `unparsable_legacy_no` 讓人看。
        """
        bases: dict[str, str] = {}
        unparsable: list[str] = []
        for r in rows:
            ln = r.get("legacy_no")
            b = self._legacy_base(ln)
            if b is None:
                if ln:
                    unparsable.append(ln)
            else:
                bases.setdefault(b, ln)
        self.unparsable_legacy_no = sorted(set(unparsable))
        if not bases:
            return []

        # 撈出所有 base 相符的既有報價單（用 LIKE 前綴，再於 Python 精確比對 base）
        found = (await self.db.execute(text(
            "SELECT legacy_quotation_no, case_code, total_price FROM erp_quotations "
            "WHERE deleted_at IS NULL AND legacy_quotation_no IS NOT NULL"
        ))).all()
        by_base: dict[str, list[tuple]] = {}
        for ln, cc, amt in found:
            b = self._legacy_base(ln)
            if b:
                by_base.setdefault(b, []).append((ln, cc, amt))

        out: list[dict[str, Any]] = []
        for b, incoming_ln in bases.items():
            for ln, cc, amt in by_base.get(b, []):
                if ln != incoming_ln:  # 完整編號相同的走既有的 to_update，不是這裡的事
                    out.append({
                        "incoming": incoming_ln, "existing": ln,
                        "base": b, "existing_case_code": cc,
                        "existing_amount": float(amt or 0),
                    })
        return out

    async def run(self, content: bytes, *, dry_run: bool = True,
                  user_id: Optional[int] = None, source_name: Optional[str] = None) -> dict[str, Any]:
        rows = parse_workbook(content)
        if not rows:
            return {"success": False, "error": "檔案裡找不到『報價單編號』欄，請確認是報價單彙整表"}

        legacy_nos = [r["legacy_no"] for r in rows]
        existing = {
            q.legacy_quotation_no: q
            for q in (await self.db.execute(
                select(ERPQuotation).where(
                    ERPQuotation.legacy_quotation_no.in_(legacy_nos),
                    ERPQuotation.deleted_at.is_(None),
                )
            )).scalars().all()
        }
        # 2026-09-03：匯出檔的「報價單編號」欄對沒有舊案號的單寫的是 QT 號（線上報價單機制）。
        # 這些列以 quotation_no 比對既有——否則匯出→匯入會把每張 QT 單再建一次。
        qt_nos = [ln for ln in legacy_nos if re.match(r"^QT\d{4}_\d{3}$", ln or "")]
        if qt_nos:
            for q in (await self.db.execute(
                select(ERPQuotation).where(ERPQuotation.quotation_no.in_(qt_nos), ERPQuotation.deleted_at.is_(None))
            )).scalars().all():
                existing.setdefault(q.quotation_no, q)

        # ⚠️ 2026-08-31：`existing` 用**完整編號**比對，而版次後綴會讓同一份
        # 估價單被判成兩張不同的單。實測後果（owner 回報「報價單匯入搞到整個
        # 系統都錯亂」後查證）：
        #
        #   03-17 匯入 `B114-B048`   ── 48 張
        #   08-20 匯入 `B114-B048-1` ── 179 張，其中 49 張與既有的同一份估價單
        #   ⇒ 03-17 那批 **48 張裡有 42 張被重新建了一次**
        #   ⇒ 全庫 44 組 / 91 張是同一份估價單的多筆紀錄，**重複 47 張、
        #      NT$6,144,188（占報價單總額 6.0%）**
        #   ⇒ 連帶 08-21 自動補建 **179 筆邀標案件**（占全部 253 筆的 71%），
        #      其中 72 筆同名同年同客戶互為分身
        #
        # 這裡**只報不合併**：版次要保留幾版是業務政策（最新版 vs 全部保留），
        # 而且 `C033a`／`C033b` 那種子碼是**分項子單、不是版次**，合併會出錯。
        # 但沉默地再建一次不是「留給人判斷」，是不給人判斷的機會。
        self.revision_candidates: list[dict[str, Any]] = await self._detect_revision_dups(rows)

        to_create, to_update, skipped = [], [], []
        seen: dict[str, dict] = {}
        conflicts: list[dict] = []
        for r in rows:
            ln = r["legacy_no"]
            if ln in seen:
                # 同一份檔案裡重複的編號。
                #
                # 114 年度的彙整表分成 5 個工作表（原始／老闆／慶忠／元宏／其他），
                # 同一張報價單會在「原始」總表與個人分表各出現一次 —— 實測 97 組重複，
                # 其中 84 組完全相同（純副本），**13 組內容不同**（多半是發票日期
                # 只填在其中一邊）。
                #
                # ⚠️ 原本的做法是「只留第一筆、其餘丟掉」，於是那 13 組裡
                # 另一邊獨有的欄位會**靜靜消失** —— 實測有 1 組保留到的是較不完整的那份。
                # 而「先遇到的比較完整」只是工作表順序剛好（原始排第一），不是設計。
                #
                # 改為**補空缺**：只把保留者為空、而重複列有值的欄位補上，
                # 不覆蓋任何已有值（兩邊都有值且不同時，仍以先遇到的為準並列入
                # conflicts 讓人看）。這樣重複不再等於丟資料。
                base = seen[ln]
                filled_from_dup, conflict_keys = [], []
                for k, v in r.items():
                    if k in ("sheet", "legacy_no"):
                        continue
                    if v in (None, "", False):
                        continue
                    cur = base.get(k)
                    if cur in (None, "", False):
                        base[k] = v
                        filled_from_dup.append(k)
                    elif str(cur) != str(v):
                        conflict_keys.append(k)
                if conflict_keys:
                    conflicts.append({
                        "legacy_no": ln,
                        "kept_sheet": base.get("sheet"),
                        "other_sheet": r.get("sheet"),
                        "conflict_fields": conflict_keys,
                    })
                skipped.append({
                    "legacy_no": ln,
                    "reason": "檔案內重複（已補空缺欄位）" if filled_from_dup else "檔案內重複（內容相同）",
                    "sheet": r.get("sheet"),
                    "filled_from_dup": filled_from_dup or None,
                    "conflict_fields": conflict_keys or None,
                })
                continue
            if not r["case_name"]:
                skipped.append({"legacy_no": ln, "reason": "缺案名", "sheet": r.get("sheet")})
                continue
            # 非案號的儲存格 —— 彙整表裡夾雜說明文字（實測有一列的「報價單編號」
            # 欄寫著「訂購通知」），照收就會產生一筆案號是中文的報價單。
            # 判準只要求「含數字」：所有真實案號都有年碼或日期
            # （`B114-B022-1`、`20260304-1`），而說明文字沒有。
            if not any(ch.isdigit() for ch in ln):
                skipped.append({"legacy_no": ln, "reason": "不是案號（無數字，疑為說明文字）",
                                "sheet": r.get("sheet")})
                continue
            seen[ln] = r
            (to_update if ln in existing else to_create).append(r)

        preview = {
            "success": True,
            "dry_run": dry_run,
            "total_rows": len(rows),
            "will_create": len(to_create),
            "will_update": len(to_update),
            "skipped": len(skipped),
            # ⚠️ 原本固定切 20 筆，而實測跳過 97 筆 —— 看的人無從核實其餘 77 筆
            #    是不是該跳過，那等於「匯入了卻不知道丟了什麼」。
            #    改為全給，並明講有沒有被截斷（上限只是防回應過大）。
            "skipped_detail": skipped[:200],
            "skipped_detail_truncated": len(skipped) > 200,
            # 兩邊都有值且不同的欄位 —— 這些是**人必須看的**，
            # 合併只補空缺、不會覆蓋，所以衝突欄位保留的是先遇到的那份。
            "conflicts": conflicts,
            "conflicts_count": len(conflicts),
            # ⚠️ 2026-08-31 新增：**同一份估價單、不同版次**。
            # 這是 08-20 那次匯入把 03-17 的 48 張裡 42 張重建一次的形狀 ——
            # 比對用完整編號，`B114-B048` 與 `B114-B048-1` 於是各成一張。
            # 全庫殘留：44 組 / 91 張，重複 47 張 NT$6,144,188（總額的 6.0%）。
            # **這個數字不為 0 時，`will_create` 就不能照單全收。**
            "revision_dups": self.revision_candidates,
            "revision_dups_count": len(self.revision_candidates),
            # 編號格式解析不出來的（例如 `B110-012-v6`、`20260512-01`）——
            # 這些不在上面的偵測範圍內，要人看一眼。
            "unparsable_legacy_no": getattr(self, "unparsable_legacy_no", []),
            "sample_create": [
                {"legacy_no": r["legacy_no"], "case_name": r["case_name"],
                 "client_name": r["client_name"], "total_price": str(r["total_price"] or ""),
                 "year": r["year"], "established": r["established"]}
                for r in to_create[:10]
            ],
        }
        # 預覽也要算「會補建幾件邀標案件」—— 這是使用者在 /pm/cases 看得到的東西，
        # 只講「新增幾張報價單」而不講案件，等於預覽沒有涵蓋一半的後果。
        # 涵蓋 to_create **與** to_update —— 這支是冪等的（已存在就不建），
        # 而只看新增的話，「報價單已匯入但案件還沒補」的情況永遠補不上。
        # step 6（2026-08-28）：案件身分的對照表 ——
        #   既有報價單 → 它現在的 case_code（A32 後是 CK 建案案號）
        #   新業務 → None（寫入時才產號，dry-run 不消耗流水號）
        code_of: dict[str, Optional[str]] = {
            r["legacy_no"]: existing[r["legacy_no"]].case_code for r in to_update
        }
        for r in to_create:
            code_of[r["legacy_no"]] = None

        preview["will_create_pm_cases"] = await self._ensure_pm_cases(
            to_create + to_update, dry_run=True, code_of=code_of)
        # 「同 base 且案名完全相同、只差版次」的候選 —— 08-20 那次靜靜建了 26 件分身，
        # 這裡把它攤在匯入前的預覽上。**只提醒不阻擋**：合併與否是人的判斷。
        preview["duplicate_candidates"] = getattr(self, "dup_candidates", [])[:200]
        preview["duplicate_candidate_count"] = len(getattr(self, "dup_candidates", []))
        preview["will_promote"] = getattr(self, "will_promote", 0)  # 寫入時會走正式 promote（不可逆）的件數
        _staff = await self._assign_staff_from_sheets(rows, dry_run=True)
        preview["will_assign_staff"] = _staff["assigned"]
        preview["staff_unmatched_sheets"] = _staff["unmatched_sheets"]
        if dry_run:
            return preview

        created = updated = 0
        for r in to_update:
            q = existing[r["legacy_no"]]
            for field, val in (
                ("case_name", r["case_name"]), ("total_price", r["total_price"]),
                ("tax_amount", r["tax_amount"]), ("year", r["year"]),
            ):
                if val is not None:
                    setattr(q, field, val)
            if r["quoted_date"] and not q.quoted_at:
                q.quoted_at = datetime.combine(r["quoted_date"], datetime.min.time())
            updated += 1

        # step 6（2026-08-28，A32 之後才成立）：case_code 是跨模組橋樑，
        # 而存量 pm_cases 已全轉為 CK 建案案號 —— 再把報價單編號寫進
        # case_code 就是把 legacy 制重新引進來（且與既有案子永遠比對不上
        # ⇒ 每次匯入都重複建案，08-20 那 36 組分身的同型）。
        #
        # 既有報價單以 legacy_quotation_no 對到（to_update，case_code 不動）；
        # to_create ＝ 真正的新業務 ⇒ 走正式產號。
        # ⚠️ 08-20 那次產號 180 筆全拿到同一個號（同交易內查不到未 commit 的），
        #    產號器已修：流水號計數器掛在 session（db.info）上，同批不撞。
        code_svc = CaseCodeService(self.db)
        for r in to_create:
            yr = r["year"] or date.today().year
            if yr and yr < 1911:
                yr += 1911
            code_of[r["legacy_no"]] = await code_svc.generate_case_code("pm", yr, "02")

        for r in to_create:
            q = ERPQuotation(
                case_code=code_of[r["legacy_no"]],
                case_name=r["case_name"],
                year=r["year"] or date.today().year,
                total_price=r["total_price"],
                tax_amount=r["tax_amount"],
                # 「是否成立=v」＝這張報價客戶接受了 ⇒ confirmed；其餘留 draft。
                # 不寫成 contracted —— 那是承攬案件的狀態，不是報價單的。
                status="confirmed" if r["established"] else "draft",
                legacy_quotation_no=r["legacy_no"],
                quoted_at=(datetime.combine(r["quoted_date"], datetime.min.time())
                           if r["quoted_date"] else None),
                created_by=user_id,
                notes=_build_notes(r),
            )
            self.db.add(q)
            created += 1

        pm_created = await self._ensure_pm_cases(to_create + to_update, code_of=code_of)
        staff_res = await self._assign_staff_from_sheets(rows)

        await self.db.commit()
        # 2026-09-03 owner「表單匯入修正機制」：收款／發票此前只寫進 notes（那時沒有結構化位置）。
        # 現在接到請款／發票，並讓成立且有金額的報價單有第一期（成案即應收）。
        finance = await self._sync_finance(to_create + to_update)
        # 2026-09-03 全景覆盤 A5：匯入是最大的寫入來源，此前只有 log。寫一筆審計：誰、何時、哪個檔、改幾筆。
        try:
            import json as _json
            await self.db.execute(text(
                "INSERT INTO audit_logs (table_name, record_id, action, changes, user_id, source, is_critical, created_at) "
                "VALUES ('erp_quotations', 0, 'import', :c, :u, 'quotation_legacy_import', true, now())"
            ), {"c": _json.dumps({"file": source_name, "total_rows": len(rows), "created": created, "updated": updated,
                                  "skipped": len(skipped), "pm_created": pm_created, "finance": finance,
                                  "promoted": getattr(self, "promoted_count", 0),
                                  "promote_failures": getattr(self, "promote_failures", [])}, ensure_ascii=False), "u": user_id})
            await self.db.commit()
        except Exception as e:
            logger.warning("匯入審計寫入失敗（不影響匯入）: %s", e)
        logger.info("報價單彙整匯入：新增 %d／更新 %d／略過 %d／補建 PM 案件 %d",
                    created, updated, len(skipped), pm_created)
        # 實際匯入時重算一次 —— preview 那次是 dry_run，兩次的 missing 可能不同
        preview["duplicate_candidates"] = getattr(self, "dup_candidates", [])[:200]
        preview["duplicate_candidate_count"] = len(getattr(self, "dup_candidates", []))
        return {**preview, "dry_run": False, "created": created, "updated": updated,
                "created_pm_cases": pm_created, "finance": finance,
                "promoted": getattr(self, "promoted_count", 0),
                "promote_failures": getattr(self, "promote_failures", []),
                "assigned_staff": staff_res["assigned"],
                "staff_unmatched_sheets": staff_res["unmatched_sheets"]}
