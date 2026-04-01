"""ERP 報價/成本服務 — 含損益即時計算

Version: 1.3.0
- v1.3.0: create/update/delete 改用 Repository 方法 (合規修正)
"""
import csv
import io
import logging
from typing import Optional, Tuple, List
from decimal import Decimal, ROUND_HALF_UP

from sqlalchemy.ext.asyncio import AsyncSession

from app.extended.models.erp import ERPQuotation
from app.repositories.erp import (
    ERPQuotationRepository, ERPInvoiceRepository,
    ERPBillingRepository, ERPVendorPayableRepository,
)
from app.schemas.erp import (
    ERPQuotationCreate, ERPQuotationUpdate, ERPQuotationResponse,
    ERPQuotationListRequest, ERPProfitSummary, ERPProfitTrendItem,
)
from app.services.case_code_service import CaseCodeService
from app.services.audit_mixin import AuditableServiceMixin

logger = logging.getLogger(__name__)

ZERO = Decimal("0")


class ERPQuotationService(AuditableServiceMixin):
    """報價管理服務 — 損益計算核心"""

    AUDIT_TABLE = "erp_quotations"

    def __init__(self, db: AsyncSession):
        self.db = db
        self.repo = ERPQuotationRepository(db)
        self.invoice_repo = ERPInvoiceRepository(db)
        self.billing_repo = ERPBillingRepository(db)
        self.payable_repo = ERPVendorPayableRepository(db)
        self.code_service = CaseCodeService(db)

    # =========================================================================
    # CRUD
    # =========================================================================

    async def generate_case_code(self, year: int, category: str = "01") -> str:
        """產生 ERP 案號"""
        return await self.code_service.generate_case_code("erp", year, category)

    async def create(self, data: ERPQuotationCreate, user_id: Optional[int] = None) -> ERPQuotationResponse:
        """建立報價 — case_code 未提供時自動產生，已有案號時驗證 PM 參照"""
        dump = data.model_dump()

        # 自動產生案號
        if not dump.get("case_code"):
            from datetime import date as _date
            year = dump.get("year") or _date.today().year
            category = "01"  # ERP 預設報價單
            dump["case_code"] = await self.code_service.generate_case_code(
                "erp", year, category,
            )
        else:
            # case_code 參照完整性驗證 — 確認 PM 案件存在
            await self._validate_case_code(dump["case_code"])

        dump["created_by"] = user_id
        quotation = await self.repo.create(dump)
        await self.audit_create(quotation.id, dump, user_id=user_id)
        return await self._to_response(quotation)

    async def get_detail(self, quotation_id: int) -> Optional[ERPQuotationResponse]:
        """取得報價詳情 (含計算欄位)"""
        quotation = await self.repo.get_by_id(quotation_id)
        if not quotation:
            return None
        return await self._to_response(quotation)

    async def update(self, quotation_id: int, data: ERPQuotationUpdate) -> Optional[ERPQuotationResponse]:
        """更新報價"""
        changes = data.model_dump(exclude_unset=True)
        quotation = await self.repo.update(quotation_id, changes)
        if not quotation:
            return None
        await self.audit_update(quotation_id, changes)
        return await self._to_response(quotation)

    async def delete(self, quotation_id: int) -> bool:
        """刪除報價"""
        result = await self.repo.delete(quotation_id)
        if result:
            await self.audit_delete(quotation_id)
        return result

    async def list_quotations(self, params: ERPQuotationListRequest) -> Tuple[List[ERPQuotationResponse], int]:
        """報價列表 — 使用批次聚合消除 N+1 查詢"""
        items, total = await self.repo.filter_quotations(
            year=params.year,
            status=params.status,
            case_code=params.case_code,
            search=params.search,
            skip=params.skip,
            limit=params.limit,
            sort_by=params.sort_by or "id",
            sort_order=params.sort_order.value if params.sort_order else "desc",
        )

        if not items:
            return [], total

        # 批次取得聚合數據 (2 queries instead of N*6)
        ids = [q.id for q in items]
        billing_agg = await self.billing_repo.get_aggregates_batch(ids)
        payable_agg = await self.payable_repo.get_aggregates_batch(ids)
        # invoice count 透過 billing count 估算或單獨批次查詢
        invoice_counts = await self._get_invoice_counts_batch(ids)

        responses = []
        for item in items:
            b = billing_agg.get(item.id, {})
            p = payable_agg.get(item.id, {})
            responses.append(self._to_response_with_aggregates(
                item,
                billing_count=b.get("count", 0),
                total_billed=b.get("total_billed", ZERO),
                total_received=b.get("total_received", ZERO),
                total_payable=p.get("total_payable", ZERO),
                total_paid=p.get("total_paid", ZERO),
                invoice_count=invoice_counts.get(item.id, 0),
            ))
        return responses, total

    async def _get_invoice_counts_batch(self, quotation_ids: List[int]) -> dict:
        """批次取得發票數量 — 委派至 ERPInvoiceRepository"""
        return await self.invoice_repo.get_counts_by_quotation_ids(quotation_ids)

    def _to_response_with_aggregates(
        self,
        quotation: ERPQuotation,
        billing_count: int,
        total_billed: Decimal,
        total_received: Decimal,
        total_payable: Decimal,
        total_paid: Decimal,
        invoice_count: int,
    ) -> ERPQuotationResponse:
        """轉換為回應格式 (使用預先批次聚合的數據，避免 N+1)"""
        profit = self.compute_profit(quotation)

        budget_limit = quotation.budget_limit
        budget_usage_pct = None
        is_over_budget = False
        if budget_limit and budget_limit > ZERO:
            usage = profit["total_cost"] / budget_limit * 100
            budget_usage_pct = usage.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            is_over_budget = profit["total_cost"] > budget_limit

        return ERPQuotationResponse(
            **{c.name: getattr(quotation, c.name) for c in quotation.__table__.columns},
            budget_usage_pct=budget_usage_pct,
            is_over_budget=is_over_budget,
            total_cost=profit["total_cost"],
            gross_profit=profit["gross_profit"],
            gross_margin=profit["gross_margin"],
            net_profit=profit["net_profit"],
            invoice_count=invoice_count,
            billing_count=billing_count,
            total_billed=total_billed,
            total_received=total_received,
            total_payable=total_payable,
            total_paid=total_paid,
        )

    async def _validate_case_code(self, case_code: str) -> None:
        """驗證 case_code 是否存在於 PM 系統 (參照完整性)"""
        try:
            from app.repositories.pm import PMCaseRepository
            pm_repo = PMCaseRepository(self.db)
            pm_case = await pm_repo.get_by_case_code(case_code)
            if not pm_case:
                logger.warning("ERP case_code '%s' 不存在於 PM 系統", case_code)
        except Exception:
            # PM 模組不可用時不阻擋 ERP 建案 (降級容錯)
            pass

    # =========================================================================
    # 損益計算
    # =========================================================================

    @staticmethod
    def compute_profit(quotation: ERPQuotation) -> dict:
        """計算毛利/淨利

        total_cost = outsourcing + personnel + overhead + other
        gross_profit = total_price - tax - total_cost
        gross_margin = gross_profit / (total_price - tax) * 100
        net_profit = gross_profit (可擴充扣除額外費用)
        """
        total_price = Decimal(str(quotation.total_price or 0))
        tax = Decimal(str(quotation.tax_amount or 0))
        outsourcing = Decimal(str(quotation.outsourcing_fee or 0))
        personnel = Decimal(str(quotation.personnel_fee or 0))
        overhead = Decimal(str(quotation.overhead_fee or 0))
        other = Decimal(str(quotation.other_cost or 0))

        total_cost = outsourcing + personnel + overhead + other
        revenue = total_price - tax
        gross_profit = revenue - total_cost

        gross_margin = None
        if revenue > ZERO:
            gross_margin = (gross_profit / revenue * 100).quantize(
                Decimal("0.01"), rounding=ROUND_HALF_UP
            )

        return {
            "total_cost": total_cost,
            "gross_profit": gross_profit,
            "gross_margin": gross_margin,
            "net_profit": gross_profit,  # 可擴充
        }

    # =========================================================================
    # 損益摘要
    # =========================================================================

    async def get_profit_summary(self, year: Optional[int] = None) -> ERPProfitSummary:
        """年度損益摘要 — 批次聚合消除 N+1"""
        items, _ = await self.repo.filter_quotations(
            year=year, skip=0, limit=9999,
        )

        total_revenue = ZERO
        total_cost = ZERO
        total_gross = ZERO

        # 批次取得請款聚合
        ids = [q.id for q in items]
        billing_agg = await self.billing_repo.get_aggregates_batch(ids) if ids else {}

        for q in items:
            profit = self.compute_profit(q)
            price = Decimal(str(q.total_price or 0))
            tax = Decimal(str(q.tax_amount or 0))
            total_revenue += price - tax
            total_cost += profit["total_cost"]
            total_gross += profit["gross_profit"]

        total_billed = sum(
            (v.get("total_billed", ZERO) for v in billing_agg.values()), ZERO,
        )
        total_received = sum(
            (v.get("total_received", ZERO) for v in billing_agg.values()), ZERO,
        )

        avg_margin = None
        if total_revenue > ZERO:
            avg_margin = (total_gross / total_revenue * 100).quantize(
                Decimal("0.01"), rounding=ROUND_HALF_UP
            )

        return ERPProfitSummary(
            total_revenue=total_revenue,
            total_cost=total_cost,
            total_gross_profit=total_gross,
            avg_gross_margin=avg_margin,
            total_billed=total_billed,
            total_received=total_received,
            total_outstanding=total_billed - total_received,
            case_count=len(items),
        )

    # =========================================================================
    # 多年度損益趨勢
    # =========================================================================

    async def get_profit_trend(self) -> List[ERPProfitTrendItem]:
        """多年度損益趨勢 — SQL 聚合 (取代全表載入)"""
        rows = await self.repo.get_yearly_trend_sql()
        return [ERPProfitTrendItem(**row) for row in rows]

    # =========================================================================
    # CSV 匯出
    # =========================================================================

    async def export_csv(self, year: Optional[int] = None) -> str:
        """匯出報價為 CSV 字串 (含損益計算)"""
        items, _ = await self.repo.filter_quotations(
            year=year, skip=0, limit=9999,
        )

        output = io.StringIO()
        output.write("\ufeff")  # BOM for Excel
        writer = csv.writer(output)
        writer.writerow([
            "案號", "案名", "年度", "總價", "稅額",
            "外包費", "人事費", "管銷費", "其他成本",
            "毛利", "毛利率(%)", "狀態",
        ])

        for item in items:
            profit = self.compute_profit(item)
            writer.writerow([
                item.case_code or "",
                item.case_name or "",
                item.year or "",
                item.total_price or "",
                item.tax_amount or "",
                item.outsourcing_fee or "",
                item.personnel_fee or "",
                item.overhead_fee or "",
                item.other_cost or "",
                profit["gross_profit"],
                profit["gross_margin"] or "",
                item.status or "",
            ])

        return output.getvalue()

    # =========================================================================
    # Excel 匯出 / 匯入
    # =========================================================================

    async def export_excel(self, year: Optional[int] = None) -> bytes:
        """匯出報價為 Excel (.xlsx)"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        items, _ = await self.repo.filter_quotations(year=year, skip=0, limit=9999)

        wb = Workbook()
        ws = wb.active
        ws.title = "報價管理"

        headers = [
            "案號", "成案編號", "案名", "年度", "總價", "稅額",
            "外包費", "人事費", "管銷費", "其他成本",
            "毛利", "毛利率(%)", "狀態", "備註",
        ]
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row_idx, item in enumerate(items, 2):
            profit = self.compute_profit(item)
            ws.cell(row=row_idx, column=1, value=item.case_code or "")
            ws.cell(row=row_idx, column=2, value=item.project_code or "")
            ws.cell(row=row_idx, column=3, value=item.case_name or "")
            ws.cell(row=row_idx, column=4, value=item.year or "")
            ws.cell(row=row_idx, column=5, value=float(item.total_price or 0))
            ws.cell(row=row_idx, column=6, value=float(item.tax_amount or 0))
            ws.cell(row=row_idx, column=7, value=float(item.outsourcing_fee or 0))
            ws.cell(row=row_idx, column=8, value=float(item.personnel_fee or 0))
            ws.cell(row=row_idx, column=9, value=float(item.overhead_fee or 0))
            ws.cell(row=row_idx, column=10, value=float(item.other_cost or 0))
            ws.cell(row=row_idx, column=11, value=float(profit["gross_profit"]))
            ws.cell(row=row_idx, column=12, value=float(profit["gross_margin"] or 0))
            ws.cell(row=row_idx, column=13, value=item.status or "")
            ws.cell(row=row_idx, column=14, value=item.notes or "")

        # 欄寬自動調整
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 15
        ws.column_dimensions["C"].width = 40  # 案名較長
        ws.freeze_panes = "A2"

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    def generate_import_template(self) -> bytes:
        """產生匯入範本 Excel"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = "報價匯入範本"

        headers = ["案號", "案名", "年度(西元)", "總價", "稅額", "外包費", "人事費", "管銷費", "其他成本", "狀態", "備註"]
        fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        font = Font(bold=True, color="FFFFFF")

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center")

        # 範例資料
        sample = ["B114-B999", "範例測量案件", 2025, 100000, 5000, 30000, 40000, 20000, 10000, "draft", "匯入測試"]
        for col, v in enumerate(sample, 1):
            ws.cell(row=2, column=col, value=v)

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 15
        ws.column_dimensions["B"].width = 30

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    async def import_from_excel(self, file_bytes: bytes, user_id: Optional[int] = None) -> dict:
        """匯入報價 Excel — 用 case_code 做 upsert"""
        import unicodedata
        import re
        from app.services.base.excel_reader import load_workbook_any

        wb = load_workbook_any(file_bytes)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))

        created = 0
        updated = 0
        errors: list = []

        def _num(v) -> Decimal:
            if v is None:
                return Decimal("0")
            if isinstance(v, (int, float)):
                return Decimal(str(v))
            s = re.sub(r'[NT$￥,\s]', '', str(v).strip())
            return Decimal(s) if s else Decimal("0")

        def _str(v) -> Optional[str]:
            if v is None:
                return None
            return unicodedata.normalize('NFKC', str(v).strip()) or None

        for idx, row in enumerate(rows, start=2):
            try:
                if not row or len(row) < 3:
                    continue
                case_code = _str(row[0])
                case_name = _str(row[1])
                if not case_code or not case_name:
                    continue

                year_val = row[2]
                if isinstance(year_val, (int, float)):
                    year = int(year_val)
                else:
                    year = int(str(year_val).strip()) if year_val else None
                # 民國年自動轉西元
                if year and year < 1911:
                    year = year + 1911

                data = {
                    "case_code": case_code,
                    "case_name": case_name,
                    "year": year,
                    "total_price": _num(row[3]) if len(row) > 3 else Decimal("0"),
                    "tax_amount": _num(row[4]) if len(row) > 4 else Decimal("0"),
                    "outsourcing_fee": _num(row[5]) if len(row) > 5 else Decimal("0"),
                    "personnel_fee": _num(row[6]) if len(row) > 6 else Decimal("0"),
                    "overhead_fee": _num(row[7]) if len(row) > 7 else Decimal("0"),
                    "other_cost": _num(row[8]) if len(row) > 8 else Decimal("0"),
                    "status": _str(row[9]) or "draft" if len(row) > 9 else "draft",
                    "notes": _str(row[10]) if len(row) > 10 else None,
                }

                # Upsert by case_code
                existing = await self.repo.get_by_case_code(case_code)
                if existing:
                    update_data = {k: v for k, v in data.items() if k != "case_code" and v is not None}
                    await self.repo.update(existing.id, update_data)
                    updated += 1
                else:
                    data["created_by"] = user_id
                    await self.repo.create(data)
                    created += 1

            except Exception as e:
                errors.append({"row": idx, "error": str(e)})

        if created > 0 or updated > 0:
            await self.db.commit()

        return {
            "total_rows": len(rows),
            "created": created,
            "updated": updated,
            "errors": errors,
        }

    # =========================================================================
    # 轉換
    # =========================================================================

    async def _to_response(self, quotation: ERPQuotation) -> ERPQuotationResponse:
        """轉換為回應格式 (含計算欄位 + 聚合)"""
        profit = self.compute_profit(quotation)

        invoices = await self.invoice_repo.get_by_quotation_id(quotation.id)
        total_billed = await self.billing_repo.get_total_billed(quotation.id)
        total_received = await self.billing_repo.get_total_received(quotation.id)
        total_payable = await self.payable_repo.get_total_payable(quotation.id)
        total_paid = await self.payable_repo.get_total_paid(quotation.id)
        billings = await self.billing_repo.get_by_quotation_id(quotation.id)

        # 預算警示計算
        budget_limit = quotation.budget_limit
        budget_usage_pct = None
        is_over_budget = False
        if budget_limit and budget_limit > ZERO:
            usage = profit["total_cost"] / budget_limit * 100
            budget_usage_pct = usage.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            is_over_budget = profit["total_cost"] > budget_limit

        return ERPQuotationResponse(
            **{c.name: getattr(quotation, c.name) for c in quotation.__table__.columns},
            budget_usage_pct=budget_usage_pct,
            is_over_budget=is_over_budget,
            total_cost=profit["total_cost"],
            gross_profit=profit["gross_profit"],
            gross_margin=profit["gross_margin"],
            net_profit=profit["net_profit"],
            invoice_count=len(invoices),
            billing_count=len(billings),
            total_billed=total_billed,
            total_received=total_received,
            total_payable=total_payable,
            total_paid=total_paid,
        )
