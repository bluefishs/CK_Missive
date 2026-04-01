# -*- coding: utf-8 -*-
"""
統一 Excel 讀取工具

自動偵測 .xls (Excel 97-2003) vs .xlsx (Excel 2007+) 格式，
統一回傳 openpyxl Workbook 物件。

.xls 處理流程：
  xlrd 讀取 → 轉換為 openpyxl Workbook（含 NFKC 正規化）

用法：
    from app.services.base.excel_reader import load_workbook_any

    wb = load_workbook_any(file_bytes, filename="data.xls")
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        ...
"""
import io
import logging
import unicodedata
from typing import Optional

import openpyxl

logger = logging.getLogger(__name__)


def load_workbook_any(
    file_bytes: bytes,
    filename: Optional[str] = None,
    data_only: bool = True,
    read_only: bool = False,
) -> openpyxl.Workbook:
    """
    讀取 Excel 檔案，自動支援 .xls 和 .xlsx。

    Args:
        file_bytes: 檔案二進位內容
        filename: 原始檔名（用於格式偵測，可選）
        data_only: 是否只讀取值（忽略公式）
        read_only: 是否唯讀模式

    Returns:
        openpyxl.Workbook

    Raises:
        ValueError: 無法辨識的檔案格式
    """
    is_xls = _is_xls_format(file_bytes, filename)

    if is_xls:
        logger.info(f"偵測到 .xls 格式，使用 xlrd 轉換: {filename or '(unnamed)'}")
        return _convert_xls_to_openpyxl(file_bytes)
    else:
        return openpyxl.load_workbook(
            io.BytesIO(file_bytes),
            data_only=data_only,
            read_only=read_only,
        )


def _is_xls_format(file_bytes: bytes, filename: Optional[str] = None) -> bool:
    """偵測是否為舊版 .xls 格式（magic bytes 或副檔名）"""
    # OLE2 Compound Document magic bytes (Excel 97-2003)
    if file_bytes[:8] == b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1':
        return True
    # 副檔名判斷（作為備用）
    if filename and filename.lower().endswith('.xls') and not filename.lower().endswith('.xlsx'):
        return True
    return False


def _convert_xls_to_openpyxl(file_bytes: bytes) -> openpyxl.Workbook:
    """
    將 .xls 檔案轉換為 openpyxl Workbook。

    xlrd 2.0+ 只支援 .xls (不含 .xlsx)。
    轉換過程中對所有字串值執行 NFKC 正規化。
    """
    import xlrd
    from datetime import datetime, date

    xls_book = xlrd.open_workbook(file_contents=file_bytes)
    wb = openpyxl.Workbook()

    for sheet_idx, sheet_name in enumerate(xls_book.sheet_names()):
        xls_sheet = xls_book.sheet_by_index(sheet_idx)

        if sheet_idx == 0:
            ws = wb.active
            ws.title = sheet_name
        else:
            ws = wb.create_sheet(title=sheet_name)

        for row_idx in range(xls_sheet.nrows):
            for col_idx in range(xls_sheet.ncols):
                cell_value = xls_sheet.cell_value(row_idx, col_idx)
                cell_type = xls_sheet.cell_type(row_idx, col_idx)

                # 型別轉換
                if cell_type == xlrd.XL_CELL_DATE:
                    # xlrd 日期 → Python datetime
                    try:
                        date_tuple = xlrd.xldate_as_tuple(cell_value, xls_book.datemode)
                        if date_tuple[3:] == (0, 0, 0):
                            cell_value = date(*date_tuple[:3])
                        else:
                            cell_value = datetime(*date_tuple)
                    except Exception:
                        pass
                elif cell_type == xlrd.XL_CELL_TEXT:
                    # 字串 NFKC 正規化
                    cell_value = unicodedata.normalize('NFKC', cell_value.strip())
                elif cell_type == xlrd.XL_CELL_NUMBER:
                    # 整數化（若無小數部分）
                    if cell_value == int(cell_value):
                        cell_value = int(cell_value)
                elif cell_type == xlrd.XL_CELL_BOOLEAN:
                    cell_value = bool(cell_value)
                elif cell_type == xlrd.XL_CELL_EMPTY:
                    cell_value = None

                ws.cell(row=row_idx + 1, column=col_idx + 1, value=cell_value)

    xls_book.release_resources()
    logger.info(f"XLS 轉換完成: {len(xls_book.sheet_names())} sheets, "
                f"{sum(xls_book.sheet_by_index(i).nrows for i in range(len(xls_book.sheet_names())))} rows")
    return wb
